# pages/promotion_viewer.py
import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# ========== دوال التنسيق ==========
def apply_excel_style(writer, sheet_name, df):
    """تطبيق التنسيقات الاحترافية على ملف Excel"""
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    header_fill = PatternFill(start_color="1F7A8C", end_color="1F7A8C", fill_type="solid")
    header_font = Font(name="Tajawal", size=12, bold=True, color="FFFFFF")
    alt_row_fill = PatternFill(start_color="E6F3F5", end_color="E6F3F5", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max(20, len(str(col_name)) + 5)
    
    for row_idx in range(2, len(df) + 2):
        is_alt = (row_idx - 2) % 2 == 1
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_alt:
                cell.fill = alt_row_fill
    
    worksheet.freeze_panes = 'A2'

@st.cache_data
def generate_empty_template():
    """توليد النموذج الفارغ"""
    output = BytesIO()
    wb = openpyxl.Workbook()
    
    header_fill = PatternFill(start_color="1F7A8C", end_color="1F7A8C", fill_type="solid")
    font_headers = Font(name="Tajawal", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws1 = wb.active
    ws1.title = "عرض خاص"
    ws1.views.sheetView[0].rightToLeft = True
    cell_1 = ws1.cell(row=1, column=1, value="  العروض الخاصة  ")
    cell_1.fill = header_fill
    cell_1.font = font_headers
    cell_1.alignment = align_center
    cell_1.border = border_thin
    ws1.column_dimensions['A'].width = 60
    ws1.row_dimensions[1].height = 24
    
    ws2 = wb.create_sheet(title="سعر مخفض")
    ws2.views.sheetView[0].rightToLeft = True
    headers_discounted = ["رمز المنتج sku", "أسم المنتج", "السعر المخفض", "تاريخ نهاية التخفيض", "العنوان الترويجي"]
    for col_idx, text in enumerate(headers_discounted, 1):
        cell = ws2.cell(row=1, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = font_headers
        cell.alignment = align_center
        cell.border = border_thin
        ws2.column_dimensions[get_column_letter(col_idx)].width = 24
    ws2.row_dimensions[1].height = 24
        
    ws3 = wb.create_sheet(title="اسعار المنتجات")
    ws3.views.sheetView[0].rightToLeft = True
    headers_prices = ["رمز المنتج sku", "أسم المنتج", "سعر المنتج", "خاضع للضريبة"]
    for col_idx, text in enumerate(headers_prices, 1):
        cell = ws3.cell(row=2, column=col_idx, value=text)
        cell.fill = header_fill
        cell.font = font_headers
        cell.alignment = align_center
        cell.border = border_thin
        ws3.column_dimensions[get_column_letter(col_idx)].width = 24
    ws3.row_dimensions[1].height = 18
    ws3.row_dimensions[2].height = 24
    
    # 🌟 إضافة الشيت الرابع: عروض لها وضع خاص
    ws4 = wb.create_sheet(title="عروض لها وضع خاص")
    ws4.views.sheetView[0].rightToLeft = True
    cell_4 = ws4.cell(row=1, column=1, value="  عروض لها وضع خاص  ")
    cell_4.fill = header_fill
    cell_4.font = font_headers
    cell_4.alignment = align_center
    cell_4.border = border_thin
    ws4.column_dimensions['A'].width = 60
    ws4.row_dimensions[1].height = 24
        
    wb.save(output)
    return output.getvalue()

# ========== دوال تحليل العروض ==========
def extract_offer_name(text):
    """استخراج اسم العرض التفصيلي مع دعم جميع الصيغ"""
    text = str(text) if not pd.isna(text) else ""
    
    # 1️⃣ خصم على القطعة الثانية
    if "خصم" in text and "القطعة الثانية" in text:
        match = re.search(r'خصم\s*(\d+)\s*%\s*على\s*القطعة\s*الثانية', text)
        if match: return f"خصم {match.group(1)}% على القطعة الثانية"
        match = re.search(r'خصم\s*(\d+)\s*%\s*على\s*القطعة', text)
        if match: return f"خصم {match.group(1)}% على القطعة الثانية"
        return "خصم على القطعة الثانية"
    
    # 2️⃣ خصم على الحبة الثانية
    if "خصم" in text and "الحبة الثانية" in text:
        match = re.search(r'خصم\s*(\d+)\s*%\s*على\s*الحبة\s*الثانية', text)
        if match: return f"خصم {match.group(1)}% على الحبة الثانية"
        match = re.search(r'خصم\s*(\d+)\s*%\s*على\s*الحبة', text)
        if match: return f"خصم {match.group(1)}% على الحبة الثانية"
        match = re.search(r'خصم\s*(\d+)\s*ريال\s*على\s*الحبة\s*الثانية', text)
        if match: return f"خصم {match.group(1)} ريال على الحبة الثانية"
        return "خصم على الحبة الثانية"
    
    # 3️⃣ عروض مجانية (2+1 مجاناً، 2 +1 مجانا، إلخ)
    if "مجاناً" in text or "مجانا" in text:
        match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجاناً?', text)
        if match: return f"عرض {match.group(1)}+{match.group(2)} مجاناً"
        match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجانا', text)
        if match: return f"عرض {match.group(1)}+{match.group(2)} مجاناً"
        match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجاناً?', text)
        if match: return f"عرض {match.group(1)}+{match.group(2)} مجاناً"
        return "عرض مجاني"
    
    # 4️⃣ صفقة اليوم
    if "صفقة اليوم" in text:
        match = re.search(r'صفقة اليوم\s*:\s*([^:]+?)(?=\s*(?:اذا اشترى|نسبة من|يبدأ بتاريخ|$))', text)
        if match:
            full_text = match.group(1).strip()
            if len(full_text) > 50:
                full_text = full_text[:50] + "..."
            return f"صفقة اليوم : {full_text}"
        return "صفقة اليوم"
    
    # 5️⃣ عروض الكميات (6حبات بسعر 77 ريال)
    if "حبة بسعر" in text or "حبات بسعر" in text:
        match = re.search(r'(\d+)\s*حبات?\s*بسعر\s*([\d.]+)\s*ريال', text)
        if match: return f"{match.group(1)}حبات بسعر {match.group(2)} ريال"
        match = re.search(r'(\d+)\s*حبات?\s*ب\s*([\d.]+)\s*ريال', text)
        if match: return f"{match.group(1)}حبات بسعر {match.group(2)} ريال"
        return "عرض كميات"
    
    # 6️⃣ عرض خاص - خصم
    if "عرض خاص" in text and "خصم" in text:
        match = re.search(r'عرض خاص\s*-\s*خصم\s*(\d+)\s*%\s*على\s*الحبة\s*الثانية', text)
        if match: return f"عرض خاص - خصم {match.group(1)}% على الحبة الثانية"
        return "عرض خاص"
    
    # 7️⃣ صيغة "2حبة بسعر 65.50 ريال"
    match = re.search(r'(\d+)\s*حبة\s*بسعر\s*([\d.]+)\s*ريال', text)
    if match: return f"{match.group(1)}حبة بسعر {match.group(2)} ريال"
    
    match = re.search(r'(\d+)\s*حبة\s*ب\s*([\d.]+)\s*ريال', text)
    if match: return f"{match.group(1)}حبة بسعر {match.group(2)} ريال"
    
    # 8️⃣ حالة خاصة: أرقام فقط بدون عرض واضح
    if "/" in text:
        parts = text.split("/")
        if len(parts) >= 2:
            potential_name = parts[0].strip()
            if re.search(r'[^\d\s\-]', potential_name):
                return potential_name
    
    return "عرض خاص"

def extract_offer_quantity(text):
    """استخراج عدد حبات العرض من نص العرض"""
    text = str(text) if not pd.isna(text) else ""
    
    # 1️⃣ عروض مجانية (2+1 مجاناً)
    match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجاناً?', text)
    if match:
        paid = int(match.group(1))
        free = int(match.group(2))
        return f"{paid + free}"
    
    # 2️⃣ صيغة "2 +1 مجانا"
    match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجانا', text)
    if match:
        paid = int(match.group(1))
        free = int(match.group(2))
        return f"{paid + free}"
    
    # 3️⃣ عروض الكميات (6حبات بسعر 77 ريال)
    match = re.search(r'(\d+)\s*حبات?\s*بسعر', text)
    if match:
        return match.group(1)
    
    match = re.search(r'(\d+)\s*حبات?\s*ب', text)
    if match:
        return match.group(1)
    
    # 4️⃣ خصم على الحبة الثانية
    if "الحبة الثانية" in text:
        return "2"
    
    # 5️⃣ خصم على القطعة الثانية
    if "القطعة الثانية" in text:
        return "2"
    
    return ""

def extract_special_offer_sku(text):
    """استخراج رقم المنتج الخاص بالعرض (المجموعة) من النص"""
    text = str(text) if not pd.isna(text) else ""
    
    # البحث عن صيغة "رقم-رقم-رقم" (مثل 5209-17164-4846)
    match = re.search(r'(\d{3,6}-\d{3,6}-\d{3,6})', text)
    if match:
        return match.group(1)
    
    # البحث عن صيغة "رقم*رقم" (مثل 16265*6)
    match = re.search(r'(\d{3,6}\*\d+)', text)
    if match:
        return match.group(1)
    
    return ""

def extract_dates(text):
    """استخراج النطاق الزمني للعروض"""
    text = str(text) if not pd.isna(text) else ""
    date_match = re.findall(r'(\d{4}-\d{2}-\d{2} \d{2}:\d{2})', text)
    start = date_match[0] if len(date_match) > 0 else None
    end = date_match[1] if len(date_match) > 1 else None
    return start, end

def extract_numbers_from_text(text):
    """عزل الأرقام الفردية للمنتجات بعد استبعاد السلاسل المركبة"""
    if pd.isna(text):
        return []
    text = str(text)
    # 🌟 استبعاد السلاسل التي تحتوي على + أو - معاً
    text_clean = re.sub(r'\d{3,6}(?:[-+]\d{3,6})+', '', text)
    text_clean = re.sub(r'\d{3,6}\s*\*\s*\d+', '', text_clean)
    
    excluded_years = {'2024', '2025', '2026', '2027', '2028', '2029', '2030'}
    pattern = r'(?:^|[^0-9])(\d{3,6})(?:[^0-9]|$)'
    matches = re.findall(pattern, text_clean)
    return [m for m in matches if m not in excluded_years and not m.startswith('20')]

def clean_sku(val):
    """تنظيف أرقام الـ SKU من الفواصل العشرية (لحل مشكلة اختفاء النسب)"""
    if pd.isna(val): return ""
    s = str(val).strip()
    if s.endswith(".0"): s = s[:-2]
    return s
    
def parse_composite_sku(sku):
    """تحليل الأرقام المركبة والسلاسل الطويلة (مع دعم + و -)"""
    sku = str(sku).strip().replace(" ", "")
    if not sku or sku == "nan" or sku == "":
        return None, None, None, [], False, False
        
    # 🌟 التعديل هنا: دعم الفواصل سواء كانت (-) أو (+)
    if '-' in sku or '+' in sku:
        parts = re.split(r'[-+]', sku)
        individual_skus = []
        for p in parts:
            if '*' in p:
                m = re.match(r'^(\d+)\*(\d+)$', p)
                if m: individual_skus.append(m.group(1))
            else:
                m = re.match(r'^(\d+)$', p)
                if m: individual_skus.append(p)
        if individual_skus:
            return individual_skus[0], sku, None, individual_skus, False, True
            
    if '*' in sku:
        match = re.match(r'^(\d+)\*(\d+)$', sku)
        if match:
            base_sku = match.group(1)
            qty = int(match.group(2))
            return base_sku, sku, qty, [base_sku], True, False

    return sku, None, None, [sku], False, False

# ========== دوال حساب الخصم والنسب ==========
def is_taxable(sku, price_map):
    """التحقق من أن المنتج خاضع للضريبة من خلال العمود D"""
    if sku in price_map:
        return price_map[sku].get("taxable", False)
    return False

def remove_tax(price_including_tax):
    """إزالة الضريبة 15% من السعر الشامل"""
    try:
        price = float(str(price_including_tax).replace(',', '').strip())
        if price <= 0:
            return price
        return price / 1.15
    except (ValueError, TypeError):
        return price_including_tax

def extract_quantity_from_promo(promo_text):
    """استخراج عدد الحبات من النص الترويجي (مثل 6حبات ب 75ريال)"""
    if not promo_text or not isinstance(promo_text, str):
        return 1
    
    patterns = [
        r'(\d+)\s*حبات?\s*ب',
        r'(\d+)\s*حبة\s*ب',
        r'(\d+)\s*حبات?\s*بسعر',
        r'(\d+)\s*حبة\s*بسعر'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, promo_text)
        if match:
            return int(match.group(1))
    
    return 1

def extract_price_from_promo(promo_text):
    """استخراج السعر من النص الترويجي"""
    if not promo_text or not isinstance(promo_text, str):
        return None
    
    patterns = [
        r'ب\s*([\d.]+)\s*ريال',
        r'بسعر\s*([\d.]+)\s*ريال',
        r'ب\s*([\d.]+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, promo_text)
        if match:
            return float(match.group(1))
    
    return None

def extract_discount_value_from_promo(promo_text):
    """استخراج قيمة الخصم بالريال من النص الترويجي (مثل خصم 17 ريال)"""
    if not promo_text or not isinstance(promo_text, str):
        return None
    
    match = re.search(r'خصم\s*([\d.]+)\s*ريال', promo_text)
    if match:
        return float(match.group(1))
    
    return None

def calculate_discount_percentage(original_price, discounted_price, promo_text="", quantity=1, taxable=False, price_map=None, sku=None):
    """
    حساب نسبة الخصم بشكل صحيح مع مراعاة الضريبة وعدد الحبات
    """
    try:
        # 1️⃣ إذا كان النص الترويجي يحتوي على نسبة مئوية مباشرة
        if promo_text and isinstance(promo_text, str):
            match = re.search(r'خصم\s*(\d+)\s*%', promo_text)
            if match:
                return f"{match.group(1)}%"
        
        # 2️⃣ التحويل إلى أرقام
        try:
            original = float(str(original_price).replace(',', '').strip()) if original_price else 0
            discounted = float(str(discounted_price).replace(',', '').strip()) if discounted_price else 0
        except (ValueError, TypeError):
            return ""
        
        if original <= 0 or discounted <= 0:
            return ""
        
        # 3️⃣ إذا كان المنتج خاضع للضريبة، نعيد السعر المخفض إلى ما قبل الضريبة
        if taxable:
            discounted_before_tax = remove_tax(discounted)
        else:
            discounted_before_tax = discounted
        
        # 4️⃣ حساب السعر الأصلي الإجمالي (السعر الأصلي × عدد الحبات)
        original_total = original * quantity
        
        # 5️⃣ حساب قيمة الخصم
        discount_amount = original_total - discounted_before_tax
        
        # 6️⃣ إذا كانت قيمة الخصم سالبة (يعني السعر المخفض أعلى من الأصلي)
        if discount_amount <= 0:
            return ""
        
        # 7️⃣ حساب النسبة المئوية
        percentage = (discount_amount / original_total) * 100
        
        return f"{round(percentage, 0)}%"
        
    except Exception:
        return ""

def calculate_discount_percentage_from_promo(original_price, promo_text, quantity=1, taxable=False):
    """
    حساب نسبة الخصم من النص الترويجي مباشرة
    """
    if not promo_text or not isinstance(promo_text, str):
        return ""
    
    try:
        original = float(str(original_price).replace(',', '').strip()) if original_price else 0
        if original <= 0:
            return ""
        
        # 1️⃣ إذا كان النص يحتوي على نسبة مئوية
        match = re.search(r'خصم\s*(\d+)\s*%', promo_text)
        if match:
            return f"{match.group(1)}%"
        
        # 2️⃣ إذا كان النص يحتوي على قيمة خصم بالريال (مثل خصم 17 ريال)
        discount_value = extract_discount_value_from_promo(promo_text)
        if discount_value:
            if taxable:
                discount_value = remove_tax(discount_value)
            original_total = original * quantity
            if original_total > 0:
                percentage = (discount_value / original_total) * 100
                return f"{round(percentage, 0)}%"
        
        # 3️⃣ إذا كان النص يحتوي على سعر محدد (مثل 6حبات ب 75ريال)
        promo_price = extract_price_from_promo(promo_text)
        if promo_price:
            if taxable:
                promo_price_before_tax = remove_tax(promo_price)
            else:
                promo_price_before_tax = promo_price
            
            original_total = original * quantity
            if original_total > 0 and promo_price_before_tax < original_total:
                discount_amount = original_total - promo_price_before_tax
                percentage = (discount_amount / original_total) * 100
                return f"{round(percentage, 0)}%"
        
        return ""
        
    except Exception:
        return ""

def extract_promo_details(promo_text, original_price, quantity=1, taxable=False):
    """
    استخراج جميع تفاصيل الخصم من النص الترويجي
    """
    if not promo_text or not isinstance(promo_text, str):
        return {"discount_percentage": "", "discount_value": "", "final_price": ""}
    
    try:
        original = float(str(original_price).replace(',', '').strip()) if original_price else 0
        if original <= 0:
            return {"discount_percentage": "", "discount_value": "", "final_price": ""}
        
        # 1️⃣ نسبة مئوية مباشرة
        match = re.search(r'خصم\s*(\d+)\s*%', promo_text)
        if match:
            return {
                "discount_percentage": f"{match.group(1)}%",
                "discount_value": "",
                "final_price": ""
            }
        
        # 2️⃣ قيمة خصم بالريال
        discount_value = extract_discount_value_from_promo(promo_text)
        if discount_value:
            if taxable:
                discount_value = remove_tax(discount_value)
            original_total = original * quantity
            if original_total > 0:
                percentage = (discount_value / original_total) * 100
                return {
                    "discount_percentage": f"{round(percentage, 0)}%",
                    "discount_value": discount_value,
                    "final_price": original_total - discount_value
                }
        
        # 3️⃣ سعر محدد (مثل 6حبات ب 75ريال)
        promo_price = extract_price_from_promo(promo_text)
        if promo_price:
            if taxable:
                promo_price_before_tax = remove_tax(promo_price)
            else:
                promo_price_before_tax = promo_price
            
            original_total = original * quantity
            if original_total > 0 and promo_price_before_tax < original_total:
                discount_amount = original_total - promo_price_before_tax
                percentage = (discount_amount / original_total) * 100
                return {
                    "discount_percentage": f"{round(percentage, 0)}%",
                    "discount_value": discount_amount,
                    "final_price": promo_price_before_tax
                }
        
        return {"discount_percentage": "", "discount_value": "", "final_price": ""}
        
    except Exception:
        return {"discount_percentage": "", "discount_value": "", "final_price": ""}

# ========== دوال التصدير ==========
def flatten_dataframe(df):
    """تسطيح الجدول وتقسيم علامات & إلى أعمدة بأسماء ديناميكية مسطحة ومنع تكرار الأسماء"""
    new_rows = []
    max_splits = {}
    for col in df.columns:
        max_splits[col] = df[col].astype(str).apply(lambda x: len(str(x).split('&')) if pd.notna(x) and '&' in str(x) else 1).max()

    custom_names = {
        "نسبة الخصم للمنتج": "نسبة الخصم (للمنتج)",
        "عدد حبات العرض للمنتج": "حبات العرض (للمنتج)",
        "سعر مخفض للمنتج": "سعر مخفض (للمنتج)",
        "العنوان الترويجي للمنتج": "ترويج (للمنتج)",
        "تاريخ نهاية التخفيض للمنتج": "نهاية تخفيض (للمنتج)",
        
        "رقم المنتج للمجموعة": "رقم المنتج (للمجموعة)",
        "اسم المنتج للمجموعة": "اسم المجموعة",
        "سعر المنتج للمجموعة": "سعر المجموعة",
        "سعر مخفض للمجموعة": "سعر مخفض للمجموعة",
        "عدد حبات المجموعة": "حبات المجموعة",
        "العنوان الترويجي للمجموعة": "ترويج المجموعة",
        "تاريخ نهاية التخفيض للمجموعة": "نهاية تخفيض المجموعة",
        "نسبة الخصم للمجموعة": "نسبة الخصم للمجموعة",
        "عدد حبات العرض للمجموعة": "حبات العرض للمجموعة"
    }

    new_cols = []
    for col in df.columns:
        if col == "اسم المنتج":
            new_cols.append("اسم (المنتج)")
        elif col == "سعر المنتج":
            new_cols.append("سعر (المنتج)")
        elif col in custom_names:
            base_name = custom_names[col]
            splits = max_splits[col]
            if splits == 1:
                new_cols.append(base_name)
            else:
                # 🌟 التعديل الجذري: ترقيم الأعمدة المتكررة لحمايتها من المسح في الإكسيل
                for i in range(1, splits + 1):
                    if "(" in base_name:
                        # مثال: "نسبة الخصم (للمنتج)" -> "نسبة الخصم (للمنتج 1)"
                        new_cols.append(base_name.replace(")", f" {i})"))
                    else:
                        # مثال: "اسم المجموعة" -> "اسم المجموعة (1)"
                        new_cols.append(f"{base_name} ({i})")
        else:
            splits = max_splits.get(col, 1)
            if splits == 1:
                new_cols.append(col)
            else:
                for i in range(1, splits + 1):
                    new_cols.append(f"{col} ({i})")

    for _, row in df.iterrows():
        new_row = []
        for col in df.columns:
            val = str(row[col]) if pd.notna(row[col]) and str(row[col]) != "nan" else ""
            splits = max_splits[col]
            if col in ["اسم المنتج", "سعر المنتج"]:
                new_row.append(val)
            elif splits > 1:
                parts = [p.strip() for p in val.split('&')]
                parts += [""] * (splits - len(parts))
                new_row.extend(parts)
            else:
                new_row.append(val)
        new_rows.append(new_row)

    return pd.DataFrame(new_rows, columns=new_cols)

def add_smart_comments(ws, df_flat):
    """إضافة تعليقات Hover الذكية للأعمدة المقسمة الخاصة بالمجموعات"""
    col_indices = {col: idx+1 for idx, col in enumerate(df_flat.columns)}
    
    for r_idx, row in enumerate(df_flat.to_dict('records'), start=2):
        group_comments = {}
        for i in range(1, 21):
            sku_col = f"رقم المنتج (للمجموعة {i})"
            if sku_col in row and row[sku_col]:
                sku = str(row[sku_col]).strip()
                name = str(row.get(f"اسم المجموعة ({i})", "")).strip()
                qty = str(row.get(f"حبات المجموعة ({i})", "")).strip()
                
                if not qty:
                    m = re.search(r'\*(\d+)', sku)
                    qty = m.group(1) if m else "1"
                    
                first_word = name.split()[0] if name else "منتج"
                comment_text = f"{first_word} - {qty}حبة - {sku}"
                group_comments[i] = comment_text
                
        for col_name, val in row.items():
            # 🌟 منع وضع تعليقات المجموعات على أعمدة المنتج الفردي
            if "للمنتج" in col_name: continue 
            
            val_str = str(val).strip()
            if not val_str or val_str == "nan": continue
            
            m = re.search(r'\(.*?(\d+)\)$', col_name)
            if m:
                g_idx = int(m.group(1))
                if g_idx in group_comments:
                    c_idx = col_indices[col_name]
                    cell = ws.cell(row=r_idx, column=c_idx)
                    comment = Comment(group_comments[g_idx], "نظام بلسم")
                    comment.width = 200
                    comment.height = 40
                    cell.comment = comment

def build_flat_excel_sheet(ws, df, is_main_sheet=False):
    """بناء الشيت بصف عناوين واحد وتطبيق الفلاتر والتعليقات والعمود المميز"""
    ws.views.sheetView[0].rightToLeft = True
    if df.empty: return

    header_fill = PatternFill(start_color="1F7A8C", end_color="1F7A8C", fill_type="solid")
    header_font = Font(name="Tajawal", size=11, bold=True, color="FFFFFF")
    
    # 🌟 ألوان العمود الأخير المميز (تنبيهي)
    special_header_fill = PatternFill(start_color="E26D5C", end_color="E26D5C", fill_type="solid") # برتقالي محمر
    special_col_fill = PatternFill(start_color="FDF3F1", end_color="FDF3F1", fill_type="solid")
    special_alt_col_fill = PatternFill(start_color="F9E4E0", end_color="F9E4E0", fill_type="solid")
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        # 🌟 تطبيق اللون المميز على عنوان العمود
        if col_name == "مجموعات بدون عروض":
            cell.fill = special_header_fill
        else:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border
        ws.column_dimensions[get_column_letter(c_idx)].width = max(20, len(str(col_name)) + 5)

    alt_row_fill = PatternFill(start_color="E6F3F5", end_color="E6F3F5", fill_type="solid")
    for r_idx, row in enumerate(df.to_dict('records'), 2):
        is_alt = (r_idx % 2 == 1)
        for c_idx, col_name in enumerate(df.columns, 1):
            val = row[col_name]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            cell.alignment = align_center
            
            # 🌟 تطبيق اللون المميز على خلايا العمود
            if col_name == "مجموعات بدون عروض":
                cell.fill = special_alt_col_fill if is_alt else special_col_fill
            else:
                if is_alt: cell.fill = alt_row_fill

    if is_main_sheet:
        add_smart_comments(ws, df)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    ws.freeze_panes = 'A2'

@st.cache_data(show_spinner=False)
def generate_excel_download_files(df_regular, df_mixed, price_map_dict):
    """توليد الملف النهائي وتجهيز الشيت الثالث (بدون عروض)"""
    output = BytesIO()
    wb = Workbook()
    wb.remove(wb.active) # مسح الشيت الافتراضي
    
    # --- 1. بناء شيت العروض الأساسية (المنتجات والعروض الخاصة) ---
    if not df_regular.empty:
        df_flat = flatten_dataframe(df_regular)
        ws_regular = wb.create_sheet("المنتجات الأساسية والعروض")
        build_flat_excel_sheet(ws_regular, df_flat, is_main_sheet=True)
        
    # --- 2. بناء شيت سعر مخفض للمجموعات (خاص فقط بما جاء من شيت سعر مخفض) ---
    if not df_mixed.empty:
        ws_mixed = wb.create_sheet("سعر مخفض للمجموعات")
        build_flat_excel_sheet(ws_mixed, df_mixed, is_main_sheet=False)
        
    # --- 3. بناء شيت المنتجات بدون عروض ---
    active_skus = set()
    
    # إضافة المنتجات من الشيت الأساسي
    for _, row in df_regular.iterrows():
        active_skus.add(str(row.get("رقم المنتج", "")).strip())
        for col in ["رقم المنتج للمجموعة", "رقم منتج العرض الخاص"]:
            val = str(row.get(col, ""))
            if val and val != "nan":
                for p in val.split('&'): # 🌟 استخدام الفاصل &
                    active_skus.add(p.strip())
                    
    # إضافة المنتجات الفردية من داخل المجموعات المعزولة
    for _, row in df_mixed.iterrows():
        val = str(row.get("رقم المنتج المجمع", "")).strip()
        if val and val != "nan":
            # تفكيك الأرقام المجمعة (مثل 2746*2-14776) لمعرفة المنتجات الفردية
            for p in re.split(r'[-+]', val):
                m = re.match(r'^(\d+)', p.strip())
                if m: active_skus.add(m.group(1))

    # مطابقة المنتجات مع شيت الأسعار الأصلي
    no_offer_records = []
    for sku, info in price_map_dict.items():
        if sku not in active_skus:
            no_offer_records.append({
                "رقم المنتج": sku,
                "اسم المنتج": info["name"],
                "سعر المنتج": info["price"],
                "خاضع للضريبة": "نعم" if info["has_tax"] else "لا"
            })
            
    df_no_offers = pd.DataFrame(no_offer_records)
    if not df_no_offers.empty:
        ws_no_offers = wb.create_sheet("منتجات بدون عروض")
        build_flat_excel_sheet(ws_no_offers, df_no_offers, is_main_sheet=False)
        
    # حماية في حال كان الملف فارغاً تماماً
    if len(wb.sheetnames) == 0:
        wb.create_sheet("لا توجد بيانات")
        
    wb.save(output)
    final_bytes = output.getvalue()
    
    return final_bytes, final_bytes
    
# ========== دالة العرض الرئيسية ==========
def show():
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800&display=swap');
        * { font-family: 'Tajawal', sans-serif; }
        .offers-header {
            background: linear-gradient(135deg, #0f4c5c 0%, #1f7a8c 50%, #16425b 100%);
            border-radius: 24px; padding: 1.5rem; color: white;
            margin-bottom: 1.5rem; text-align: center;
        }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="offers-header">
        <h1>🛍️ مركز معالجة وإدارة عروض المتجر</h1>
        <p>فصل أوتوماتيكي لسلاسل العروض الخاصة، تفعيل كشافات البحث الفوري، وتأمين الترويجات والتواريخ</p>
    </div>
    """, unsafe_allow_html=True)

    st.subheader("📂 رفع ملف التقرير الشامل")
    
    col_upload, col_template = st.columns([3, 1])
    with col_upload:
        uploaded_file = st.file_uploader("قم برفع ملف المبيعات والعروض المشترك", type=["xlsx", "xls"])
    with col_template:
        st.markdown("<div style='text-align: center; padding-top: 1.6rem;'>", unsafe_allow_html=True)
        template_bytes = generate_empty_template()
        st.download_button(
            label="📥 نموذج ملف العروض",
            data=template_bytes,
            file_name="نموذج_جميع_عروض_المتجر_الفعالة.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.markdown("</div>", unsafe_allow_html=True)
    
    if uploaded_file is not None:
        excel_data = pd.ExcelFile(uploaded_file)
        sheet_names = excel_data.sheet_names
        
        offers_sheet = None
        discounted_sheet = None
        prices_sheet = None
        special_offers_sheet = None  # 🌟 تعريف الشيت الرابع
        
        for sheet in sheet_names:
            if "وضع خاص" in sheet:
                special_offers_sheet = sheet
            elif "عرض خاص" in sheet or "عروض" in sheet:
                if "وضع خاص" not in sheet:  # لمنع التداخل
                    offers_sheet = sheet
            if "سعر مخفض" in sheet or "مخفض" in sheet: discounted_sheet = sheet
            if "اسعار المنتجات" in sheet or "سعر المنتج" in sheet or "أسعار المنتجات" in sheet: prices_sheet = sheet
        
        if not offers_sheet: offers_sheet = sheet_names[0]
        
        df_raw = pd.read_excel(uploaded_file, sheet_name=offers_sheet, header=None)
        df_discounted = pd.read_excel(uploaded_file, sheet_name=discounted_sheet) if discounted_sheet else pd.DataFrame()
        df_regular_prices = pd.read_excel(uploaded_file, sheet_name=prices_sheet) if prices_sheet else pd.DataFrame()
        
        # 🌟 قراءة شيت "الوضع الخاص" ليتم معالجته لاحقاً بقواعد الأعمدة A و D و E
        df_special = pd.DataFrame()
        if special_offers_sheet:
            df_special = pd.read_excel(uploaded_file, sheet_name=special_offers_sheet, header=None)
            if not df_special.empty:
                # نأخذ العمود الأول فقط لضمان دقة الدمج
                df_raw = pd.concat([df_raw.iloc[:, :1], df_special.iloc[:, :1]], ignore_index=True)
        
        df_raw = pd.read_excel(uploaded_file, sheet_name=offers_sheet, header=None)
        df_discounted = pd.read_excel(uploaded_file, sheet_name=discounted_sheet) if discounted_sheet else pd.DataFrame()
        df_regular_prices = pd.read_excel(uploaded_file, sheet_name=prices_sheet) if prices_sheet else pd.DataFrame()
        
        # ========== معالجة الأسعار العادية مع عمود الضريبة ==========
        price_map = {}  # {sku: {name, price, has_tax}}
        if not df_regular_prices.empty:
            sku_col_idx, name_col_idx, price_col_idx, tax_col_idx = 0, 1, 2, 3
            for i, c in enumerate(df_regular_prices.columns):
                c_str = str(c).lower()
                if 'sku' in c_str or 'رمز' in c_str: sku_col_idx = i
                elif 'اسم' in c_str or 'أسم' in c_str: name_col_idx = i
                elif 'سعر' in c_str: price_col_idx = i
                elif 'ضريبة' in c_str or 'tax' in c_str or 'خاضع' in c_str: tax_col_idx = i
            
            sku_col_choice = df_regular_prices.columns[sku_col_idx] if sku_col_idx < len(df_regular_prices.columns) else df_regular_prices.columns[0]
            name_col_choice = df_regular_prices.columns[name_col_idx] if name_col_idx < len(df_regular_prices.columns) else df_regular_prices.columns[1]
            price_col_choice = df_regular_prices.columns[price_col_idx] if price_col_idx < len(df_regular_prices.columns) else df_regular_prices.columns[2]
            tax_col_choice = df_regular_prices.columns[tax_col_idx] if tax_col_idx < len(df_regular_prices.columns) else None
            
            for _, row in df_regular_prices.iterrows():
                # 🌟 تطبيق دالة التنظيف هنا
                sku = clean_sku(row[sku_col_choice])
                if sku and sku != "nan":
                    price_map[sku] = {
                        "name": row[name_col_choice] if pd.notna(row[name_col_choice]) else "",
                        "price": row[price_col_choice] if pd.notna(row[price_col_choice]) else "",
                        "has_tax": str(row[tax_col_choice]).strip().lower() in ["نعم", "yes", "true", "1", "خاضع"] if tax_col_choice else False
                    }
        
        # ========== دوال حساب النسبة مع مراعاة الضريبة ==========
        def calculate_discount_percentage_advanced(original_price, discounted_price, promo_text="", quantity=1, has_tax=False):
            """
            حساب نسبة الخصم مع مراعاة الضريبة (15%)
            - إذا كان المنتج خاضعاً للضريبة، يتم إعادة السعر المخفض إلى ما قبل الضريبة
            - ثم حساب النسبة من السعر الأصلي
            """
            try:
                original = float(str(original_price).replace(',', '').strip()) if original_price else 0
                discounted = float(str(discounted_price).replace(',', '').strip()) if discounted_price else 0
            except (ValueError, TypeError):
                return "", 0
            
            if original <= 0 or discounted <= 0:
                return "", 0
            
            # إذا كان المنتج خاضعاً للضريبة (15%)، نعيد السعر المخفض إلى ما قبل الضريبة
            if has_tax:
                discounted_before_tax = discounted / 1.15
            else:
                discounted_before_tax = discounted
            
            # السعر الأصلي الإجمالي (السعر الأصلي × عدد الحبات)
            total_original = original * quantity
            
            # إذا كان السعر المخفض الإجمالي أكبر من السعر الأصلي الإجمالي، نستخدم السعر المخفض الفردي
            if discounted_before_tax > total_original:
                # محاولة استخدام السعر المخفض الفردي
                if discounted_before_tax <= original:
                    discount_amount = original - discounted_before_tax
                    percentage = (discount_amount / original) * 100
                else:
                    return "", 0
            else:
                discount_amount = total_original - discounted_before_tax
                percentage = (discount_amount / total_original) * 100
            
            # التقريب إلى أقرب عدد صحيح
            return f"{round(percentage, 0)}%", round(percentage, 0)
        
        def extract_discount_percentage_full(text, original_price=None, discounted_price=None, promo_text="", quantity=1, has_tax=False):
            """
            استخراج نسبة الخصم من جميع المصادر الممكنة مع مراعاة الضريبة والكميات
            """
            text = str(text) if not pd.isna(text) else ""
            
            # 🌟 خصم على الحبة/القطعة الثانية
            match_second = re.search(r'خصم\s*(\d+)\s*(?:%|بالمائة)\s*على\s*(?:الحبة|القطعة)\s*الثانية', text)
            if not match_second:
                match_second = re.search(r'خصم\s*(\d+)\s*%\s*على\s*القطعة', text)
                
            if match_second:
                discount_second = float(match_second.group(1))
                percentage = discount_second / 2 
                return f"{round(percentage, 0)}%", 2

            # 1️⃣ خصم بنسبة مئوية مباشرة (دعم كلمة خصم أو الرقم المئوي مباشرة)
            match = re.search(r'(?:خصم)?\s*(\d+)\s*%', text)
            if match and not match_second:
                return f"{match.group(1)}%", quantity
            
            # 2️⃣ خصم بقيمة ريال (مثل خصم 17 ريال)
            match = re.search(r'خصم\s*([\d.]+)\s*ريال', text)
            if match and original_price:
                try:
                    discount_value = float(match.group(1))
                    original = float(str(original_price).replace(',', '').strip())
                    total_original = original * quantity
                    
                    if total_original > 0:
                        if has_tax:
                            discount_before_tax = discount_value / 1.15
                        else:
                            discount_before_tax = discount_value
                            
                        percentage = (discount_before_tax / total_original) * 100
                        return f"{round(percentage, 0)}%", quantity
                except (ValueError, TypeError):
                    pass
            
            # 3️⃣ عروض الكميات (مثل 12حبة ب39.5ريال)
            qty_match = re.search(r'(\d+)\s*(?:حبة|حبات)\s*(?:ب|بسعر)\s*([\d.]+)', text)
            if qty_match and original_price:
                try:
                    qty = int(qty_match.group(1))
                    promo_price = float(qty_match.group(2))
                    original = float(str(original_price).replace(',', '').strip())
                    
                    if original > 0 and promo_price > 0:
                        total_original = original * qty
                        if has_tax:
                            promo_before_tax = promo_price / 1.15
                        else:
                            promo_before_tax = promo_price
                        
                        if promo_before_tax < total_original:
                            discount_amount = total_original - promo_before_tax
                            percentage = (discount_amount / total_original) * 100
                            return f"{round(percentage, 0)}%", qty
                except (ValueError, TypeError):
                    pass
            
            # 4️⃣ حساب النسبة من السعر الأصلي والمخفض مباشرة
            if original_price and discounted_price:
                try:
                    original = float(str(original_price).replace(',', '').strip())
                    discounted = float(str(discounted_price).replace(',', '').strip())
                    total_original = original
                    if discounted >= original and quantity > 1:
                        total_original = original * quantity

                    if total_original > 0 and discounted > 0 and total_original > discounted:
                        discount_amount = total_original - discounted
                        percentage = (discount_amount / total_original) * 100
                        # 🌟 التعديل هنا: إرجاع كمية العرض الصحيحة بدلاً من إرجاع صفر
                        return f"{round(percentage, 0)}%", quantity
                except (ValueError, TypeError):
                    pass
            
            # 5️⃣ عروض مجانية (تطبيق المنطق الحسابي السليم 100%)
            match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجاناً?', text)
            if match:
                paid_qty = int(match.group(1))
                free_qty = int(match.group(2))
                total_qty = paid_qty + free_qty
                percentage = (free_qty / total_qty) * 100
                return f"{round(percentage, 0)}%", total_qty
            
            # 6️⃣ صيغة "2 +1 مجانا"
            match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجانا', text)
            if match:
                paid_qty = int(match.group(1))
                free_qty = int(match.group(2))
                total_qty = paid_qty + free_qty
                percentage = (free_qty / total_qty) * 100
                return f"{round(percentage, 0)}%", total_qty
            
            return "", 0
        
        def extract_offer_quantity_advanced(text):
            """استخراج عدد حبات العرض من نص العرض"""
            text = str(text) if not pd.isna(text) else ""
            
            # 1️⃣ عروض مجانية (2+1 مجاناً)
            match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجاناً?', text)
            if match:
                paid = int(match.group(1))
                free = int(match.group(2))
                return f"{paid + free}"
            
            # 2️⃣ صيغة "2 +1 مجانا"
            match = re.search(r'(\d+)\s*\+\s*(\d+)\s*مجانا', text)
            if match:
                paid = int(match.group(1))
                free = int(match.group(2))
                return f"{paid + free}"
            
            # 3️⃣ عروض الكميات (6حبات بسعر 77 ريال)
            match = re.search(r'(\d+)\s*(?:حبة|حبات)\s*(?:ب|بسعر)', text)
            if match:
                return match.group(1)
            
            # 4️⃣ عروض الكميات (6حبات ب 77 ريال)
            match = re.search(r'(\d+)\s*(?:حبة|حبات)\s*(?:ب|بسعر)', text)
            if match:
                return match.group(1)
            
            # 5️⃣ خصم على الحبة الثانية
            if "الحبة الثانية" in text:
                return "2"
            
            # 6️⃣ خصم على القطعة الثانية
            if "القطعة الثانية" in text:
                return "2"
            
            # 7️⃣ إذا كان هناك رقم قبل "حبة" أو "حبات"
            match = re.search(r'(\d+)\s*حبات?', text)
            if match:
                return match.group(1)
            
            return "1"  # الافتراضي هو حبة واحدة
        
        # ========== معالجة الأسعار المخفضة ==========
        individual_discount_map = {}
        group_discount_map = {}
        sku_master = {}
        complex_discount_records = [] # 🌟 قائمة جديدة لعزل منتجات شيت السعر المخفض المجمعة
        
        # تعريف المتغيرات مسبقاً
        disc_sku_col = None
        disc_price_col = None
        disc_end_col = None
        disc_promo_col = None
        
        if not df_discounted.empty:
            st.subheader("🔧 تحديد أعمدة شيت 'سعر مخفض'")
            c1, c2, c3, c4 = st.columns(4)
            
            discounted_options = ["لا يوجد"] + list(df_discounted.columns)
            
            disc_sku_idx, disc_price_idx, disc_end_idx, disc_promo_idx = 0, 2, 3, 4
            for i, c in enumerate(df_discounted.columns):
                c_str = str(c).lower()
                if 'sku' in c_str or 'رمز' in c_str: disc_sku_idx = i
                elif 'مخفض' in c_str or 'سعر' in c_str: disc_price_idx = i
                elif 'نهاية' in c_str or 'تاريخ' in c_str: disc_end_idx = i
                elif 'ترويج' in c_str or 'عنوان' in c_str: disc_promo_idx = i
            
            default_sku_col = df_discounted.columns[disc_sku_idx] if disc_sku_idx < len(df_discounted.columns) else df_discounted.columns[0]
            default_price_col = df_discounted.columns[disc_price_idx] if disc_price_idx < len(df_discounted.columns) else df_discounted.columns[0]
            default_end_col = df_discounted.columns[disc_end_idx] if disc_end_idx < len(df_discounted.columns) else "لا يوجد"
            default_promo_col = df_discounted.columns[disc_promo_idx] if disc_promo_idx < len(df_discounted.columns) else "لا يوجد"
            
            with c1: disc_sku_col = st.selectbox("عمود معرف المنتج (SKU)", options=list(df_discounted.columns), index=list(df_discounted.columns).index(default_sku_col))
            with c2: disc_price_col = st.selectbox("عمود السعر المخفض", options=list(df_discounted.columns), index=list(df_discounted.columns).index(default_price_col))
            with c3: disc_end_col = st.selectbox("عمود نهاية التخفيض", options=discounted_options, index=discounted_options.index(default_end_col))
            with c4: disc_promo_col = st.selectbox("عمود العنوان الترويجي", options=discounted_options, index=discounted_options.index(default_promo_col))
            
            if disc_sku_col is None:
                st.warning("⚠️ لم يتم العثور على عمود معرف المنتج (SKU) في شيت 'سعر مخفض'")
            else:
                for _, row in df_discounted.iterrows():
                    # تنظيف رقم المنتج
                    sku_raw = clean_sku(row[disc_sku_col])
                    if not sku_raw or sku_raw == "nan": continue
                    
                    price_val = row[disc_price_col] if disc_price_col and disc_price_col in row and pd.notna(row[disc_price_col]) else ""
                    promo_val = row[disc_promo_col] if disc_promo_col and disc_promo_col != "لا يوجد" and disc_promo_col in row and pd.notna(row[disc_promo_col]) else ""
                    
                    # 🌟 التصفية الذكية: إذا كان السعر المخفض فارغاً والعنوان الترويجي فارغاً، تجاهل هذا المنتج تماماً لتسريع النظام
                    p_str = str(price_val).strip()
                    pr_str = str(promo_val).strip()
                    if (not p_str or p_str == "nan") and (not pr_str or pr_str == "nan"):
                        continue
                    
                    end_val = row[disc_end_col] if disc_end_col and disc_end_col != "لا يوجد" and disc_end_col in row and pd.notna(row[disc_end_col]) else ""
                    
                    base_sku, group_sku, group_qty, individual_skus, is_star, is_dash = parse_composite_sku(sku_raw)
                    
                    product_info = price_map.get(base_sku, {})
                    original_price = product_info.get("price", "")
                    has_tax = product_info.get("has_tax", False)
                    
                    quantity = 1
                    if is_star and group_qty:
                        quantity = group_qty
                    else:
                        qty_match = re.search(r'(\d+)\s*حبات?\s*ب', promo_val)
                        if not qty_match:
                            qty_match = re.search(r'(\d+)\s*حبة\s*بسعر', promo_val)
                        if qty_match:
                            quantity = int(qty_match.group(1))
                    
                    discount_percentage, extracted_qty = extract_discount_percentage_full(
                        text=promo_val, original_price=original_price, discounted_price=price_val,
                        promo_text=promo_val, quantity=quantity, has_tax=has_tax
                    )
                    
                    if extracted_qty > 0: quantity = extracted_qty
                    
                    offer_quantity = extract_offer_quantity_advanced(promo_val)
                    if is_star and group_qty: offer_quantity = str(group_qty)
                    
                    disc_payload = {
                        "discounted_price": price_val, "end_date": end_val,
                        "promo_title": promo_val, "discount_percentage": discount_percentage,
                        "offer_quantity": offer_quantity
                    }
                    
                    # 🌟 التعديل الجذري: العزل المباشر للمنتجات المجمعة في شيت السعر المخفض فقط
                    if is_dash:
                        complex_discount_records.append({
                            "رقم المنتج المجمع": sku_raw,
                            "السعر المخفض": price_val,
                            "العنوان الترويجي": promo_val,
                            "تاريخ نهاية التخفيض": end_val,
                            "نسبة الخصم": discount_percentage,
                            "عدد حبات العرض": offer_quantity
                        })
                    elif is_star:
                        group_discount_map[group_sku] = disc_payload
                        for ind_sku in individual_skus:
                            if ind_sku not in sku_master:
                                sku_master[ind_sku] = {"groups": set(), "special_offer_skus": set(), "offers": []}
                            sku_master[ind_sku]["groups"].add(group_sku)
                    else:
                        individual_discount_map[base_sku] = disc_payload
                        if base_sku not in sku_master:
                            sku_master[base_sku] = {"groups": set(), "special_offer_skus": set(), "offers": []}

        def expand_ranges_to_individuals(text):
            """توسيع النطاقات (مثل 1000:1020) وتحويل السلاسل المرتبطة بها إلى أرقام فردية"""
            if pd.isna(text): return ""
            text_str = str(text)
            
            def replace_complex_list(match):
                block = match.group(0)
                if ':' in block:
                    # فصل السلسلة بناءً على الشرطات
                    parts = re.split(r'-', block)
                    res = []
                    for p in parts:
                        if ':' in p:
                            sub_parts = p.split(':')
                            if len(sub_parts) == 2 and sub_parts[0].strip().isdigit() and sub_parts[1].strip().isdigit():
                                start, end = int(sub_parts[0].strip()), int(sub_parts[1].strip())
                                # توليد جميع الأرقام بين البداية والنهاية
                                if start < end and (end - start) <= 5000:
                                    res.extend([str(i) for i in range(start, end + 1)])
                                else:
                                    res.append(p)
                        else:
                            res.append(p)
                    # إرجاع الأرقام كقائمة مفصولة بمسافات ليتعرف عليها النظام كمنتجات فردية
                    return " " + " ".join(res) + " "
                return block
                
            # البحث عن الكتل التي تحتوي على أرقام ونقطتين وشرطات (مثل 18822:18860-18889)
            expanded_text = re.sub(r'(?:\d{3,6}[:-])+\d{3,6}', replace_complex_list, text_str)
            return expanded_text
    
        # ========== معالجة العروض ==========
        with st.spinner("🧠 جاري معالجة العروض..."):
            offers_col = df_raw.columns[0]
            # 🌟 المعالجة المخصصة لعروض شيت "وضع خاص"
            if not df_special.empty:            
                for idx, row in df_special.iterrows():
                    # التأكد من وجود أعمدة كافية (العمود A=0, D=3, E=4)
                    if len(row) <= 3: continue 
                    
                    sku_val = str(row[0]).strip()
                    # تجاهل العناوين أو الخلايا الفارغة
                    if not sku_val or sku_val == "nan" or "رقم" in sku_val or "sku" in sku_val.lower(): continue

                    promo_title = str(row[3]).strip() if len(row) > 3 and pd.notna(row[3]) else ""
                    offer_logic = str(row[4]).strip() if pd.notna(row[4]) else ""
                    
                    # إذا كان العنوان الترويجي فارغاً، نعوضه بمنطق العرض
                    if not promo_title or promo_title == "nan":
                        promo_title = offer_logic
                        
                    start_date, end_date = extract_dates(offer_logic)
                    is_daily_deal = "صفقة اليوم" in offer_logic
                    special_offer_sku = extract_special_offer_sku(sku_val)
                    
                    # فك النطاقات (مثل 1000:1020)
                    text_expanded = expand_ranges_to_individuals(sku_val)
                    
                    # 1. استخراج عروض المجموعات
                    groups_star = re.findall(r'(\d{3,6})\s*\*\s*(\d+)', text_expanded)
                    for base_sku, qty in groups_star:
                        group_sku = f"{base_sku}*{qty.strip()}"
                        
                        g_original = price_map.get(group_sku, {}).get("price", "")
                        g_tax = price_map.get(group_sku, {}).get("has_tax", False)
                        
                        g_discount_percentage, _ = extract_discount_percentage_full(
                            text=offer_logic, original_price=g_original, has_tax=g_tax
                        )
                        g_offer_quantity = extract_offer_quantity_advanced(offer_logic)
                        if not g_discount_percentage:
                            match = re.search(r'خصم\s*(\d+)\s*%', offer_logic)
                            if match: g_discount_percentage = f"{match.group(1)}%"
                            
                        if base_sku not in sku_master:
                            sku_master[base_sku] = {"groups": set(), "special_offer_skus": set(), "offers": []}
                        sku_master[base_sku]["groups"].add(group_sku)
                        
                        if group_sku not in group_discount_map:
                            group_discount_map[group_sku] = {
                                "promo_title": promo_title, "discount_percentage": g_discount_percentage,
                                "offer_quantity": g_offer_quantity, "end_date": end_date, "discounted_price": ""
                            }
                        else:
                            g_disc = group_discount_map[group_sku]
                            if promo_title and promo_title not in str(g_disc.get("promo_title", "")):
                                g_disc["promo_title"] = f"{g_disc.get('promo_title', '')} & {promo_title}".strip(" &")
                            if g_discount_percentage and not g_disc.get("discount_percentage"):
                                g_disc["discount_percentage"] = g_discount_percentage
                            if g_offer_quantity and g_disc.get("offer_quantity", "1") == "1":
                                g_disc["offer_quantity"] = g_offer_quantity
                    
                    # 2. استخراج السلاسل المفصولة بـ - أو + (العروض الخاصة المجمعة)
                    groups_dash = re.findall(r'(\d{3,6}(?:[-+]\d{3,6})+)', text_expanded)
                    for d_seq in groups_dash:
                        individual_skus = re.split(r'[-+]', d_seq)
                        g_discount_percentage, _ = extract_discount_percentage_full(text=offer_logic)
                        g_offer_quantity = extract_offer_quantity_advanced(offer_logic)
                        
                        if not g_discount_percentage:
                            match = re.search(r'خصم\s*(\d+)\s*%', offer_logic)
                            if match: g_discount_percentage = f"{match.group(1)}%"
                            
                        if d_seq not in group_discount_map:
                            group_discount_map[d_seq] = {
                                "promo_title": promo_title, "discount_percentage": g_discount_percentage,
                                "offer_quantity": g_offer_quantity, "end_date": end_date, "discounted_price": ""
                            }
                        
                        for sku in individual_skus:
                            if sku not in sku_master:
                                sku_master[sku] = {"groups": set(), "special_offer_skus": set(), "offers": []}
                            sku_master[sku]["special_offer_skus"].add(d_seq)
                    
                    # 3. استخراج الأرقام الفردية للمنتج وتوجيهها لأعمدة المنتج الفردي 
                    numbers = extract_numbers_from_text(text_expanded)
                    for sku in numbers:
                        original_price = price_map.get(sku, {}).get("price", "")
                        has_tax = price_map.get(sku, {}).get("has_tax", False)
                        
                        discount_percentage, _ = extract_discount_percentage_full(
                            text=offer_logic, original_price=original_price, has_tax=has_tax
                        )
                        offer_quantity = extract_offer_quantity_advanced(offer_logic)
                        if not discount_percentage:
                            match = re.search(r'خصم\s*(\d+)\s*%', offer_logic)
                            if match: discount_percentage = f"{match.group(1)}%"
                        
                        offer_payload = {
                            "name": promo_title, 
                            "start": start_date,
                            "end": end_date,
                            "is_daily_deal": is_daily_deal,
                            "discount_percentage": discount_percentage,
                            "offer_quantity": offer_quantity,
                            "special_offer_sku": special_offer_sku
                        }
                        
                        if sku not in sku_master:
                            sku_master[sku] = {"groups": set(), "special_offer_skus": set(), "offers": []}
                        sku_master[sku]["offers"].append(offer_payload)
            
            # ========== بناء النتيجة النهائية ==========
            final_results = []
            
            for base_sku in sorted(sku_master.keys()):
                data = sku_master[base_sku]
                product_info = price_map.get(base_sku, {"name": "", "price": "", "has_tax": False})
                
                ind_disc = individual_discount_map.get(base_sku, {})
                disc_price_item = ind_disc.get("discounted_price", "")
                disc_promo_item = ind_disc.get("promo_title", "")
                disc_end_item = ind_disc.get("end_date", "")
                
                unique_offers = {}
                for o in data["offers"]:
                    key = o["name"]
                    if key not in unique_offers:
                        unique_offers[key] = o
                
                offer_names = " & ".join([o["name"] for o in unique_offers.values()])
                offer_starts = " & ".join([o["start"] for o in unique_offers.values() if o["start"]])
                offer_ends = " & ".join([o["end"] for o in unique_offers.values() if o["end"]])
                is_daily = "نعم" if any(o["is_daily_deal"] for o in unique_offers.values()) else ""
                
                discount_percentages = []
                offer_quantities = []
                special_offer_skus = []
                
                for o in unique_offers.values():
                    if o.get("discount_percentage"): discount_percentages.append(o["discount_percentage"])
                    if o.get("offer_quantity"): offer_quantities.append(o["offer_quantity"])
                    if o.get("special_offer_sku"): special_offer_skus.append(o["special_offer_sku"])
                
                if ind_disc.get("discount_percentage"): discount_percentages.append(ind_disc["discount_percentage"])
                if ind_disc.get("offer_quantity"): offer_quantities.append(ind_disc["offer_quantity"])
                
                discount_percentage_str = " & ".join(list(dict.fromkeys([str(x) for x in discount_percentages if x])))
                offer_quantity_str = " & ".join(list(dict.fromkeys([str(x) for x in offer_quantities if x])))
                special_offer_sku_str = " & ".join(list(dict.fromkeys([str(x) for x in special_offer_skus if x])))
                
                g_skus_list, g_names_list, g_prices_list, g_qtys_list = [], [], [], []
                g_disc_prices_list, g_promos_list, g_ends_list = [], [], []
                g_discount_percentages, g_offer_quantities = [], []
                
                for g_sku in sorted(list(data["groups"])):
                    g_skus_list.append(g_sku)
                    g_info = price_map.get(g_sku, {})
                    g_name = g_info.get("name", "")
                    g_price = g_info.get("price", "")
                    
                    if not g_name:
                        base_m = re.match(r'^(\d+)', g_sku)
                        if base_m:
                            b_info = price_map.get(base_m.group(1), {})
                            g_name = b_info.get("name", "")
                            g_price = b_info.get("price", "")
                    
                    g_names_list.append(str(g_name))
                    g_prices_list.append(str(g_price))
                    
                    qty_val = ""
                    q_match = re.search(r'\*(\d+)', g_sku)
                    if q_match: qty_val = q_match.group(1)
                    g_qtys_list.append(qty_val)
                    
                    g_disc = group_discount_map.get(g_sku, {})
                    g_disc_prices_list.append(str(g_disc.get("discounted_price", "")))
                    g_promos_list.append(str(g_disc.get("promo_title", "")))
                    g_ends_list.append(str(g_disc.get("end_date", "")))
                    g_discount_percentages.append(str(g_disc.get("discount_percentage", "")))
                    g_offer_quantities.append(str(g_disc.get("offer_quantity", "")))
                
                group_sku_str = " & ".join(g_skus_list) if g_skus_list else ""
                group_name_str = " & ".join(g_names_list) if any(x != "" for x in g_names_list) else ""
                group_price_str = " & ".join(g_prices_list) if any(x != "" for x in g_prices_list) else ""
                group_disc_price_str = " & ".join([x for x in g_disc_prices_list if x != ""]) if g_disc_prices_list else ""
                group_qty_str = " & ".join(g_qtys_list) if any(x != "" for x in g_qtys_list) else ""
                group_promo_str = " & ".join([x for x in g_promos_list if x != ""]) if g_promos_list else ""
                group_end_str = " & ".join([x for x in g_ends_list if x != ""]) if g_ends_list else ""
                group_discount_percentage_str = " & ".join([x for x in g_discount_percentages if x != ""]) if g_discount_percentages else ""
                group_offer_quantity_str = " & ".join([x for x in g_offer_quantities if x != ""]) if g_offer_quantities else ""
                
                # 🌟 الكشاف: البحث عن المجموعات التي ليس لها عروض
                unoffered_groups = []
                prefix = f"{base_sku}*"
                for p_sku in price_map.keys():
                    if str(p_sku).startswith(prefix) and str(p_sku) not in data["groups"]:
                        unoffered_groups.append(str(p_sku))
                unoffered_groups_str = " ، ".join(unoffered_groups)
                
                final_results.append({
                    "رقم المنتج": base_sku,
                    "رقم المنتج للمجموعة": group_sku_str,
                    "رقم منتج العرض الخاص": special_offer_sku_str,
                    "اسم المنتج": product_info["name"],
                    "سعر المنتج": product_info["price"],
                    "اسم المنتج للمجموعة": group_name_str,
                    "سعر المنتج للمجموعة": group_price_str,
                    "اسم العرض الخاص": offer_names,
                    "نسبة الخصم للمنتج": discount_percentage_str,          
                    "عدد حبات العرض للمنتج": offer_quantity_str,           
                    "نسبة الخصم للمجموعة": group_discount_percentage_str,  
                    "عدد حبات العرض للمجموعة": group_offer_quantity_str,   
                    "بداية العرض": offer_starts,
                    "نهاية العرض": offer_ends,
                    "سعر مخفض للمنتج": disc_price_item,
                    "سعر مخفض للمجموعة": group_disc_price_str,
                    "عدد حبات المجموعة": group_qty_str,
                    "العنوان الترويجي للمنتج": disc_promo_item,
                    "العنوان الترويجي للمجموعة": group_promo_str,
                    "تاريخ نهاية التخفيض للمنتج": disc_end_item,
                    "تاريخ نهاية التخفيض للمجموعة": group_end_str,
                    "صفقة اليوم": is_daily,
                    "مجموعات بدون عروض": unoffered_groups_str # 👈 العمود الجديد
                })
            
            df_final = pd.DataFrame(final_results)
            df_final = df_final[df_final["رقم المنتج"].notna() & (df_final["رقم المنتج"] != "nan") & (df_final["رقم المنتج"] != "")]
            df_final = df_final[~df_final["رقم المنتج"].isin(["2024", "2025", "2026", "2027", "2028", "2029", "2030"])]
            df_final = df_final.drop_duplicates(subset=["رقم المنتج", "رقم المنتج للمجموعة", "رقم منتج العرض الخاص"], keep="first")
            
            column_order = [
                "رقم المنتج", "رقم المنتج للمجموعة", "رقم منتج العرض الخاص",
                "اسم المنتج", "سعر المنتج", "اسم المنتج للمجموعة", "سعر المنتج للمجموعة",
                "اسم العرض الخاص", 
                "نسبة الخصم للمنتج", "عدد حبات العرض للمنتج",           
                "نسبة الخصم للمجموعة", "عدد حبات العرض للمجموعة",       
                "بداية العرض", "نهاية العرض",
                "سعر مخفض للمنتج", "سعر مخفض للمجموعة", "عدد حبات المجموعة",
                "العنوان الترويجي للمنتج", "العنوان الترويجي للمجموعة",
                "تاريخ نهاية التخفيض للمنتج", "تاريخ نهاية التخفيض للمجموعة",
                "صفقة اليوم",
                "مجموعات بدون عروض" # 👈 تم وضعه في نهاية الترتيب
            ]
            df_final = df_final[[col for col in column_order if col in df_final.columns]]
            
        st.success(f"✅ تم معالجة {len(df_final)} منتج بنجاح")
        
        # عرض الإحصائيات
        col1, col2, col3, col4, col5 = st.columns(5)
        with col1: st.metric("📦 إجمالي المنتجات", f"{len(df_final):,}")
        with col2: st.metric("🏷️ مع عروض", f"{(df_final['اسم العرض الخاص'].astype(str).str.strip().str.len().gt(0)).sum():,}")
        
        # 🌟 التعديل هنا: فحص الأعمدة الجديدة المخصصة للمنتج والمجموعة معاً
        discount_mask = (df_final['نسبة الخصم للمنتج'].astype(str).str.strip().str.len().gt(0)) | \
                        (df_final['نسبة الخصم للمجموعة'].astype(str).str.strip().str.len().gt(0))
        with col3: st.metric("💰 نسبة الخصم", f"{discount_mask.sum():,}")
        
        qty_mask = (df_final['عدد حبات العرض للمنتج'].astype(str).str.strip().str.len().gt(0)) | \
                   (df_final['عدد حبات العرض للمجموعة'].astype(str).str.strip().str.len().gt(0))
        with col4: st.metric("🎯 عدد حبات العرض", f"{qty_mask.sum():,}")
        
        with col5: st.metric("📊 أنواع العروض", f"{df_final['اسم العرض الخاص'].nunique():,}")
        
        # ========== تصفية ==========
        st.subheader("🔍 استعلام وبحث سريع")
        search_term = st.text_input("أدخل رقم صنف، رقم مجموعة، أو سلسلة العرض الخاص للبحث:", placeholder="مثال: 5929")
        if search_term:
            search_term = search_term.strip()
            filtered_df = df_final[
                (df_final["رقم المنتج"].astype(str) == search_term) |
                (df_final["رقم المنتج للمجموعة"].astype(str).str.contains(search_term, na=False, regex=False)) |
                (df_final["رقم منتج العرض الخاص"].astype(str).str.contains(search_term, na=False, regex=False)) |
                (df_final["اسم المنتج"].str.contains(search_term, na=False, regex=False))
            ]
            if not filtered_df.empty:
                st.dataframe(filtered_df, use_container_width=True)
            else:
                st.warning(f"⚠️ لم يتم العثور على نتائج مطابقة لـ: '{search_term}'")
        
        # ========== عرض الجدول ==========
        st.subheader("📋 شاشة العرض والمراقبة الشاملة")
        st.dataframe(df_final, use_container_width=True, height=400)
        
        # ========== خيارات التحميل ==========
        st.subheader("💾 استخراج وحفظ التقارير")
        
        df_complex_discounts = pd.DataFrame(complex_discount_records)
        if not df_complex_discounts.empty:
            with st.expander("📦 استعراض المجموعات المعزولة (سعر مخفض)", expanded=False):
                st.dataframe(df_complex_discounts, use_container_width=True)
                
        col_dl1, col_dl2 = st.columns(2)
        
        # تمرير الجداول الثلاثة (بما فيها price_map) לדالة التصدير 🌟
        simple_bytes, detailed_bytes = generate_excel_download_files(df_final, df_complex_discounts, price_map)
        
        with col_dl1:
            st.download_button(
                label="📥 تحميل الملف المفصل (مع شيتات منفصلة)",
                data=detailed_bytes,
                file_name="تقرير_العروض_المفصل.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with col_dl2:
            st.download_button(
                label="📥 تحميل الملف المبسط (جدول واحد)",
                data=simple_bytes,
                file_name="العروض_والمنتجات_الموحد.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        # ========== تصفية حسب نوع العرض ==========
        if "اسم العرض الخاص" in df_final and len(df_final["اسم العرض الخاص"].dropna().unique()) > 0:
            st.subheader("📊 تصفية حسب نوع العرض")
            offer_types = ["الكل"] + sorted(df_final["اسم العرض الخاص"].dropna().unique().tolist())
            selected_type = st.selectbox("اختر نوع العرض:", offer_types)
            if selected_type != "الكل":
                st.dataframe(df_final[df_final["اسم العرض الخاص"] == selected_type], use_container_width=True)
    else:
        st.info("📂 يرجى رفع ملف Excel لبدء المعالجة")

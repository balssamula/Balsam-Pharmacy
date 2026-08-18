import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime
from utils.excel_processor import update_balances

def show():
    st.markdown(
        """
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800&display=swap');
            * { font-family: 'Tajawal', sans-serif; }
        </style>
        <div class="hero">
            <h1>🔄 تحديث أرصدة الفروع</h1>
            <p>رفع ملفات ABC و Salla لتحديث الأرصدة وبناء مستند التحديث الثنائي المتوافق مع المنصة</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    col1, col2 = st.columns(2)
    with col1:
        abc_file = st.file_uploader("📊 رفع ملف ABC (يبدأ من الصف 5)", type=["xlsx"], key="abc_balances")
    with col2:
        salla_file = st.file_uploader("📋 رفع ملف Salla", type=["xlsx"], key="salla_balances")
    
    if abc_file and salla_file:
        if st.button("🔄 تنفيذ تحديث الأرصدة", use_container_width=True):
            with st.spinner("جاري تحديث الأرصدة وبناء الترويسات الثنائية..."):
                result_df, result = update_balances(abc_file, salla_file)
                if result_df is not None:
                    st.success(f"✅ تم التحديث بنجاح! عدد الأصناف المحدثة والمعدلة: {result:,}")
                    st.dataframe(result_df.head(20), use_container_width=True)
                    
                    # 🧠 بناء المصنف سحابياً ببنية الصفين المتطابقة مع منصة سلة
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Salla Product Quantities Sheet"
                    
                    # إعدادات التنسيقات والخطوط الاحترافية
                    font_title = Font(name="Tajawal", size=11, bold=True, color="FFFFFF")
                    font_headers = Font(name="Tajawal", size=10, bold=True, color="FFFFFF")
                    font_data = Font(name="Tajawal", size=10)
                    
                    fill_title_prod = PatternFill(start_color="16425B", end_color="16425B", fill_type="solid")
                    fill_title_qty = PatternFill(start_color="1F7A8C", end_color="1F7A8C", fill_type="solid")
                    fill_headers = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
                    
                    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    border_thin = Border(
                        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
                    )
                    
                    # 🏢 [الصف الأول]: دمج الترويسات الكبرى
                    ws.merge_cells("A1:E1")
                    ws["A1"] = "بيانات المنتج"
                    ws.merge_cells("F1:AO1")
                    ws["F1"] = "الكميات"
                    
                    for col in range(1, 6):
                        cell = ws.cell(row=1, column=col)
                        cell.font = font_title
                        cell.fill = fill_title_prod
                        cell.alignment = align_center
                        cell.border = border_thin
                        
                    for col in range(6, 42):
                        cell = ws.cell(row=1, column=col)
                        cell.font = font_title
                        cell.fill = fill_title_qty
                        cell.alignment = align_center
                        cell.border = border_thin
                        
                    # 📋 [الصف الثاني]: حقن فروع صيدليات بلسم العلا الـ 17 والمؤشرات بالترتيب
                    headers_row2 = [
                        "No.", "النوع", "أسم المنتج", "رمز المنتج sku", "غير محدود الكمية", 
                        "الكمية في فرع تبوك القادسية وباقي المدن", "العرض في فرع تبوك القادسية وباقي المدن", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 8 النظيم", "العرض في فرع تبوك - صيدلية بلسم العلا 8 النظيم", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 9", "العرض في فرع العلا - صيدلية بلسم العلا 9", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 11 البوادي", "العرض في فرع تبوك - صيدلية بلسم العلا 11 البوادي", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 15 الصفا", "العرض في فرع تبوك - صيدلية بلسم العلا 15 الصفا", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 16 النخيل", "العرض في فرع تبوك - صيدلية بلسم العلا 16 النخيل", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 10 الريان", "العرض في فرع تبوك - صيدلية بلسم العلا 10 الريان", 
                        "الكمية في فرع تبوك صيدلية بلسم العلا 13 القادسية", "العرض في فرع تبوك صيدلية بلسم العلا 13 القادسية", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 12 العليا", "العرض في فرع تبوك - صيدلية بلسم العلا 12 العليا", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 14 المصيف", "العرض في فرع تبوك - صيدلية بلسم العلا 14 المصيف", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 1", "العرض في فرع العلا - صيدلية بلسم العلا 1", 
                        "الكمية في فرع العلا -- صيدلية بلسم العلا 2", "العرض في فرع العلا -- صيدلية بلسم العلا 2", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 3", "العرض في فرع العلا - صيدلية بلسم العلا 3", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 4 العذيب", "العرض في فرع العلا - صيدلية بلسم العلا 4 العذيب", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 5", "العرض في فرع العلا - صيدلية بلسم العلا 5", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 6 ابو راكة", "العرض في فرع العلا - صيدلية بلسم العلا 6 ابو راكة", 
                        "الكمية في فرع العلا - صيدلية بلسم العلا 7", "العرض في فرع العلا - صيدلية بلسم العلا 7", 
                        "الكمية في فرع تبوك - صيدلية بلسم العلا 17 الروضة", "العرض في فرع تبوك - صيدلية بلسم العلا 17 الروضة"
                    ]
                    
                    for col_idx, h_text in enumerate(headers_row2, 1):
                        cell = ws.cell(row=2, column=col_idx, value=h_text)
                        cell.font = font_headers
                        cell.fill = fill_headers
                        cell.alignment = align_center
                        cell.border = border_thin
                        
                    # 📊 [الصف الثالث فصاعداً]: صب البيانات وتحديث العروض (نعم / لا)
                    # بناء قائمة بأرقام الأعمدة التي تمثل "العرض في فرع" للتحكم בה
                    offer_columns = [i + 1 for i, val in enumerate(headers_row2) if str(val).startswith("العرض في")]
                    col_13_offer_idx = headers_row2.index("العرض في فرع تبوك صيدلية بلسم العلا 13 القادسية") + 1
                    col_7_offer_idx = headers_row2.index("العرض في فرع العلا - صيدلية بلسم العلا 7") + 1

                    for r_idx, row_values in enumerate(result_df.values, start=3):
                        for c_idx, val in enumerate(row_values, start=1):
                            
                            # 💡 التعديل: إذا كان العمود الحالي هو عمود "عرض"، نقوم بتغيير قيمته
                            if c_idx in offer_columns:
                                if c_idx == col_13_offer_idx or c_idx == col_7_offer_idx:
                                    val = "لا"
                                else:
                                    val = "نعم"
                            else:
                                # بالنسبة للأعمدة العادية
                                if isinstance(val, (np.integer, np.floating)):
                                    val = val.item()
                                elif pd.isna(val):
                                    val = ""
                                    
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.font = font_data
                            cell.alignment = align_center
                            cell.border = border_thin
                            
                    ws.row_dimensions[1].height = 28
                    ws.row_dimensions[2].height = 24
                    
                    for col_cols in ws.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col_cols)
                        col_letter = openpyxl.utils.get_column_letter(col_cols[0].column)
                        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 32)
                    
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    st.download_button(
                        "📥 تحميل ملف Salla المحدث",
                        data=output,
                        file_name=f"salla_updated_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    st.error(f"❌ خطأ في التحديث: {result}")

import streamlit as st
import pandas as pd
# 💡 [إصلاح مشكلة حجم الجدول]: زيادة الحد الأقصى للخلايا التي يمكن تلوينها في الواجهة
pd.set_option("styler.render.max_elements", 5000000)
from io import BytesIO
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from utils.database import (
    get_latest_upload_summary, get_all_sessions, get_session_items, 
    lock_session, unlock_session, activate_session, delete_session,
    fetch_active_items, get_all_last_logins, get_completed_items,
    reopen_case_by_item_key, hide_item_from_pharmacy, unhide_item_from_pharmacy,
    lock_item, unlock_item, save_case_note,
    get_manager_last_login, get_login_history,
    get_old_orders, get_old_orders_stats,
    get_old_invoices, get_old_invoices_stats,
    move_item_to_branch, get_available_branches,
    check_duplicate_across_branches
)
from utils.helpers import (
    is_cancelled_or_returned_status, is_pending_payment_status,
    get_tab_label, numeric_value
)
from utils.excel_processor import process_excel

def export_to_excel(dataframes_dict: dict) -> bytes:
    output = BytesIO()
    tab_colors = {
        "الإضافات": "4472C4", "الإرجاعات": "ED7D31",
        "طلبات_بدون_فاتورة": "70AD47", "فواتير_بدون_طلب": "FFC000",
        "فواتير_بعد_آخر_طلب": "9B59B6", "بانتظار_الدفع": "3498DB",
        "ملغي_ومسترجع": "E74C3C", "تم_الانتهاء": "27AE60",
        "مقارنة_الجلسات": "2A5298"
    }
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            if df is not None and not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                worksheet = writer.sheets[sheet_name[:31]]
                header_fill = PatternFill(start_color=tab_colors.get(sheet_name, "2A5298"), 
                                         end_color=tab_colors.get(sheet_name, "2A5298"), fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=12)
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for row in range(1, len(df) + 2):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if worksheet.cell(row=1, column=col).value == "الفرق":
                            diff_value = cell.value
                            if diff_value:
                                if diff_value > 0:
                                    cell.font = Font(color="008000", bold=True)
                                elif diff_value < 0:
                                    cell.font = Font(color="FF0000", bold=True)
            else:
                empty_df = pd.DataFrame({"ملاحظة": ["لا توجد بيانات"]})
                empty_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    output.seek(0)
    return output.getvalue()

def compare_sessions(session1_id: str, session2_id: str) -> pd.DataFrame:
    import sqlite3
    from utils.database import DB_PATH
    conn = sqlite3.connect(DB_PATH)
    query = """
        SELECT order_number, invoice_number, sku, product_name, pharmacy_name,
               salla_qty, abc_qty, (salla_qty - abc_qty) as difference,
               case_type, order_status, abc_pharmacist_name
        FROM reconciliation_items 
        WHERE upload_batch_id IN (?, ?) AND active = 1
    """
    df = pd.read_sql_query(query, conn, params=(session1_id, session2_id))
    conn.close()
    df['required_action'] = df['difference'].apply(
        lambda x: 'إضافة' if x > 0 else ('إرجاع' if x < 0 else 'مطابق')
    )
    return df.rename(columns={
        "order_number": "رقم الطلب", "invoice_number": "رقم الفاتورة",
        "sku": "SKU", "product_name": "المنتج", "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة", "abc_qty": "كمية ABC",
        "difference": "الفرق", "order_status": "حالة الطلب",
        "abc_pharmacist_name": "الصيدلي"
    })

def styled_dataframe(input_df):
    """تنسيق DataFrame مع تمييز الألوان حسب نوع الحالة"""
    if input_df.empty:
        return None
        
    # 💡 [الإصلاح الجذري]: إعادة ضبط الفهارس لتجنب انهيار Streamlit عند التلوين
    display_df = input_df.copy().reset_index(drop=True)
    
    # 💡 [حماية إضافية]: إزالة أي أعمدة متكررة قد توقف عملية العرض
    display_df = display_df.loc[:, ~display_df.columns.duplicated()]

    def highlight_rows(row):
        # تحويل القيم إلى سلاسل نصية للتعامل الآمن
        case_type = str(row.get('case_type', '')).strip() if 'case_type' in row else ''
        status = str(row.get('status', '')).strip() if 'status' in row else ''
        
        # 1. الحالات المكتملة (أخضر فاتح)
        if status == "تم":
            return ['background-color: #d4edda; color: #155724;'] * len(row)
            
        # 2. تمييز الحالات داخل التبويبات المدمجة
        if case_type in ['orphan_salla', 'orphan_abc']:
            return ['background-color: #fff3cd; color: #856404;'] * len(row)
        elif case_type == 'return':
            return ['background-color: #ffe0df; color: #491217;'] * len(row)
        elif case_type == 'addition':
            return ['background-color: #dff1ff; color: #084298;'] * len(row)
            
        return [''] * len(row)
    
    # توحيد روابط المسميات لمنع سقوط الإحصائيات في الداشبورد الإداري
    if 'case_type' in display_df.columns:
        display_df['label_type'] = display_df['case_type'].map({
            'addition': 'إضافة عادية ➕',
            'return': 'إرجاع عادي 🔄',
            'orphan_salla': 'طلب بدون فاتورة (سلة) 🛒',
            'orphan_abc': 'فاتورة بدون طلب (ABC) 📄',
            'post_cutoff_abc': 'فاتورة بعد آخر طلب ⏰'
        }).fillna(display_df['case_type'])
    
    display_df = display_df.rename(columns={
        "order_number": "رقم الطلب",
        "invoice_number": "رقم الفاتورة",
        "sku": "SKU",
        "product_name": "المنتج",
        "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة",
        "abc_qty": "كمية ABC",
        "difference": "الفرق",
        "order_status": "حالة الطلب",
        "type_label": "نوع الحالة",
        "status": "الحالة"
    })
    
    # التأكد مرة أخرى من عدم وجود أعمدة تحمل نفس الاسم بعد الترجمة
    display_df = display_df.loc[:, ~display_df.columns.duplicated()]
    
    return display_df.style.apply(highlight_rows, axis=1)

def render_table_with_click(df, tab_name, allow_move: bool = True):
    """عرض جدول مع إمكانية تحديد الصف وإظهار إجراءات منبثقة مع حماية الذاكرة التامة"""
    if df.empty:
        st.success("لا توجد بيانات في هذا القسم.")
        return
    
    # 💡 [الحل الجذري والنهائي لمنع انهيار السيرفر والذاكرة]:
    # نحدد سقفاً أقصى للعرض المرئي داخل المتصفح، مع بقاء زر التصدير محتفظاً بالبيانات كاملة
    MAX_ROWS_TO_RENDER = 500
    if len(df) > MAX_ROWS_TO_RENDER:
        st.warning(f"⚠️ يحتوي هذا القسم على {len(df)} صف. لحماية أداء التطبيق ومنع الانهيار المجمّع، تم عرض أول {MAX_ROWS_TO_RENDER} صف فقط بالجدول التفاعلي. يمكنك تصدير الملف بصيغة Excel لقراءة التقرير الكامل بكل سلاسة.")
        display_subset_df = df.head(MAX_ROWS_TO_RENDER).copy()
    else:
        display_subset_df = df.copy()
        
    styled_df = styled_dataframe(display_subset_df)
    if styled_df is not None:
        event = st.dataframe(
            styled_df,
            use_container_width=True,
            height=400,
            selection_mode="single-row",
            on_select="rerun"
        )
        
        if event.selection.rows:
            selected_idx = event.selection.rows[0]
            if 0 <= selected_idx < len(display_subset_df):
                row = display_subset_df.iloc[selected_idx]
                item_key = row.get('item_key', row.get('id', ''))
                if pd.isna(item_key) or item_key == '':
                    item_key = f"old_row_{selected_idx}"
                               
# ========== التحقق من وجود مكررات (النسخة المحمية من المسافات البادئة) ==========
                order_number = str(row.get('order_number', ''))
                sku = str(row.get('sku', ''))
                current_pharmacy = row.get('pharmacy_name', '')
                
                duplicate_warning = ""
                try:
                    from utils.database import check_duplicate_across_branches
                    duplicates = check_duplicate_across_branches(order_number, sku, current_pharmacy)
                    
                    if duplicates:
                        duplicate_warning = (
                            '<div style="background:#fff3cd; border-right:4px solid #ff9800; padding:0.75rem; margin-top:0.75rem; border-radius:10px; margin-bottom:0.75rem; direction:rtl; text-align:right;">'
                            '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:0.5rem;">'
                            '<span style="font-size:1.2rem;">⚠️</span>'
                            f'<span style="color:#856404; font-weight:bold;">تنبيه: يوجد نفس المنتج (SKU: {sku}) في فروع أخرى بموجب نفس رقم الطلب!</span>'
                            '</div>'
                            '<div style="margin-right:1.5rem;">'
                        )
                        for dup in duplicates:
                            dup_pharmacy = dup.get('pharmacy', 'غير معروف')
                            dup_status = dup.get('status', 'غير معروف')
                            dup_case = dup.get('case_type', 'غير معروف')
                            dup_invoice = dup.get('invoice_date', '')
                            invoice_str = f' | تاريخ الفاتورة: {dup_invoice[:16]}' if dup_invoice else ''
                            
                            duplicate_warning += (
                                '<div style="font-size:0.85rem; margin-bottom:0.4rem; padding:0.3rem 0; border-bottom:1px dashed #ffe0a3; color:#66521a;">'
                                f'🏥 <strong>{dup_pharmacy}</strong> | الحالة: {dup_status} | النوع: {dup_case}{invoice_str}'
                                '</div>'
                            )
                        duplicate_warning += '</div></div>'
                except Exception as e:
                    pass
                
                # عرض التنبيه بشكل سليم وبدون مسافات بادئة تكسر مفسر الماركداون
                if duplicate_warning:
                    st.markdown(duplicate_warning, unsafe_allow_html=True)
                
                # صندوق الإجراءات الإدارية الذي يليه مباشرة
                st.markdown(f"""
                <div style="background:#f0f2f6;border-radius:10px;padding:1rem;margin-top:1rem;border-right:4px solid #1f7a8c;">
                    <h4 style="margin:0 0 0.5rem 0;">🛠️ إجراءات الصف المحدد (الأرشيف التاريخي)</h4>
                    <p><strong>📋 مستند الحالة:</strong> {row.get('order_number', row.get('invoice_number', 'N/A'))} | 
                    <strong>🏷️ SKU:</strong> {row.get('sku', 'N/A')} | 
                    <strong>📦 المنتج:</strong> {str(row.get('product_name', 'N/A'))[:60]}</p>
                </div>
                """, unsafe_allow_html=True)
                
                # الصف الأول من الأزرار
                col1, col2, col3, col4 = st.columns(4)
                
                # زر الإخفاء/الإظهار
                is_hidden = row.get('hidden_from_pharmacy', 0) == 1
                if col1.button("🙈 إخفاء" if not is_hidden else "👁️ إظهار", key=f"hide_{tab_name}_{selected_idx}", use_container_width=True):
                    if is_hidden:
                        unhide_item_from_pharmacy(item_key)
                    else:
                        hide_item_from_pharmacy(item_key, st.session_state.username)
                    st.rerun()
                
                # زر القفل/الفتح
                is_locked = row.get('is_item_locked', 0) == 1
                if col2.button("🔒 قفل" if not is_locked else "🔓 فتح", key=f"lock_{tab_name}_{selected_idx}", use_container_width=True):
                    if is_locked:
                        unlock_item(item_key)
                    else:
                        lock_item(item_key, st.session_state.username)
                    st.rerun()
                
                # زر إعادة الفتح (للحالات المكتملة فقط)
                if row['status'] == "تم":
                    if col3.button("🔄 إعادة فتح", key=f"reopen_{tab_name}_{selected_idx}", use_container_width=True):
                        reopen_case_by_item_key(item_key)
                        st.rerun()
                
                # الصف الثاني من الأزرار (النقل والملحوظة)
                if allow_move and row['status'] != "تم":
                    st.markdown("---")
                    col_a, col_b, col_c, col_d = st.columns([2, 1, 2, 1])
                    
                    current_branch = row.get('pharmacy_name', '')
                    branches = get_available_branches(current_branch)
                    
                    if branches:
                        selected_branch = col_a.selectbox(
                            "🏥 نقل إلى فرع",
                            branches,
                            key=f"move_branch_{tab_name}_{selected_idx}"
                        )
                        
                        if col_b.button("🚚 نقل", key=f"move_{tab_name}_{selected_idx}", use_container_width=True):
                            if move_item_to_branch(item_key, selected_branch, st.session_state.username):
                                st.success(f"✅ تم نقل العنصر إلى {selected_branch}")
                                st.rerun()
                            else:
                                st.error("❌ فشل نقل العنصر")
                    
                    note = col_c.text_input("📝 ملحوظة", value=row.get('pharmacist_note', ''), key=f"note_{tab_name}_{selected_idx}")
                    
                    if col_d.button("💾 حفظ", key=f"save_note_{tab_name}_{selected_idx}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], row['pharmacy_name'], row['case_type'], note)
                        st.rerun()
                else:
                    # إذا كان النقل غير مسموح، نعرض فقط الملحوظة
                    st.markdown("---")
                    col_a, col_b = st.columns([3, 1])
                    note = col_a.text_input("📝 ملحوظة", value=row.get('pharmacist_note', ''), key=f"note_{tab_name}_{selected_idx}")
                    if col_b.button("💾 حفظ", key=f"save_note_{tab_name}_{selected_idx}", use_container_width=True):
                        save_case_note(row['order_number'], row['sku'], row['pharmacy_name'], row['case_type'], note)
                        st.rerun()

def render_old_items_table(df, title, is_orders=True):
    if df.empty:
        st.success(f"🎉 لا توجد {title} قديمة (أكثر من 6 أشهر)")
        return
    
    def highlight_old_rows(row):
        if is_cancelled_or_returned_status(row.get('order_status', '')):
            return ['background-color: #ffe5e5; color: #333'] * len(row)
        
        order_date = row.get('order_date', '')
        invoice_date = row.get('invoice_date', '')
        check_date = invoice_date if not is_orders else order_date
        
        if check_date and check_date != '':
            try:
                date_obj = datetime.strptime(str(check_date)[:10], "%Y-%m-%d")
                from datetime import timedelta
                if (datetime.now() - date_obj) > timedelta(days=180):
                    return ['background-color: #1a1a1a; color: white; font-weight: bold'] * len(row)
            except:
                pass
        
        if is_pending_payment_status(row.get('order_status', '')):
            return ['background-color: #fff4d6; color: #333'] * len(row)
        
        return [''] * len(row)
    
    display_df = df.copy()
    display_df = display_df.rename(columns={
        "order_number": "رقم الطلب",
        "invoice_number": "رقم الفاتورة",
        "sku": "SKU",
        "product_name": "المنتج",
        "pharmacy_name": "الفرع",
        "salla_qty": "كمية سلة",
        "abc_qty": "كمية ABC",
        "difference": "الفرق",
        "case_label": "نوع الحالة",
        "order_status": "حالة الطلب",
        "order_date": "تاريخ الطلب",
        "invoice_date": "تاريخ الفاتورة",
        "days_old": "عدد الأيام"
    })
    
    styled_df = display_df.style.apply(highlight_old_rows, axis=1)
    st.dataframe(styled_df, use_container_width=True, height=400)
    
    st.markdown("---")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(f"📊 إجمالي {title} القديمة", len(df))
    with col2:
        avg_days = df['days_old'].mean() if 'days_old' in df.columns else 0
        st.metric("📅 متوسط عدد الأيام", f"{avg_days:.0f} يوم")
    with col3:
        max_days = df['days_old'].max() if 'days_old' in df.columns else 0
        st.metric("⏰ أقدم عنصر", f"{max_days:.0f} يوم")

def show():
    st.markdown("""
    <div class="hero">
        <h1>👑 لوحة التحكم الإدارية</h1>
        <p>إدارة الطلبات والفواتير - متابعة الإضافات والإرجاعات - إدارة الجلسات</p>
    </div>
    """, unsafe_allow_html=True)
    
    manager_info = get_manager_last_login()
    login_history = get_login_history(10)
    
    col1, col2 = st.columns(2)
    with col1:
        st.markdown(f"""
        <div class="note-card">
            <strong>👑 آخر دخول للمدير العام:</strong><br>
            📅 {manager_info['last_login']}<br>
            🌐 IP: {manager_info['last_ip']}<br>
            👤 {manager_info['pharmacist_name']}
        </div>
        """, unsafe_allow_html=True)
    with col2:
        if not login_history.empty:
            st.markdown("### 📋 آخر محاولات الدخول")
            st.dataframe(login_history, use_container_width=True)
    
    st.markdown("---")
    
    with st.expander("📂 رفع ملف الطلبات والفواتير", expanded=True):
        uploaded_file = st.file_uploader("اختر ملف Excel", type=["xlsx"])
        if uploaded_file:
            if st.button("🔄 معالجة الملف", use_container_width=True, type="primary"):
                with st.spinner("جاري معالجة الملف..."):
                    results, upload_batch_id = process_excel(uploaded_file, st.session_state.username)
                if results is not None:
                    st.success(f"✅ تمت المعالجة بنجاح!")
                    st.balloons()
                    st.rerun()
    
    latest = get_latest_upload_summary()
    if latest:
        batch_id, file_name, uploaded_by, uploaded_at, total_cases, additions, returns, orphan_salla, orphan_abc, post_cutoff, is_locked, session_name = latest
        lock_status = "🔒 مقفلة" if is_locked else "🔓 مفتوحة"
        st.markdown(f"""
        <div class="note-card">
            <strong>📋 الجلسة النشطة:</strong> {session_name or 'غير مسماة'} &nbsp; | &nbsp;
            <strong>الملف:</strong> {file_name} &nbsp; | &nbsp;
            <strong>بواسطة:</strong> {uploaded_by} &nbsp; | &nbsp;
            <strong>التاريخ:</strong> {uploaded_at[:16] if uploaded_at else ''} &nbsp; | &nbsp;
            <strong>الحالة:</strong> {lock_status}
        </div>
        """, unsafe_allow_html=True)
    
    with st.expander("📋 إدارة الجلسات السابقة", expanded=False):
        sessions_df = get_all_sessions()
        if not sessions_df.empty:
            for _, session in sessions_df.iterrows():
                col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 2])
                session_name_val = session.get('session_name', session['upload_batch_id'][:8])
                with col1:
                    st.markdown(f'<div class="session-card"><strong>📅 {session_name_val}</strong><br><small>{session["file_name"][:35]}</small></div>', unsafe_allow_html=True)
                with col2:
                    is_active = session.get('is_active', 0)
                    active_badge = "✅ نشطة" if is_active else "⏸ غير نشطة"
                    st.markdown(f'<div class="session-card"><small>👤 {session["uploaded_by"]}<br>📅 {session["uploaded_at"][:16]}<br>{active_badge}</small></div>', unsafe_allow_html=True)
                with col3:
                    st.markdown(f'<div class="session-card"><small>📊 {int(session.get("total_cases", 0))} حالة<br>➕ {int(session.get("total_additions", 0))}<br>➖ {int(session.get("total_returns", 0))}</small></div>', unsafe_allow_html=True)
                with col4:
                    is_locked = session.get('is_locked', 0)
                    lock_text = "🔒 مقفلة" if is_locked else "🔓 مفتوحة"
                    st.markdown(f'<div class="session-card" style="text-align:center;"><small>{lock_text}</small></div>', unsafe_allow_html=True)
                with col5:
                    btn1, btn2, btn3, btn4 = st.columns(4)
                    with btn1:
                        if not is_locked:
                            if st.button(f"🔒", key=f"lock_{session['upload_batch_id']}"):
                                lock_session(session['upload_batch_id'], st.session_state.username)
                                st.rerun()
                        else:
                            if st.button(f"🔓", key=f"unlock_{session['upload_batch_id']}"):
                                unlock_session(session['upload_batch_id'])
                                st.rerun()
                    with btn2:
                        if not is_active:
                            if st.button(f"⭐", key=f"activate_{session['upload_batch_id']}"):
                                activate_session(session['upload_batch_id'])
                                st.rerun()
                    with btn3:
                        if st.button(f"👁️", key=f"view_{session['upload_batch_id']}"):
                            st.session_state.view_session_id = session['upload_batch_id']
                            st.rerun()
                    with btn4:
                        if st.button(f"🗑️", key=f"delete_{session['upload_batch_id']}"):
                            delete_session(session['upload_batch_id'])
                            st.rerun()
            st.markdown("---")
        else:
            st.info("لا توجد جلسات سابقة")

    active_df = fetch_active_items()
    if not active_df.empty:
        st.markdown("### 📈 التحليل الإحصائي الفوري للجلسة النشطة")
        chart_col1, chart_col2 = st.columns(2)
    
        with chart_col1:
            case_counts = active_df['case_label'].value_counts().reset_index()
            case_counts.columns = ['نوع الحالة', 'العدد']
            fig_pie = px.pie(case_counts, values='العدد', names='نوع الحالة', 
                             hole=0.4, color_discrete_sequence=['#4472C4', '#ED7D31', '#70AD47', '#FFC000'])
            fig_pie.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=280, showlegend=True)
            st.plotly_chart(fig_pie, use_container_width=True)
        
        with chart_col2:
            branch_counts = active_df['pharmacy_name'].value_counts().reset_index()
            branch_counts.columns = ['الفرع', 'عدد الحالات']
            fig_bar = px.bar(branch_counts.head(7), x='عدد الحالات', y='الفرع', orientation='h',
                             color='عدد الحالات', color_continuous_scale='Viridis')
            fig_bar.update_layout(margin=dict(t=20, b=20, l=20, r=20), height=280, coloraxis_showscale=False)
            st.plotly_chart(fig_bar, use_container_width=True)
        
    with st.expander("🔄 مقارنة الجلسات", expanded=False):
        sessions_list = get_all_sessions()
        if not sessions_list.empty:
            session_options = {f"{row['session_name']} ({row['uploaded_at'][:16]})": row['upload_batch_id'] 
                              for _, row in sessions_list.iterrows()}
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                session1 = st.selectbox("اختر الجلسة الأولى", list(session_options.keys()), key="session1")
            with col2:
                session2 = st.selectbox("اختر الجلسة الثانية", list(session_options.keys()), key="session2")
            with col3:
                if st.button("📊 مقارنة", use_container_width=True):
                    with st.spinner("جاري المقارنة..."):
                        comparison_df = compare_sessions(session_options[session1], session_options[session2])
                        st.session_state.comparison_result = comparison_df
                        st.success(f"✅ تمت المقارنة!")
            if st.session_state.get('comparison_result') is not None:
                st.dataframe(st.session_state.comparison_result, use_container_width=True)
                excel_data = export_to_excel({"مقارنة_الجلسات": st.session_state.comparison_result})
                st.download_button("📥 تحميل تقرير المقارنة", data=excel_data, type="primary", 
                    file_name=f"session_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        else:
            st.info("لا توجد جلسات للمقارنة")
    
    if st.session_state.get('view_session_id'):
        with st.expander("📄 عرض الجلسة المحددة", expanded=True):
            session_items = get_session_items(st.session_state.view_session_id)
            if not session_items.empty:
                st.dataframe(session_items, use_container_width=True)
            if st.button("إغلاق العرض", use_container_width=True):
                del st.session_state.view_session_id
                st.rerun()
    
    df = fetch_active_items(include_hidden=True)
    if df.empty:
        st.info("📂 لا توجد بيانات فعالة بعد. ارفع ملف Excel من الأعلى لبدء التحليل.")
        return
    
    active_mask = ~df["order_status"].apply(is_cancelled_or_returned_status)
    active_df = df[active_mask]
    
    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)
    with col1:
        st.metric("📊 إجمالي الحالات", len(active_df))
    with col2:
        st.metric("➕ إضافات", len(active_df[active_df["case_type"] == "addition"]))
    with col3:
        st.metric("➖ إرجاعات", len(active_df[active_df["case_type"] == "return"]))
    with col4:
        st.metric("📦 طلبات بدون فاتورة", len(active_df[active_df["case_type"] == "orphan_salla"]))
    with col5:
        st.metric("🧾 فواتير بدون طلب", len(active_df[active_df["case_type"] == "orphan_abc"]))
    with col6:
        st.metric("⏰ فواتير بعد آخر طلب", len(active_df[active_df["case_type"] == "post_cutoff_abc"]))
    with col7:
        st.metric("✅ تم إنجازها", len(df[df["status"] == "تم"]))
    
    col1, col2 = st.columns([1, 6])
    with col1:
        if st.button("🔄 تحديث الصفحة", use_container_width=True):
            st.rerun()
    with col2:
        if st.button("📥 تصدير كل التقارير الحالية إلى ملف Excel موحد"):
            excel_data = export_to_excel({
                "الإضافات والطلبات المفقودة": additions_filtered,
                "الإرجاعات والفواتير المعلقة": returns_filtered,
                "فواتير معلقة بين الفروع": conflicts_filtered, # ⚡ سيصدر الجدول المدمج بالطلب القديم المعلق هنا
                "فواتير بعد آخر طلب": post_cutoff_filtered,
                "بانتظار الدفع": payment_filtered,
                "الملغيات والمسترجعات": cancelled_filtered,
                "الطلبات القديمة التاريخية": old_orders_df, # 📥 إضافة تبويب الطلبات القديمة كاملاً بالأرشيف
                "الفواتير القديمة التاريخية": old_invoices_df # 📥 إضافة تبويب الفواتير القديمة كاملاً بالأرشيف
            })
    
            st.download_button(
                label="💾 اضغط هنا لتحميل ملف Excel الموحد للإدارة",
                data=excel_data,
                file_name=f"Balsam_Admin_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # ========== الفلاتر المتقدمة وعزل الجلسات ==========
    sessions_list = get_all_sessions()
    
    # صف الفلاتر الأول: يحتوي على فلتر عزل الجلسات لمنع تكدس صفوف الملفات القديمة
    col_sess, col1, col2 = st.columns(3)
    with col_sess:
        if not sessions_list.empty:
            session_options = {"📋 كل الجلسات التاريخية": "الكل"}
            for _, s_row in sessions_list.iterrows():
                s_name = s_row.get('session_name', s_row['upload_batch_id'][:8]) or "جلسة غير مسماة"
                session_options[f"📂 {s_name} ({str(s_row['uploaded_at'])[:10]})"] = s_row['upload_batch_id']
            selected_session_label = st.selectbox("📂 فلتر الجلسة / الملف المرفوع", list(session_options.keys()))
            selected_session_id = session_options[selected_session_label]
        else:
            selected_session_id = "الكل"
            st.selectbox("📂 فلتر الجلسة / الملف المرفوع", ["لا توجد جلسات مرفوعة"])
            
    with col1:
        branch_options = ["الكل"] + sorted(df["pharmacy_name"].dropna().astype(str).unique().tolist())
        selected_branch = st.selectbox("🏥 فلتر الفرع", branch_options)
    with col2:
        status_filter = st.selectbox("📌 فلتر حالة الإجراء", ["الكل", "قيد المتابعة", "تم"])

    # صف الفلاتر الثاني
    col3, col4, col5, col6 = st.columns(4)
    with col3:
        order_status_options = ["الكل", "تم التوصيل", "ملغي", "مسترجع", "بانتظار الدفع", "تم الاستلام من فرع"]
        selected_order_status = st.selectbox("📋 فلتر حالة الطلب", order_status_options)
    with col4:
        search_order = st.text_input("🔢 رقم الطلب", placeholder="بحث برقم الطلب...")
    with col5:
        search_invoice = st.text_input("🧾 رقم الفاتورة", placeholder="بحث برقم الفاتورة...")
    with col6:
        search_sku = st.text_input("🏷️ SKU", placeholder="بحث بـ SKU...")

    # تطبيق الفلاتر ديناميكياً (تم حذق السطر المتكرر المسبب للمشكلة التاريخية)
    filtered_df = df.copy()
    if selected_session_id != "الكل":
        filtered_df = filtered_df[filtered_df["upload_batch_id"] == selected_session_id]

    if selected_branch != "الكل":
        filtered_df = filtered_df[filtered_df["pharmacy_name"] == selected_branch]
    if status_filter != "الكل":
        filtered_df = filtered_df[filtered_df["status"] == ("تم" if status_filter == "تم" else "قيد المتابعة")]
    if selected_order_status != "الكل":
        if selected_order_status == "تم الاستلام من فرع":
            filtered_df = filtered_df[filtered_df["order_status"].str.contains("تم الاستلام من فرع", na=False)]
        else:
            filtered_df = filtered_df[filtered_df["order_status"] == selected_order_status]
    if search_order:
        filtered_df = filtered_df[filtered_df["order_number"].astype(str).str.contains(search_order, na=False)]
    if search_invoice:
        filtered_df = filtered_df[filtered_df["invoice_number"].astype(str).str.contains(search_invoice, na=False)]
    if search_sku:
        filtered_df = filtered_df[filtered_df["sku"].astype(str).str.contains(search_sku, na=False)]

    active_mask_filtered = ~filtered_df["order_status"].apply(is_cancelled_or_returned_status)
    payment_mask = filtered_df["order_status"].apply(is_pending_payment_status)
    cancelled_mask = filtered_df["order_status"].apply(is_cancelled_or_returned_status)
    
    completed_df = get_completed_items()
    if selected_branch != "الكل":
        completed_df = completed_df[completed_df["pharmacy_name"] == selected_branch]
    
    # ========== العناصر القديمة ==========
    selected_branch_name = None if selected_branch == "الكل" else selected_branch
    
    old_orders_data = get_old_orders(pharmacy_name=selected_branch_name, months=6)
    old_invoices_data = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
    
    old_order_numbers = set(old_orders_data['order_number'].astype(str).tolist()) if not old_orders_data.empty else set()
    old_invoice_numbers = set(old_invoices_data['invoice_number'].astype(str).tolist()) if not old_invoices_data.empty else set()
    
    def exclude_old_items(temp_df, exclude_orders=True, exclude_invoices=True):
        if temp_df.empty:
            return temp_df
        result_temp = temp_df.copy()
        if exclude_orders and 'order_number' in result_temp.columns:
            result_temp = result_temp[~result_temp['order_number'].astype(str).isin(old_order_numbers)]
        if exclude_invoices and 'invoice_number' in result_temp.columns:
            result_temp = result_temp[~result_temp['invoice_number'].astype(str).isin(old_invoice_numbers)]
        return result_temp
    
# ========== حساب الإحصائيات الصحيحة للتبويبات بناءً على الشروط الدقيقة والمحدثة ==========
    filtered_df['status_clean'] = filtered_df['status'].astype(str).str.strip()
    active_mask_filtered = filtered_df['status_clean'].isin(['قيد المتابعة', 'بانتظار المراجعة', 'معلق', ''])

    # 1️⃣ بناء أقنعة الفرز الصارمة بناءً على حالة الطلب (لمنع تداخل الصفوف تاريخياً)
    is_cancelled_returned = filtered_df["order_status"].apply(is_cancelled_or_returned_status)
    is_pending_payment = filtered_df["order_status"].apply(is_pending_payment_status)
    is_normal_order = ~(is_cancelled_returned | is_pending_payment)

    # 2️⃣ تبويب الإضافات والطلبات المفقودة: أصناف سلة بدون فواتير أو كميتها أكبر (شرط أن تكون عادية وليست ملغية/معلقة)
    additions_base_df = filtered_df[
        is_normal_order & 
        active_mask_filtered & 
        (
            (filtered_df['case_type'] == 'orphan_salla') | 
            ((filtered_df['case_type'] == 'addition') & (filtered_df['difference'] > 0))
        )
    ].copy()

    additions_merged_df = exclude_old_items(additions_base_df)
    total_additions_merged = len(additions_merged_df)
    completed_additions_merged = int((additions_merged_df['status_clean'] == 'تم').sum())

    # 3️⃣ تبويب الإرجاعات والفواتير المعلقة: فواتير ABC بدون طلب أو كميتها أكبر + فواتير الطلبات الملغية أو المسترجعة لـتأكيد الإرجاع
    returns_base_df = filtered_df[
        active_mask_filtered & (
            (is_normal_order & ((filtered_df['case_type'] == 'orphan_abc') | ((filtered_df['case_type'] == 'return') & (filtered_df['difference'] < 0)))) |
            (is_cancelled_returned & filtered_df['case_type'].isin(['orphan_abc', 'return']))
        )
    ].copy()

    returns_merged_df = exclude_old_items(returns_base_df)
    total_returns_merged = len(returns_merged_df)
    completed_returns_merged = int((returns_merged_df['status_clean'] == 'تم').sum())

    # 3️⃣ تبويب فواتير معلقة بين الفروع (البيانات النشطة الحالية)
    conflicts_filtered = filtered_df[filtered_df["case_type"] == "branch_conflict"].copy()
    conflicts_filtered = exclude_old_items(conflicts_filtered)
    
    # جلب الطلبات القديمة التي تمتلك تنبيه تداخل فروع ودمجها في التبويب
    old_orders_df = get_old_orders(months=6)  # جلب الطلبات القديمة من قاعدة البيانات
    
    if not old_orders_df.empty:
        old_orders_filtered_for_tab = old_orders_df.copy()
        
        # 💡 [تم الإصلاح]: تغيير selected_pharmacy إلى المتغير الصحيح المعرّف في صفحتك selected_branch
        if selected_branch != "الكل":
            old_orders_filtered_for_tab = old_orders_filtered_for_tab[old_orders_filtered_for_tab["pharmacy_name"] == selected_branch]
        
        # الفلترة الذكية: اختيار العناصر التي لها تنبيه تداخل فروع
        old_conflicts = old_orders_filtered_for_tab[
            (old_orders_filtered_for_tab["case_type"] == "branch_conflict") | 
            (old_orders_filtered_for_tab["order_number"].isin(conflicts_filtered["order_number"]))
        ].copy()
        
        # دمج الطلبات القديمة المكررة مع الفواتير المعلقة النشطة في جدول واحد
        if not old_conflicts.empty:
            conflicts_filtered = pd.concat([conflicts_filtered, old_conflicts], ignore_index=True)
            if 'item_key' in conflicts_filtered.columns:
                conflicts_filtered = conflicts_filtered.drop_duplicates(subset=['item_key'])

    # حساب الإجمالي الكلي للتبويب بعد عملية الدمج والتصفية
    total_conflicts = len(conflicts_filtered)
    completed_conflicts = len(conflicts_filtered[conflicts_filtered["status"] == "تم"])
    
    # 4️⃣ تبويب فواتير بعد آخر طلب (ABC)
    post_cutoff_filtered = filtered_df[(filtered_df["case_type"] == "post_cutoff_abc") & active_mask_filtered].copy()
    post_cutoff_filtered = exclude_old_items(post_cutoff_filtered)
    total_post_cutoff = len(post_cutoff_filtered)
    completed_post_cutoff = len(post_cutoff_filtered[post_cutoff_filtered["status"] == "تم"])
    
    # 5️⃣ تبويب بانتظار الدفع (تم إضافة شرط استبعاد الفواتير المعلقة المتداخلة)
    payment_filtered = filtered_df[is_pending_payment & (filtered_df["case_type"] != "branch_conflict") & active_mask_filtered].copy()
    payment_filtered = exclude_old_items(payment_filtered)
    total_payment = len(payment_filtered)
    
    # 6️⃣ تبويب ملغي ومسترجع (تم إضافة شرط استبعاد الفواتير المعلقة المتداخلة)
    cancelled_filtered = filtered_df[is_cancelled_returned & (filtered_df["case_type"] != "branch_conflict")].copy()
    cancelled_filtered = exclude_old_items(cancelled_filtered)
    total_cancelled = len(cancelled_filtered)
    
    completed_filtered = completed_df.copy()
    completed_filtered = exclude_old_items(completed_filtered)
    total_completed = len(completed_filtered)
    
    old_orders_filtered = get_old_orders(pharmacy_name=selected_branch_name, months=6)
    old_invoices_filtered = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
    
    # ========== تصدير Excel ==========
    if st.session_state.get('show_export', False):
        old_orders_for_export = get_old_orders(pharmacy_name=selected_branch_name, months=6)
        old_invoices_for_export = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
        
        stats_data = []
        if not old_orders_for_export.empty:
            stats_data.append(["إجمالي الطلبات القديمة", len(old_orders_for_export)])
            stats_data.append(["إضافات قديمة (طلبات)", len(old_orders_for_export[old_orders_for_export["case_type"] == "addition"])])
            stats_data.append(["إرجاعات قديمة (طلبات)", len(old_orders_for_export[old_orders_for_export["case_type"] == "return"])])
            stats_data.append(["طلبات بدون فاتورة قديمة", len(old_orders_for_export[old_orders_for_export["case_type"] == "orphan_salla"])])
        else:
            stats_data.append(["إجمالي الطلبات القديمة", 0])
            stats_data.append(["إضافات قديمة (طلبات)", 0])
            stats_data.append(["إرجاعات قديمة (طلبات)", 0])
            stats_data.append(["طلبات بدون فاتورة قديمة", 0])
        
        if not old_invoices_for_export.empty:
            stats_data.append(["إجمالي الفواتير القديمة", len(old_invoices_for_export)])
            stats_data.append(["إضافات قديمة (فواتير)", len(old_invoices_for_export[old_invoices_for_export["case_type"] == "addition"])])
            stats_data.append(["إرجاعات قديمة (فواتير)", len(old_invoices_for_export[old_invoices_for_export["case_type"] == "return"])])
            stats_data.append(["فواتير بدون طلب قديمة", len(old_invoices_for_export[old_invoices_for_export["case_type"] == "orphan_abc"])])
        else:
            stats_data.append(["إجمالي الفواتير القديمة", 0])
            stats_data.append(["إضافات قديمة (فواتير)", 0])
            stats_data.append(["إرجاعات قديمة (فواتير)", 0])
            stats_data.append(["فواتير بدون طلب قديمة", 0])
        
        stats_data.append(["إجمالي العناصر القديمة", len(old_orders_for_export) + len(old_invoices_for_export)])
        stats_df = pd.DataFrame(stats_data, columns=["المقياس", "القيمة"])
        
        export_data = {
            "01_الإضافات_والطلبات_بدون_فاتورة": additions_merged_df,
            "02_الإرجاعات_والفواتير_بدون_طلب": returns_merged_df,
            "03_فواتير_بعد_آخر_طلب": post_cutoff_filtered,
            "04_بانتظار_الدفع": payment_filtered,
            "05_ملغي_ومسترجع": cancelled_filtered,
            "06_تم_الانتهاء": completed_filtered,
            "07_الطلبات_القديمة": old_orders_for_export,
            "08_الفواتير_القديمة": old_invoices_for_export,
            "09_إحصائيات_قديمة": stats_df
        }
        excel_data = export_to_excel(export_data)
        st.download_button(
            "📥 تحميل التقرير الشامل",
            data=excel_data,
            file_name=f"balsam_full_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            use_container_width=True,
            type="primary"
        )
        st.session_state.show_export = False
    
    st.markdown("""
    <style>
    .stTabs [data-baseweb="tab-list"] button:nth-child(1) { background-color: #4472C4; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(2) { background-color: #ED7D31; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(3) { background-color: #9B59B6; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(4) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(5) { background-color: #3498DB; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(6) { background-color: #E74C3C; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(7) { background-color: #27AE60; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(8) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(9) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button:nth-child(10) { background-color: #6c757d; color: white; border-radius: 10px 10px 0 0; }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="true"] { transform: translateY(-2px) !important; box-shadow: 0 4px 8px rgba(0,0,0,0.2) !important; }
    .stTabs [data-baseweb="tab-list"] button[aria-selected="false"] { opacity: 0.85 !important; }
    .stTabs [data-baseweb="tab-list"] button:hover { transform: translateY(-2px) !important; opacity: 1 !important; }
    </style>
    """, unsafe_allow_html=True)
    
    # ========== التبويبات المدمجة الجديدة ==========
    tab_additions, tab_returns, tab_conflicts, tab_post_cutoff, tab_payment, tab_cancelled, tab_completed, tab_old_orders, tab_old_invoices, tab_old_stats = st.tabs([
        f"📥 الإضافات والطلبات المفقودة ({completed_additions_merged}/{total_additions_merged})" if total_additions_merged > 0 else "📥 الإضافات والطلبات المفقودة (0)",
        f"📤 الإرجاعات والفواتير المعلقة ({completed_returns_merged}/{total_returns_merged})" if total_returns_merged > 0 else "📤 الإرجاعات والفواتير المعلقة (0)",
        f"📊 فواتير معلقة بين الفروع ({completed_conflicts}/{total_conflicts})" if total_conflicts > 0 else f"📊 فواتير معلقة بين الفروع ({total_conflicts})", # التبويب المضاف لـ الإدارة
        f"⏰ فواتير بعد آخر طلب ({completed_post_cutoff}/{total_post_cutoff})" if total_post_cutoff > 0 else f"⏰ فواتير بعد آخر طلب ({total_post_cutoff})",
        f"💰 بانتظار الدفع ({total_payment})",
        f"⚠️ ملغي/مسترجع ({total_cancelled})",
        f"✅ تم الانتهاء ({total_completed})",
        f"📅 طلبات قديمة ({len(old_orders_filtered)})",
        f"🧾 فواتير قديمة ({len(old_invoices_filtered)})",
        "📊 إحصائيات قديمة"
    ])
    
       
    with tab_additions:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #0f4c5c10, #1f7a8c10); padding: 0.75rem; border-radius: 12px; margin-bottom: 1rem;">
            <p style="margin: 0; font-weight: 500; font-size:0.9rem;">
                <strong>📥 قسم الإضافات المدمج:</strong> تم تصفية العداد والجدول بالكامل لحظر الحالات غير النشطة.
            </p>
        </div>
        """, unsafe_allow_html=True)
    
        if not additions_merged_df.empty:
            additions_merged_df['case_label'] = additions_merged_df['case_type'].map({
                'addition': '➕ إضافة مخزنية عادية',
                'orphan_salla': '🛒 سلة: طلب بدون فاتورة'
            }).fillna(additions_merged_df['case_label'])
        
            render_table_with_click(additions_merged_df, "additions_merged", allow_move=True)
        else:
            st.success("🎉 لا توجد طلبات إضافات أو نواقص قائمة حالياً.")
    
    with tab_returns:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #dc354510, #e74c3c10); padding: 0.75rem; border-radius: 12px; margin-bottom: 1rem;">
            <p style="margin: 0; font-weight: 500;">📤 <strong>الفواتير التي تحتاج إلى إرجاع أو معالجة نقص الطلبات</strong><br>
            <span style="font-size: 0.85rem;">🔴 الإرجاعات العادية: كمية الفاتورة أعلى من الطلب | 🟡 الفواتير بدون طلب: فاتورة موجودة في ABC وغير موجودة في سلة</span>
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        if not returns_merged_df.empty:
            col1, col2, col3 = st.columns(3)
            with col1:
                normal_returns = len(returns_merged_df[returns_merged_df['case_type'] == 'return'])
                st.metric("🔄 الإرجاعات العادية", normal_returns)
            with col2:
                orphan_abc_count = len(returns_merged_df[returns_merged_df['case_type'] == 'orphan_abc'])
                st.metric("📄 الفواتير بدون طلب", orphan_abc_count)
            with col3:
                st.metric("✅ المنجز", f"{completed_returns_merged}/{total_returns_merged}")
            st.markdown("---")
            render_table_with_click(returns_merged_df, "returns_merged", allow_move=True)
        else:
            st.success("🎉 لا توجد طلبات إرجاعات أو فواتير بدون طلب حالياً.")

    with tab_conflicts:
        st.markdown(f"### 📊 فواتير معلقة بسبب التداخل والتكرار بين الفروع")
        # تم تغيير المتغير إلى conflicts_filtered لإنهاء الخطأ وعرض البيانات المدمجة
        render_table_with_click(conflicts_filtered, "branch_conflict", allow_move=True)
        
    with tab_post_cutoff:
        render_table_with_click(post_cutoff_filtered, "post_cutoff", allow_move=True)
    
    with tab_payment:
        render_table_with_click(payment_filtered, "payment", allow_move=False)
    
    with tab_cancelled:
        render_table_with_click(cancelled_filtered, "cancelled", allow_move=False)
    
    with tab_completed:
        render_table_with_click(completed_filtered, "completed", allow_move=False)
    
    with tab_old_orders:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #1a1a1a, #333); padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
            <h3 style="color: white; margin: 0;">📅 الطلبات القديمة (أكثر من 6 أشهر)</h3>
            <p style="color: #ccc; margin: 0.5rem 0 0 0;">⚠️ الطلبات التي مر عليها أكثر من 6 أشهر والموجودة في سلة ولكن لم تكتمل</p>
        </div>
        """, unsafe_allow_html=True)
        
        if selected_branch != "الكل":
            st.info(f"🏥 عرض الطلبات القديمة للفرع: {selected_branch}")
        
        months_orders = st.slider("عدد الأشهر للبحث (طلبات)", min_value=3, max_value=24, value=6, step=3, key="old_orders_months")
        old_orders_filtered_dynamic = get_old_orders(pharmacy_name=selected_branch_name, months=months_orders)
        
        if not old_orders_filtered_dynamic.empty:
            st.warning(f"⚠️ يوجد {len(old_orders_filtered_dynamic)} طلب قديم (أكثر من {months_orders} أشهر)")
            render_table_with_click(old_orders_filtered_dynamic, "old_orders", allow_move=True)
            
            if st.button("📥 تصدير الطلبات القديمة إلى Excel", use_container_width=True, key="export_old_orders"):
                excel_data = export_to_excel({"الطلبات_القديمة": old_orders_filtered_dynamic})
                st.download_button("📥 تحميل التقرير", data=excel_data, 
                                  file_name=f"old_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
                                  use_container_width=True)
        else:
            st.success(f"🎉 لا توجد طلبات قديمة (أكثر من {months_orders} أشهر)")
    
    with tab_old_invoices:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #1a1a1a, #333); padding: 1rem; border-radius: 10px; margin-bottom: 1rem;">
            <h3 style="color: white; margin: 0;">🧾 الفواتير القديمة (أكثر من 6 أشهر)</h3>
            <p style="color: #ccc; margin: 0.5rem 0 0 0;">⚠️ الفواتير التي مر عليها أكثر من 6 أشهر والموجودة في ABC ولكن لم تكتمل</p>
        </div>
        """, unsafe_allow_html=True)
        
        if selected_branch != "الكل":
            st.info(f"🏥 عرض الفواتير القديمة للفرع: {selected_branch}")
        
        months_invoices = st.slider("عدد الأشهر للبحث (فواتير)", min_value=3, max_value=24, value=6, step=3, key="old_invoices_months")
        old_invoices_filtered_dynamic = get_old_invoices(pharmacy_name=selected_branch_name, months=months_invoices)
        
        if not old_invoices_filtered_dynamic.empty:
            st.warning(f"⚠️ يوجد {len(old_invoices_filtered_dynamic)} فاتورة قديمة (أكثر من {months_invoices} أشهر)")
            render_table_with_click(old_invoices_filtered_dynamic, "old_invoices", allow_move=True)
            
            if st.button("📥 تصدير الفواتير القديمة إلى Excel", use_container_width=True, key="export_old_invoices"):
                excel_data = export_to_excel({"الفواتير_القديمة": old_invoices_filtered_dynamic})
                st.download_button("📥 تحميل التقرير", data=excel_data, 
                                  file_name=f"old_invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", 
                                  use_container_width=True)
        else:
            st.success(f"🎉 لا توجد فواتير قديمة (أكثر من {months_invoices} أشهر)")
    
    with tab_old_stats:
        st.markdown("### 📊 إحصائيات العناصر القديمة")
        st.markdown("---")
        
        if selected_branch != "الكل":
            st.info(f"🏥 عرض الإحصائيات للفرع: {selected_branch}")
        
        old_orders_stats_data = get_old_orders(pharmacy_name=selected_branch_name, months=6)
        old_invoices_stats_data = get_old_invoices(pharmacy_name=selected_branch_name, months=6)
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 📅 الطلبات القديمة")
            st.markdown(f"""
            - **إجمالي الطلبات القديمة:** {len(old_orders_stats_data)}
            - **إضافات قديمة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "addition"])}
            - **إرجاعات قديمة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "return"])}
            - **طلبات بدون فاتورة:** {len(old_orders_stats_data[old_orders_stats_data["case_type"] == "orphan_salla"])}
            """)
            
            if selected_branch == "الكل" and not old_orders_stats_data.empty:
                st.markdown("#### 🏥 التوزيع حسب الفرع")
                for branch, count in old_orders_stats_data.groupby("pharmacy_name").size().items():
                    st.markdown(f"- {branch}: {count} طلب")
        
        with col2:
            st.markdown("#### 🧾 الفواتير القديمة")
            st.markdown(f"""
            - **إجمالي الفواتير القديمة:** {len(old_invoices_stats_data)}
            - **إضافات:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "addition"])}
            - **إرجاعات:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "return"])}
            - **فواتير بدون طلب:** {len(old_invoices_stats_data[old_invoices_stats_data["case_type"] == "orphan_abc"])}
            """)
            
            if selected_branch == "الكل" and not old_invoices_stats_data.empty:
                st.markdown("#### 🏥 التوزيع حسب الفرع")
                for branch, count in old_invoices_stats_data.groupby("pharmacy_name").size().items():
                    st.markdown(f"- {branch}: {count} فاتورة")
        
        st.markdown("---")
        total_old = len(old_orders_stats_data) + len(old_invoices_stats_data)
        if total_old > 0:
            st.warning(f"⚠️ إجمالي العناصر القديمة (طلبات + فواتير): {total_old}")
            st.info("💡 هذه العناصر تم استبعادها تلقائياً من التبويبات الأخرى (الإضافات، الإرجاعات، إلخ)")
        else:
            st.success("🎉 لا توجد عناصر قديمة (طلبات أو فواتير)")
    
    st.markdown('<div class="section-title">👥 آخر دخول للصيدليات</div>', unsafe_allow_html=True)
    last_logins = get_all_last_logins()
    if not last_logins.empty:
        cols = st.columns(4)
        for idx, (_, row) in enumerate(last_logins.head(8).iterrows()):
            with cols[idx % 4]:
                st.markdown(f"""
                <div class="note-card">
                    <strong>🏥 {row['pharmacy_name'][-10:]}</strong><br>
                    <span>👤 {row['pharmacist_name'] or 'غير مسجل'}</span><br>
                    <span>📅 {row['last_login'][:16] if row['last_login'] else 'لم يدخل'}</span><br>
                    <span>🌐 IP: {row['last_ip'] or 'غير معروف'}</span>
                </div>
                """, unsafe_allow_html=True)
        
        with st.expander("📋 عرض جميع الصيدليات"):
            st.dataframe(last_logins[['pharmacy_name', 'pharmacist_name', 'last_login', 'last_ip']], use_container_width=True)
    else:
        st.info("لا توجد سجلات دخول للصيدليات بعد")

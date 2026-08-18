import streamlit as st
import pandas as pd
import sqlite3
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from utils.database import (
    fetch_active_items, get_completed_items, get_tab_completed_counts, 
    get_old_orders, get_old_invoices, get_old_invoices_stats,
    save_case_note, mark_case_done, get_setting, DB_PATH
)
from utils.helpers import (
    is_cancelled_or_returned_status, is_pending_payment_status, 
    get_branch_number, get_branch_location, get_tab_label, numeric_value,
    get_saudi_time
)
from utils.ui_components import render_metrics, render_completed_table

def to_safe_int(val):
    if pd.isna(val) or str(val).strip() in ["", "nan", "None"]:
        return 0
    try:
        return int(float(str(val).strip()))
    except:
        return 0
        
def export_to_excel(dataframes_dict: dict, pharmacy_name: str) -> bytes:
    output = BytesIO()
    
    tab_colors = {
        "الاضافات والطلبات المفقودة": "4472C4", "الارجاعات والزيادات": "ED7D31",
        "فواتير معلقة بين الفروع": "9B59B6", "فواتير بعد اخر طلب": "9B59B6",
        "بانتظار الدفع": "3498DB", "الملغيات والمسترجعات": "E74C3C", "تم الانتهاء": "2ECC71"
    }
    
    columns_mapping = {
        'order_date': 'تاريخ الطلب', 'order_number': 'رقم الطلب', 'invoice_date': 'تاريخ الفاتورة',
        'invoice_number': 'رقم الفاتورة', 'customer_phone': 'رقم جوال العميل', 'sku': 'رقم المنتج SKU',
        'product_name': 'اسم المنتج', 'salla_qty': 'كمية سلة', 'abc_qty': 'كمية ABC',
        'difference': 'الفرق', 'order_status': 'حالة الطلب', 'profile_type': 'نوع البروفايل',
        'abc_pharmacist_name': 'الصيدلي المسؤول', 'pharmacist_note': 'ملاحظات الصيدلية',
        'status': 'حالة التسوية', 'case_type': 'نوع الحالة', 'case_reason': 'سبب الحالة',
        'city': 'المدينة', 'item_key': 'المفتاح الشامل للصنف'
    }

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            if df is not None and not df.empty:
                available_cols = [col for col in columns_mapping.keys() if col in df.columns]
                df_filtered = df[available_cols].copy()
                df_filtered = df_filtered.rename(columns=columns_mapping)
                df_filtered.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                worksheet = writer.sheets[sheet_name[:31]]
                header_fill = PatternFill(start_color=tab_colors.get(sheet_name, "2A5298"), fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col in range(1, len(df_filtered.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    worksheet.column_dimensions[get_column_letter(col)].width = 22
            else:
                pd.DataFrame({"ملاحظة": ["لا توجد بيانات لهذا التبويب"]}).to_excel(writer, sheet_name=sheet_name[:31], index=False)
    output.seek(0)
    return output.getvalue()

def export_to_excel_brief(dataframes_dict: dict) -> bytes:
    output = BytesIO()
    allowed_tabs = ["الاضافات والطلبات المفقودة", "الارجاعات والزيادات", "فواتير معلقة بين الفروع", "بانتظار الدفع", "تم الانتهاء"]
    
    tab_colors = {
        "الاضافات والطلبات المفقودة": "4472C4", "الارجاعات والزيادات": "ED7D31",
        "فواتير معلقة بين الفروع": "9B59B6", "بانتظار الدفع": "3498DB", "تم الانتهاء": "2ECC71"
    }
    
    # 💡 إضافة الأعمدة الجديدة المطلوبة للملف المختصر
    target_columns = {
        'order_date': 'تاريخ الطلب', 'invoice_date': 'تاريخ الفاتورة', 'order_number': 'رقم الطلب',
        'invoice_number': 'رقم الفاتورة', 'customer_phone': 'رقم جوال العميل', 'sku': 'رقم المنتج',
        'product_name': 'اسم المنتج', 'salla_qty': 'كمية سلة', 'abc_qty': 'كمية abc',
        'diff_qty': 'الفرق', 'order_status': 'حالة الطلب', 'abc_pharmacist_name': 'اسم الصيدلي',
        'profile_type': 'نوع البروفايل', 'status': 'حالة التسوية',
        'performed_by': 'تم التنفيذ بواسطة', 'performed_at': 'وقت التنفيذ'
    }
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name in allowed_tabs:
            df = dataframes_dict.get(sheet_name)
            if df is not None and not df.empty:
                df_brief = df.copy()
                salla_q = pd.to_numeric(df_brief.get('salla_qty', 0), errors='coerce').fillna(0).astype(int)
                abc_q = pd.to_numeric(df_brief.get('abc_qty', 0), errors='coerce').fillna(0).astype(int)
                df_brief['diff_qty'] = salla_q - abc_q
                
                for col in target_columns.keys():
                    if col not in df_brief.columns: df_brief[col] = "N/A"
                        
                df_brief = df_brief[list(target_columns.keys())].rename(columns=target_columns)
                df_brief.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                worksheet = writer.sheets[sheet_name[:31]]
                header_fill = PatternFill(start_color=tab_colors.get(sheet_name, "2A5298"), fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col in range(1, len(df_brief.columns) + 1):
                    worksheet.cell(row=1, column=col).fill = header_fill
                    worksheet.cell(row=1, column=col).font = header_font
                    worksheet.column_dimensions[get_column_letter(col)].width = 20
            else:
                pd.DataFrame({"ملاحظة": ["لا توجد بيانات معلقة حالياً"]}).to_excel(writer, sheet_name=sheet_name[:31], index=False)
                
    output.seek(0)
    return output.getvalue()

def render_single_case_card(row, idx, allow_actions, pharmacist_name, pharmacy_name, tab_id=""):
    salla_numeric = int(row.get('salla_qty', 0)) if pd.notna(row.get('salla_qty', 0)) else 0
    abc_numeric = int(row.get('abc_qty', 0)) if pd.notna(row.get('abc_qty', 0)) else 0
    diff_value = salla_numeric - abc_numeric
    case_type = row.get('case_type', '')

    if case_type == 'branch_conflict':
        badge_text, badge_color, badge_text_color, diff_style, required_action = "⚠️ تداخل معلق بين الفروع", "#f8d7da", "#721c24", "color: #dc3545; font-weight: bold;", "<span style='color: #dc3545; font-weight: bold;'>مراجعة وتأكيد</span>"
    elif diff_value > 0:
        badge_text, badge_color, badge_text_color, diff_style, required_action = "إضافة مخزنية عادية ➕", "#dff1ff", "#084298", "color: #28a745; font-weight: bold;", "<span style='color: #28a745; font-weight: bold;'>إضافة</span>"
    elif diff_value < 0:
        badge_text, badge_color, badge_text_color, diff_style, required_action = "إرجاع مخزني عادي 🔄", "#ffe0df", "#491217", "color: #dc3545; font-weight: bold;", "<span style='color: #dc3545; font-weight: bold;'>إرجاع</span>"
    else:
        diff_style, required_action = "color: #6c757d; font-weight: bold;", "<span style='color: #6c757d; font-weight: bold;'>مطابق</span>"
        if case_type == 'orphan_salla': badge_text, badge_color, badge_text_color = "طلب مبيعات مفقود الفاتورة 🛒", "#fff3cd", "#856404"
        elif case_type == 'orphan_abc': badge_text, badge_color, badge_text_color = "فاتورة توريد مفقودة الطلب 📄", "#f8d7da", "#721c24"
        else: badge_text, badge_color, badge_text_color = "حالة تسوية عامة", "#e2e8f0", "#475569"
   
    order_status = row.get('order_status', 'غير متوفرة')
    invoice_date = row.get('invoice_date', '')
    order_date = row.get('order_date', '')
    order_number = str(row.get('order_number', '')).strip()
    sku = str(row.get('sku', '')).strip()
    
    exact_duplicates = []  
    shared_order_duplicates = []  

    try: 
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT pharmacy_name, invoice_number, status 
            FROM reconciliation_items 
            WHERE TRIM(order_number) = ? 
              AND TRIM(sku) = ? 
              AND TRIM(pharmacy_name) != ?
        """, (order_number, sku, pharmacy_name.strip()))
        
        for r in cursor.fetchall():
            exact_duplicates.append({
                "pharmacy": r["pharmacy_name"],
                "invoice_number": r["invoice_number"] or "بدون فاتورة",
                "status": r["status"]
            })
        
        cursor.execute("""
            SELECT DISTINCT pharmacy_name, sku, product_name, status 
            FROM reconciliation_items 
            WHERE TRIM(order_number) = ? 
              AND TRIM(pharmacy_name) != ? 
              AND TRIM(sku) != ?
        """, (order_number, pharmacy_name.strip(), sku))
        
        for r in cursor.fetchall():
            shared_order_duplicates.append({
                "pharmacy": r["pharmacy_name"],
                "sku": r["sku"],
                "product_name": r["product_name"] or "غير محدد",
                "status": r["status"]
            })
        conn.close()
    except Exception as e:
        pass
    
    with st.container():
        st.markdown(f"""
        <div style="background:#f8f9fa; border-radius:16px; padding:1rem; margin-bottom:1rem; border-right:5px solid #1f7a8c; box-shadow:0 2px 8px rgba(0,0,0,0.05);">
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:0.75rem; flex-wrap:wrap;">
                <span style="background:{badge_color}; color:{badge_text_color}; padding:0.25rem 0.75rem; border-radius:20px; font-size:0.8rem; font-weight:bold;">{badge_text}</span>
                <div>
                    <span style="color:#6c757d; font-size:0.75rem;">📅 الطلب: {order_date[:16] if order_date else 'غير محدد'}</span>
                    <span style="color:#6c757d; font-size:0.75rem; margin-right:0.5rem;">📅 الفاتورة: {invoice_date[:16] if invoice_date else 'غير محدد'}</span>
                </div>
            </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([2, 1, 1.5])
        with col1:
            st.markdown(f"- **📋 رقم الطلب:** {row.get('order_number', 'N/A')}\n- **🏷️ SKU / رقم المنتج:** {row.get('sku', 'N/A')}\n- **📦 المنتج:** {str(row.get('product_name', 'N/A'))[:60]}")
            if case_type in ['addition', 'orphan_salla', 'branch_conflict']:
                phone = row.get('customer_phone', 'N/A')
                if pd.notna(phone) and str(phone).strip() not in ["", "nan", "None"]: st.markdown(f"- **📱 جوال العميل:** `{phone}`")
        with col2:
            st.markdown(f"- **🛒 كمية سلة:** {salla_numeric}\n- **📄 كمية ABC:** {abc_numeric}\n- **📊 الفرق:** <span style='{diff_style}'>{'+' if diff_value > 0 else ''}{diff_value}</span>\n- **🎯 المطلوب:** {required_action}", unsafe_allow_html=True)
        with col3:
            st.markdown(f"- **🧾 رقم الفاتورة:** {row.get('invoice_number', 'N/A')}\n- **👤 الصيدلي:** {row.get('abc_pharmacist_name', 'غير معروف')}\n- **⚙️ حالة الطلب:** {order_status}")
            if case_type in ['return', 'orphan_abc', 'branch_conflict']:
                profile = row.get('profile_type', 'N/A')
                if pd.notna(profile) and str(profile).strip() not in ["", "nan", "None"]: st.markdown(f"- **📄 نوع البروفايل:** `{profile}`")
            
        if exact_duplicates:
            dup_exact_html = '<div style="background:#f8d7da; border-right:4px solid #dc3545; padding:0.75rem; margin-top:0.75rem; border-radius:10px; margin-bottom:0.5rem;"><span style="color:#721c24; font-weight:bold;">🚨 تنبيه تكرار الصنف الشامل: هذا الصنف مسجل في الفروع التالية لنفس الطلب:</span>'
            for ed in exact_duplicates:
                status_lbl = "✅ تمت إضافتها واعتمادها" if ed.get("status") == "تم" else "⏳ معلقة لم تُعتمد بعد"
                dup_exact_html += f'<div style="font-size:0.85rem; color:#491217; margin-top:4px;">🏥 <strong>{ed.get("pharmacy")}</strong> | فاتورة: {ed.get("invoice_number")} | الإجراء بالفرع الآخر: <span style="font-weight:bold;">{status_lbl}</span></div>'
            st.markdown(dup_exact_html + '</div>', unsafe_allow_html=True)
            
        if shared_order_duplicates:
            dup_warning_html = '<div style="background:#fff3cd; border-right:4px solid #ff9800; padding:0.75rem; margin-top:0.5rem; border-radius:10px; margin-bottom:0.5rem;"><span style="color:#856404; font-weight:bold;">⚠️ تنبيه الطلب المشترك: يوجد فروع أخرى أصدرت فواتير لنفس رقم هذا الطلب لأصناف أخرى:</span>'
            for dup in shared_order_duplicates: 
                status_lbl = "✅ تمت تسويتها واكتمالها" if dup.get("status") == "تم" else "⏳ معلقة"
                dup_warning_html += f'<div style="font-size:0.85rem; color:#66521a; margin-top:4px;">🏥 <strong>{dup.get("pharmacy")}</strong> | الصنف الآخر: {dup.get("sku")} ({str(dup.get("product_name"))[:35]}...) | الحالة: [{status_lbl}]</div>'
            st.markdown(dup_warning_html + '</div>', unsafe_allow_html=True)
            
        if row.get("status") == "تم":
            action_word = "الاضافة" if case_type in ['addition', 'orphan_salla'] else ("الارجاع" if case_type in ['return', 'orphan_abc'] else "التسوية")
            performed_user = row.get('performed_by', pharmacist_name)
            if not performed_user: performed_user = pharmacist_name
            
            if row.get("pharmacist_note"):
                st.markdown(f"<div style='margin-top:10px; color:#495057;'><strong>📝 ملحوظة الصيدلي:</strong> {row.get('pharmacist_note')}</div>", unsafe_allow_html=True)
                
            st.markdown(f"""
            <div style="background-color: #d1e7dd; color: #0f5132; padding: 12px; border-radius: 8px; text-align: center; border: 1px solid #badbcc; font-weight: bold; margin-top: 15px;">
                تمت {action_word} بنجاح عن طريق {performed_user} ✅
            </div>
            """, unsafe_allow_html=True)
            
        else:
            note_key = f"note_{tab_id}_{case_type}_{idx}_{row.get('order_number', '')}_{row.get('sku', '')}"
            note_value = st.text_area("📝 ملحوظة الصيدلي", value=row.get("pharmacist_note", "") or "", key=note_key, height=60)
            
            btn_col1, btn_col2, btn_col3 = st.columns([1, 1.5, 1.5])
            with btn_col1:
                if st.button("💾 حفظ الملحوظة", key=f"save_{note_key}", use_container_width=True):
                    # 💡 تمرير اسم وصلاحية الصيدلي بشكل صريح
                    save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, note_value, pharmacist_name, st.session_state.user_role)
                    st.toast("📋 تم حفظ الملاحظة بنجاح!", icon="💾")
                    st.rerun()
                    
            if allow_actions:
                if case_type == "branch_conflict":
                    with btn_col2:
                        if st.button("📥 تأكيد الإضافة (تم البيع من فرعي)", key=f"conf_add_{note_key}", use_container_width=True):
                            save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, f"[فرع صحيح] | {note_value}", pharmacist_name, st.session_state.user_role)
                            mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                            st.toast("✅ تم الاعتماد!"); st.rerun()
                    with btn_col3:
                        if st.button("🔄 تأكيد الإرجاع (ليس من فرعي)", key=f"conf_ret_{note_key}", use_container_width=True):
                            save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, f"[فرع مخطئ] | {note_value}", pharmacist_name, st.session_state.user_role)
                            mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                            st.toast("🔄 تم العكس!"); st.rerun()
                elif case_type in {"addition", "orphan_salla", "return", "orphan_abc"}:
                    button_label = "✅ تأكيد الإضافة" if case_type in {"addition", "orphan_salla"} else "🔄 تأكيد الإرجاع"
                    with btn_col2:
                        if st.button(button_label, key=f"done_{note_key}", use_container_width=True):
                            save_case_note(row['order_number'], row['sku'], pharmacy_name, case_type, note_value, pharmacist_name, st.session_state.user_role)
                            mark_case_done(row['order_number'], row['sku'], pharmacy_name, case_type, pharmacist_name)
                            st.toast("🚀 تم التأكيد!"); st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

def render_case_cards_pharmacy(df, allow_actions, pharmacist_name, pharmacy_name, tab_id=""):
    if df is not None and not df.empty:
        for i, (orig_idx, row) in enumerate(df.iterrows()):
            render_single_case_card(row, i, allow_actions, pharmacist_name, pharmacy_name, tab_id=tab_id)
    else:
        st.success("🎉 ممتاز! التبويب الحالي متطابق بالكامل ولا توجد أي فروقات مخزنية معلقة هنا.")

def show():
    pharmacy_name = st.session_state.username
    pharmacist_name = st.session_state.get("pharmacist_name", "") or ""
    branch_number = get_branch_number(pharmacy_name)
    branch_location = get_branch_location(branch_number)

    st.markdown(f"""
    <div class="hero">
        <h1>🏥 {pharmacy_name}</h1>
        <p>فرع رقم {branch_number} | الموقع: {branch_location} | الصيدلي: {pharmacist_name}</p>
        <p>🕐 آخر تحديث: {get_saudi_time()}</p>
    </div>
    """, unsafe_allow_html=True)

    conn = sqlite3.connect(DB_PATH)
    history_df = pd.read_sql_query(
        "SELECT login_time as 'وقت الدخول', user_agent as 'اسم الصيدلي', ip_address as 'IP الجهاز' FROM login_history WHERE username = ? ORDER BY login_time DESC LIMIT 5", 
        conn, params=(pharmacy_name,)
    )
    conn.close()

    st.markdown("### 🕒 سجل الدخول الخاص بالصيدلية")
    st.dataframe(history_df, use_container_width=True)

    col1, col2, col3 = st.columns([1, 1.2, 1.5])
    with col1:
        if st.button("🔄 تحديث الصفحة", use_container_width=True): st.rerun()
    with col2:
        if st.button("📥 تصدير الملف الكامل Excel", use_container_width=True): st.session_state.show_export_pharmacy = True
    with col3:
        if st.button("📋 تصدير ملف مختصر Excel", use_container_width=True): st.session_state.show_export_brief_pharmacy = True

    df = fetch_active_items(pharmacy_name, include_hidden=False)
    old_invoices_df = get_old_invoices(pharmacy_name=pharmacy_name, months=6)
    old_orders_df = get_old_orders(pharmacy_name=pharmacy_name, months=6)

    if df.empty:
        st.info("📭 لا توجد حالات نشطة لهذا الفرع حاليًا.")
        return

    st.markdown("### 🔍 بحث وتصفية")
    search_col1, search_col2, search_col3 = st.columns(3)
    search_order = search_col1.text_input("🔢 بحث برقم الطلب")
    search_invoice = search_col2.text_input("🧾 بحث برقم الفاتورة")
    search_sku = search_col3.text_input("🏷️ بحث بـ SKU")

    if search_order:
        df = df[df["order_number"].astype(str).str.contains(search_order, na=False, case=False)]
    if search_invoice:
        df = df[df["invoice_number"].astype(str).str.contains(search_invoice, na=False, case=False)]
    if search_sku:
        df = df[df["sku"].astype(str).str.contains(search_sku, na=False, case=False)]

    is_locked = False
    if not df.empty and 'is_item_locked' in df.columns:
        is_locked = df['is_item_locked'].iloc[0] == 1
        
    allow_actions = not is_locked

    active_mask = ~df["order_status"].apply(is_cancelled_or_returned_status)
    payment_mask = df["order_status"].apply(is_pending_payment_status)
    active_df = df[active_mask].copy()

    branch_add_df = df[df['case_type'].isin(['addition', 'orphan_salla']) & active_mask & ~payment_mask].copy()
    total_additions_merged = len(branch_add_df)
    completed_additions_merged = len(branch_add_df[branch_add_df["status"] == "تم"])
    
    branch_ret_df = df[df['case_type'].isin(['return', 'orphan_abc']) & active_mask].copy()
    total_returns_merged = len(branch_ret_df)
    completed_returns_merged = len(branch_ret_df[branch_ret_df["status"] == "تم"])
    
    conflicts_df = df[df["case_type"] == "branch_conflict"].copy()
    total_conflicts = len(conflicts_df)
    completed_conflicts = len(conflicts_df[conflicts_df["status"] == "تم"])

    post_cutoff_df = df[(df["case_type"] == "post_cutoff_abc") & active_mask].copy()
    total_post_cutoff = len(post_cutoff_df)
    completed_post_cutoff = len(post_cutoff_df[post_cutoff_df["status"] == "تم"])
    
    payment_df = df[df["order_status"].apply(is_pending_payment_status) & (df["case_type"] != "branch_conflict") & active_mask].copy()
    total_payment = len(payment_df)
    
    cancelled_df = df[df["order_status"].apply(is_cancelled_or_returned_status) & (df["case_type"] != "branch_conflict")].copy()
    total_cancelled = len(cancelled_df)
    
    completed_df = get_completed_items(pharmacy_name)
    if search_order:
        completed_df = completed_df[completed_df["order_number"].astype(str).str.contains(search_order, na=False, case=False)]
    if search_invoice:
        completed_df = completed_df[completed_df["invoice_number"].astype(str).str.contains(search_invoice, na=False, case=False)]
    if search_sku:
        completed_df = completed_df[completed_df["sku"].astype(str).str.contains(search_sku, na=False, case=False)]
        
    total_completed = len(completed_df)

    if st.session_state.get('show_export_pharmacy', False):
        excel_sheets = {
            "الاضافات والطلبات المفقودة": branch_add_df,
            "الارجاعات والزيادات": branch_ret_df,
            "فواتير معلقة بين الفروع": conflicts_df,
            "فواتير بعد اخر طلب": post_cutoff_df,
            "بانتظار الدفع": payment_df,
            "الملغيات والمسترجعات": cancelled_df,
            "تم الانتهاء": completed_df
        }
        excel_data = export_to_excel(excel_sheets, pharmacy_name)
        st.download_button(label="💾 تحميل ملف Excel الموحد", data=excel_data, file_name=f"Full_Report_{pharmacy_name}.xlsx", use_container_width=True, type="primary")
        st.session_state.show_export_pharmacy = False

    if st.session_state.get('show_export_brief_pharmacy', False):
        excel_sheets_brief = {
            "الاضافات والطلبات المفقودة": branch_add_df,
            "الارجاعات والزيادات": branch_ret_df,
            "فواتير معلقة بين الفروع": conflicts_df,
            "بانتظار الدفع": payment_df,
            "تم الانتهاء": completed_df
        }
        excel_data_brief = export_to_excel_brief(excel_sheets_brief)
        st.download_button(label="📊 تحميل ملف Excel المختصر", data=excel_data_brief, file_name=f"Brief_Report_{pharmacy_name}.xlsx", use_container_width=True, type="primary")
        st.session_state.show_export_brief_pharmacy = False

    st.markdown("""
    <div style="background-color: #ffeb3b; padding: 15px; border-radius: 10px; text-align: center; margin-bottom: 20px; border: 2px solid #fbc02d;">
        <h2 style="color: #d32f2f; margin: 0; font-weight: bold;">رجاء الإرجاع أولاً لتوفير الرصيد اللازم .. ثم البدء في الإضافات</h2>
    </div>
    """, unsafe_allow_html=True)
    
    # 💡 بناء قائمة التبويبات المتاحة بناءً على إعدادات الإدارة
    available_tabs = []
    
    if get_setting("show_tab_additions", "1") == "1":
        available_tabs.append(("additions", f"📥 الإضافات والطلبات ({completed_additions_merged}/{total_additions_merged})"))
        
    if get_setting("show_tab_returns", "1") == "1":
        available_tabs.append(("returns", f"📤 الإرجاعات والزيادات ({completed_returns_merged}/{total_returns_merged})"))
        
    if get_setting("show_tab_conflicts", "1") == "1":
        available_tabs.append(("conflicts", f"📊 فواتير معلقة بين الفروع ({completed_conflicts}/{total_conflicts})"))
        
    if get_setting("show_tab_post_cutoff", "1") == "1":
        available_tabs.append(("post_cutoff", f"⏰ فواتير بعد آخر طلب ({completed_post_cutoff}/{total_post_cutoff})"))
        
    if get_setting("show_tab_payment", "1") == "1":
        available_tabs.append(("payment", f"💰 بانتظار الدفع ({total_payment})"))
        
    if get_setting("show_tab_cancelled", "1") == "1":
        available_tabs.append(("cancelled", f"⚠️ ملغي/مسترجع ({total_cancelled})"))
        
    if get_setting("show_tab_completed", "1") == "1":
        available_tabs.append(("completed", f"✅ تم الانتهاء ({total_completed})"))

    # إنشاء وعرض التبويبات المتاحة فقط
    if available_tabs:
        tabs = st.tabs([t[1] for t in available_tabs])
        for idx, (tab_id, tab_title) in enumerate(available_tabs):
            with tabs[idx]:
                if tab_id == "additions":
                    render_case_cards_pharmacy(branch_add_df, allow_actions, pharmacist_name, pharmacy_name, tab_id="add")
                elif tab_id == "returns":
                    render_case_cards_pharmacy(branch_ret_df, allow_actions, pharmacist_name, pharmacy_name, tab_id="ret")
                elif tab_id == "conflicts":
                    render_case_cards_pharmacy(conflicts_df, allow_actions, pharmacist_name, pharmacy_name, tab_id="conf")
                elif tab_id == "post_cutoff":
                    render_case_cards_pharmacy(post_cutoff_df, False, pharmacist_name, pharmacy_name, tab_id="cutoff")
                elif tab_id == "payment":
                    render_case_cards_pharmacy(payment_df, False, pharmacist_name, pharmacy_name, tab_id="pay")
                elif tab_id == "cancelled":
                    render_case_cards_pharmacy(cancelled_df, False, pharmacist_name, pharmacy_name, tab_id="cancel")
                elif tab_id == "completed":
                    render_completed_table(completed_df, is_admin=False)
    else:
        st.warning("⚠️ لا توجد تبويبات متاحة حالياً. تم إخفاء جميع التبويبات من قبل الإدارة.")

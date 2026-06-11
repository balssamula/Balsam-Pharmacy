import re
from io import BytesIO
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import streamlit as st


def normalize_text(text):
    """تنظيف وتوحيد النص ليتعامل مع الأخطاء الإملائية، الاختصارات، والكلمات التسويقية الزائدة"""
    text = str(text).lower().strip()

    # إزالة الكلمات التسويقية الزائدة التي قد تشوش على المعادلات
    words_to_remove = ["حصري", "اونلاين", "اون لاين", "فقط", "عرض", "مميز"]
    for word in words_to_remove:
        text = text.replace(word, "")

    # توحيد المسافات والحروف
    text = re.sub(r"\s+", " ", text)  # إزالة المسافات الزائدة
    text = re.sub(r"[أإآا]", "ا", text)  # توحيد الألف
    text = re.sub(r"ة\b", "ه", text)  # توحيد التاء المربوطة والهاء
    text = re.sub(r"\bع\b", "على", text)  # تحويل "ع" إلى "على"

    # توحيد الكلمات الدلالية للحبات
    text = re.sub(r"حبات?", "حبة", text)
    text = re.sub(r"الحبات?", "الحبة", text)

    return text.strip()


def create_template_excel():
    """إنشاء ملف نموذج فارغ يحتوي على الأعمدة المطلوبة بالتنسيق الاحترافي"""
    template_data = {
        "رقم المنتج (SKU)": ["1001", "1002", "1003", "1004"],
        "اسم المنتج": [
            "بندول نايت 20 قرص",
            "فيتامين سي 1000 ملجم",
            "معجون أسنان كولجيت",
            "حليب أطفال نيدو 400 جرام",
        ],
        "عنوان العرض": [
            "عرض 1+1 مجانا حصري",
            "خصم 50% ع الحبه الثانيه",
            "خصم 20% اونلاين",
            "اشتر 3 حبات بسعر حبة",
        ],
        "السعر غير شامل الضريبة": [100.00, 50.00, 150.00, 60.00],
        "الضريبة": [0.15, 0.15, 0.15, 0.15],
    }
    df_template = pd.DataFrame(template_data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_template.to_excel(writer, index=False, sheet_name="نموذج العروض")
        worksheet = writer.sheets["نموذج العروض"]

        header_fill = PatternFill(start_color="1F7A8C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Segoe UI")
        for col in range(1, len(df_template.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            worksheet.column_dimensions[get_column_letter(col)].width = 25

    output.seek(0)
    return output.getvalue()


def style_excel_professionally(df, output_bytesio):
    """تطبيق تنسيق احترافي، تثبيت الصف الأول وتفعيل الفلترة التلقائية"""
    output_styled = BytesIO()

    with pd.ExcelWriter(output_styled, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="تقرير العروض الذكي")
        worksheet = writer.sheets["تقرير العروض الذكي"]

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(
            start_color="1F7A8C", end_color="1F7A8C", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Segoe UI")
        bold_price_font = Font(
            bold=True, size=11, name="Segoe UI", color="0F4C5C"
        )
        normal_font = Font(size=11, name="Segoe UI")
        center_align = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )

        price_columns_keywords = ["سعر", "السعر", "الخصم", "قيمة"]

        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                col_name = df.columns[col_idx - 1]

                if any(
                    keyword in col_name for keyword in price_columns_keywords
                ):
                    cell.font = bold_price_font
                else:
                    cell.font = normal_font

                cell.border = thin_border
                cell.alignment = center_align

        for col in worksheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 5, 18)

    output_styled.seek(0)
    return output_styled.getvalue()


def calculate_promotions_stream(uploaded_file):
    """حساب وتفصيل العروض الترويجية الذكية بالاعتماد على معالجة النصوص المرنة"""
    df = pd.read_excel(uploaded_file)

    required_cols = [
        "رقم المنتج (SKU)",
        "اسم المنتج",
        "عنوان العرض",
        "السعر غير شامل الضريبة",
        "الضريبة",
    ]
    for col in required_cols:
        if col not in df.columns:
            raise ValueError(
                f"العمود الأساسي المطلوب '{col}' مفقود! يرجى تحميل النموذج وتعبئته."
            )

    total_qty_list = []
    discount_pct_list = []
    price_before_list = []
    discount_val_list = []
    price_after_list = []

    for idx, row in df.iterrows():
        p_base = float(row.get("السعر غير شامل الضريبة", 0))
        tax_pct = float(row.get("الضريبة", 0))
        original_promo_text = str(row["عنوان العرض"])

        # تنظيف وتحضير النص برمجياً للتعامل مع المتغيرات
        promo_text = normalize_text(original_promo_text)

        p_tax = p_base * (1 + tax_pct)
        qty, disc_pct, p_before, p_after = 1, 0.0, p_tax, p_tax
        processed = False

        # --------------------------------------------------
        # 1. نمط: اشتر X بسعر Y (يدعم: اشتر حبتين بسعر حبة / 3 حبات بسعر 2 .. إلخ)
        # --------------------------------------------------
        match_buy_pay = re.search(
            r"(?:اشتر|اشتري)\s*(\d+|حبتين|3\s*حبات|حبة)\s*(?:حبة)?\s*بسعر\s*(\d+|حبتين|حبة)",
            promo_text,
        )
        if match_buy_pay and not processed:
            buy_str = match_buy_pay.group(1)
            pay_str = match_buy_pay.group(2)

            # تحويل الكلمات النصية الشائعة إلى أرقام
            buy_qty = (
                2
                if "حبتين" in buy_str
                else (1 if "حبة" in buy_str else int(re.search(r"\d+", buy_str).group()))
            )
            pay_qty = (
                2
                if "حبتين" in pay_str
                else (1 if "حبة" in pay_str else int(re.search(r"\d+", pay_str).group()))
            )

            if buy_qty > 0:
                qty = buy_qty
                p_before = p_tax * qty
                p_after = p_tax * pay_qty
                disc_pct = (qty - pay_qty) / qty
                processed = True

        # --------------------------------------------------
        # 2. نمط: العروض الحسابية المباشرة (مثل: 1+1 مجانا، 2+1 مجانا)
        # --------------------------------------------------
        match_free = re.search(r"(\d+)\s*(?:حبة)?\s*\+\s*(\d+)\s*مجانا", promo_text)
        if match_free and not processed:
            buy_qty = int(match_free.group(1))
            free_qty = int(match_free.group(2))
            qty = buy_qty + free_qty
            p_before = p_tax * qty
            p_after = p_tax * buy_qty
            disc_pct = (free_qty / qty) if qty > 0 else 0.0
            processed = True

        # --------------------------------------------------
        # 3. نمط: خصم نسبة على الحبة الثانية (مثل: خصم 50% على الحبة الثانية / ع الثانية)
        # --------------------------------------------------
        if "ثانيه" in promo_text and "%" in promo_text and not processed:
            try:
                pct_val = (
                    float(re.search(r"خصم\s*(\d+)%", promo_text).group(1)) / 100
                )
                qty = 2
                p_before = p_tax * 2
                p_after = p_tax + (p_tax * (1 - pct_val))
                disc_pct = (
                    ((p_before - p_after) / p_before) if p_before > 0 else 0.0
                )
                processed = True
            except:
                pass

        # --------------------------------------------------
        # 4. نمط: حبات محددة بسعر ثابت (مثل: 2 حبة بسعر 99.95 ريال)
        # --------------------------------------------------
        if "بسعر" in promo_text and not processed:
            qty_match = re.search(r"(\d+)\s*حبة", promo_text)
            price_match = re.search(r"بسعر\s*([\d\.]+)", promo_text)
            if qty_match and price_match:
                qty = int(qty_match.group(1))
                p_after = float(price_match.group(1))
                p_before = p_tax * qty
                disc_pct = (
                    ((p_before - p_after) / p_before) if p_before > 0 else 0.0
                )
                processed = True

        # --------------------------------------------------
        # 5. نمط: خصم نسبة مئوية مباشرة على الصنف (مثل: خصم 20%)
        # --------------------------------------------------
        if "خصم" in promo_text and "%" in promo_text and not processed:
            try:
                pct_val = (
                    float(re.search(r"خصم\s*(\d+)%", promo_text).group(1)) / 100
                )
                qty = 1
                p_before = p_tax
                p_after = p_tax * (1 - pct_val)
                disc_pct = pct_val
                processed = True
            except:
                pass

        # حساب صافي قيمة الخصم النهائي
        disc_val = p_before - p_after

        total_qty_list.append(qty)
        discount_pct_list.append(round(disc_pct * 100, 2))
        price_before_list.append(round(p_before, 2))
        discount_val_list.append(round(disc_val, 2))
        price_after_list.append(round(p_after, 2))

    df["عدد حبات العرض"] = total_qty_list
    df["نسبة الخصم الإجمالية %"] = discount_pct_list
    df["السعر قبل العرض شامل الضريبة"] = price_before_list
    df["قيمة الخصم شامل الضريبة"] = discount_val_list
    df["السعر بعد العرض شامل الضريبة"] = price_after_list

    styled_excel_bytes = style_excel_professionally(df, None)
    return styled_excel_bytes, df


def show():
    st.markdown(
        "<div class='hero'><h1>🏷️ حاسبة ومُنسّق العروض الترويجية المطور الذكي</h1><p>تصدير تقارير محاسبية معالجة ذكياً ضد النصوص العشوائية والكلمات التسويقية الزائدة</p></div>",
        unsafe_allow_html=True,
    )

    st.markdown("### 📋 الخطوة 1: تنزيل نموذج الهيكل الجديد")
    st.info(
        "قم بتحميل الملف المفرغ أدناه والمطور، والذي يضم أعمدة (رقم المنتج SKU واسم المنتج):"
    )

    template_bytes = create_template_excel()
    st.download_button(
        label="📥 تحميل نموذج العروض المطور والمعدّل (Excel Template)",
        data=template_bytes,
        file_name="Promotions_Enhanced_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    st.markdown("---")

    st.markdown("### 📂 الخطوة 2: رفع وتحليل ملف العروض")
    uploaded_file = st.file_uploader(
        "ارفع ملف العروض المكتمل بعد تعبئته بناءً على هيكل النموذج الجديد",
        type=["xlsx"],
    )

    if uploaded_file is not None:
        with st.spinner(
            "⏳ جاري تفكيك العروض بالذكاء الموسّع وتطبيق التنسيقات الفاخرة..."
        ):
            try:
                excel_bytes, result_df = calculate_promotions_stream(
                    uploaded_file
                )
                st.success(
                    "✨ تم الحساب بنجاح وتجاوز الكلمات الزائدة وتثبيت الفلاتر!"
                )

                st.markdown("### 📊 معاينة تفاعلية سريعة للملف الجاهز للتنزيل:")
                st.dataframe(result_df.head(10), use_container_width=True)

                st.download_button(
                    label="💾 تحميل الملف الاحترافي المنسق والجاهز للإدارة (Excel)",
                    data=excel_bytes,
                    file_name="Styled_Promotions_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
            except ValueError as ve:
                st.error(f"⚠️ {str(ve)}")
            except Exception as e:
                st.error(
                    f"❌ حدث خطأ غير متوقع أثناء التصميم والحساب: {str(e)}"
                )

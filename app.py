import os
import re
import sqlite3
import uuid
from datetime import datetime
from io import BytesIO

import pandas as pd
import numpy as np
import streamlit as st

try:
    from pandas.io.formats.style import Styler as PandasStyler
except Exception:
    PandasStyler = None


st.set_page_config(
    page_title="نظام بلسم العلا - مطابقة الطلبات والفواتير",
    layout="wide",
    initial_sidebar_state="expanded",
)


DB_DIR = "data"
DB_PATH = os.path.join(DB_DIR, "pharmacy_reconciliation.db")
PHARMACY_COUNT = 17
SPECIAL_ORDER_NUMBERS = {"0", "123456"}
EXCLUDED_PROFILE = "FREE GIFTS FOR CUSTOMERS"
CASE_LABELS = {
    "addition": "إضافة",
    "return": "إرجاع",
    "orphan_salla": "طلب بدون فاتورة",
    "orphan_abc": "فاتورة بدون طلب",
    "post_cutoff_abc": "فاتورة بعد آخر طلب",
    "branch_mismatch": "اختلاف فرع",
    "special_review": "مراجعة رقم طلب خاص",
}
STATUS_DONE = "تم"
STATUS_PENDING = "قيد المتابعة"


st.markdown(
    """
<style>
    @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@300;400;500;700;800&display=swap');

    * { font-family: 'Tajawal', sans-serif; }

    .hero {
        background:
            radial-gradient(circle at top right, rgba(255,255,255,0.18), transparent 28%),
            linear-gradient(135deg, #0f4c5c 0%, #1f7a8c 50%, #16425b 100%);
        border-radius: 24px;
        padding: 2.2rem;
        color: white;
        margin-bottom: 1.6rem;
        box-shadow: 0 18px 40px rgba(22, 66, 91, 0.20);
    }

    .hero h1 {
        margin: 0;
        font-size: 2.2rem;
        font-weight: 800;
    }

    .hero p {
        margin-top: 0.6rem;
        font-size: 1rem;
        opacity: 0.95;
    }

    .section-title {
        font-size: 1.15rem;
        font-weight: 800;
        color: #16425b;
        border-right: 5px solid #1f7a8c;
        padding-right: 0.65rem;
        margin: 1rem 0 0.8rem;
    }

    .note-card {
        background: linear-gradient(135deg, #f4fbfc 0%, #ffffff 100%);
        border: 1px solid #d7ebef;
        border-radius: 16px;
        padding: 1rem 1.1rem;
        margin-bottom: 1rem;
    }

    .action-card {
        background: white;
        border: 1px solid #e4eef1;
        border-right: 6px solid #1f7a8c;
        border-radius: 18px;
        padding: 1rem;
        margin-bottom: 0.85rem;
        box-shadow: 0 8px 22px rgba(15, 76, 92, 0.07);
    }

    .action-card-alert {
        background: #fff1f1;
        border-color: #f1c5c5;
        border-right-color: #d9534f;
    }

    .action-card-payment {
        background: #fff7e8;
        border-color: #f3d08a;
        border-right-color: #d48b00;
    }

    .action-card-completed {
        background: #d4edda;
        border-color: #c3e6cb;
        border-right-color: #28a745;
    }

    .pill {
        display: inline-block;
        padding: 0.28rem 0.75rem;
        border-radius: 999px;
        font-size: 0.78rem;
        font-weight: 700;
    }

    .pill-green { background: #dff7e8; color: #0f7a3a; }
    .pill-amber { background: #fff0c2; color: #8a5b00; }
    .pill-red { background: #ffe0df; color: #a32929; }
    .pill-blue { background: #dff1ff; color: #0f5488; }
    .pill-slate { background: #eef3f5; color: #445b66; }
    .pill-cancel { background: #ffd8d8; color: #8f1f1f; }
    .pill-payment { background: #fff0c2; color: #8a5b00; }
    .pill-completed { background: #28a745; color: white; }

    .date-accent {
        color: #0f5488;
        font-weight: 700;
    }

    .diff-positive {
        color: #0f7a3a;
        font-weight: 800;
    }

    .diff-negative {
        color: #a32929;
        font-weight: 800;
    }

    .metric-box {
        background: white;
        border-radius: 18px;
        padding: 1rem;
        border: 1px solid #e6eef0;
        box-shadow: 0 8px 20px rgba(15, 76, 92, 0.06);
    }

    .stButton button {
        width: 100%;
        border-radius: 10px;
        font-weight: 800;
    }
    
    .session-card {
        background: #f8f9fa;
        border-radius: 12px;
        padding: 0.8rem;
        margin: 0.3rem 0;
        border-right: 3px solid #1f7a8c;
    }
    
    .lock-badge {
        display: inline-block;
        padding: 0.2rem 0.6rem;
        border-radius: 20px;
        font-size: 0.7rem;
        font-weight: bold;
    }
    .lock-closed { background: #d9534f; color: white; }
    .lock-open { background: #5cb85c; color: white; }
    
    .info-box {
        background: #e8f4f8;
        padding: 0.8rem;
        border-radius: 12px;
        margin: 0.5rem 0;
        border-right: 4px solid #1f7a8c;
    }
</style>
""",
    unsafe_allow_html=True,
)


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def pharmacy_names():
    return [f"Balsam Alula Pharmacy {i:02d}" for i in range(1, PHARMACY_COUNT + 1)]


def upgrade_database():
    """ترقية قاعدة البيانات إلى أحدث إصدار"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    cur.execute("PRAGMA table_info(uploads)")
    existing_columns = [row[1] for row in cur.fetchall()]
    
    new_columns = {
        "session_name": "TEXT DEFAULT ''",
        "is_locked": "INTEGER DEFAULT 0",
        "locked_by": "TEXT DEFAULT ''",
        "locked_at": "TEXT DEFAULT ''",
        "is_active": "INTEGER DEFAULT 0"
    }
    
    for col_name, col_type in new_columns.items():
        if col_name not in existing_columns:
            try:
                cur.execute(f"ALTER TABLE uploads ADD COLUMN {col_name} {col_type}")
            except:
                pass
    
    cur.execute("PRAGMA table_info(reconciliation_items)")
    existing_cols_items = [row[1] for row in cur.fetchall()]
    
    new_items_columns = {
        "profile_type": "TEXT DEFAULT ''",
        "profile_type_from_abc": "TEXT DEFAULT ''",
        "receipt_classification": "TEXT DEFAULT ''",
        "all_abc_pharmacies": "TEXT DEFAULT ''",
        "other_branch_details": "TEXT DEFAULT ''",
        "pharmacist_note": "TEXT DEFAULT ''",
        "hidden_from_pharmacy": "INTEGER DEFAULT 0",
        "hidden_by": "TEXT DEFAULT ''",
        "hidden_at": "TEXT DEFAULT ''"
    }
    
    for col_name, col_type in new_items_columns.items():
        if col_name not in existing_cols_items:
            try:
                cur.execute(f"ALTER TABLE reconciliation_items ADD COLUMN {col_name} {col_type}")
            except:
                pass
    
    conn.commit()
    conn.close()


def ensure_database():
    """تهيئة قاعدة البيانات مع دعم الترقية"""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            role TEXT NOT NULL,
            pharmacist_name TEXT DEFAULT '',
            last_login TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS last_access (
            pharmacy_name TEXT PRIMARY KEY,
            last_login TEXT DEFAULT '',
            pharmacist_name TEXT DEFAULT ''
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS uploads (
            upload_batch_id TEXT PRIMARY KEY,
            file_name TEXT,
            uploaded_by TEXT,
            uploaded_at TEXT,
            total_cases INTEGER DEFAULT 0,
            total_additions INTEGER DEFAULT 0,
            total_returns INTEGER DEFAULT 0,
            total_orphan_salla INTEGER DEFAULT 0,
            total_orphan_abc INTEGER DEFAULT 0,
            total_branch_mismatch INTEGER DEFAULT 0,
            total_special_review INTEGER DEFAULT 0
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS reconciliation_items (
            item_key TEXT PRIMARY KEY,
            upload_batch_id TEXT,
            order_number TEXT,
            invoice_number TEXT,
            sku TEXT,
            product_name TEXT,
            salla_product_name TEXT,
            abc_product_name TEXT,
            pharmacy_name TEXT,
            salla_pharmacy_name TEXT,
            abc_pharmacy_name TEXT,
            branch_number TEXT,
            salla_qty REAL DEFAULT 0,
            abc_qty REAL DEFAULT 0,
            difference REAL DEFAULT 0,
            case_type TEXT,
            case_label TEXT,
            case_reason TEXT,
            status TEXT DEFAULT 'قيد المتابعة',
            performed_by TEXT DEFAULT '',
            performed_at TEXT DEFAULT '',
            customer_name TEXT DEFAULT '',
            customer_phone TEXT DEFAULT '',
            city TEXT DEFAULT '',
            order_status TEXT DEFAULT '',
            order_date TEXT DEFAULT '',
            invoice_date TEXT DEFAULT '',
            total_amount REAL DEFAULT 0,
            first_seen_at TEXT,
            last_seen_at TEXT,
            active INTEGER DEFAULT 1
        )
        """
    )

    for index, name in enumerate(pharmacy_names(), start=1):
        cur.execute(
            """
            INSERT OR IGNORE INTO users (username, password, role, pharmacist_name, last_login)
            VALUES (?, ?, 'pharmacy', '', '')
            """,
            (name, f"balsam{index}"),
        )

    cur.execute(
        """
        INSERT OR IGNORE INTO users (username, password, role, pharmacist_name, last_login)
        VALUES ('admin', 'admin123', 'admin', 'Manager', '')
        """
    )

    conn.commit()
    conn.close()
    
    upgrade_database()


def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_city(value) -> str:
    city = normalize_text(value).upper()
    city = city.replace("-", " ").replace("_", " ")
    city = re.sub(r"\s+", " ", city)
    return city


def is_cancelled_or_returned_status(status_text: str) -> bool:
    status = normalize_text(status_text)
    return any(token in status for token in ["ملغي", "مسترجع"])


def is_pending_payment_status(status_text: str) -> bool:
    status = normalize_text(status_text)
    return "بانتظار الدفع" in status


def cancel_status_label(status_text: str) -> str:
    status = normalize_text(status_text)
    if "مسترجع" in status:
        return "مسترجع"
    if "ملغي" in status:
        return "ملغي"
    return ""


def normalize_order_number(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = text.replace(".0", "")
    if text.lower() in {"nan", "none", "null"}:
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def normalize_sku(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace(".0", "")
    invalid_skus = {"0", "1", "200"}
    if text in invalid_skus:
        return ""
    if re.fullmatch(r"\d+", text):
        return text
    return ""


def numeric_value(value) -> float:
    return float(pd.to_numeric(pd.Series([value]), errors="coerce").fillna(0).iloc[0])


def extract_branch_from_status(status_text):
    if not status_text or pd.isna(status_text):
        return ""
    match = re.search(r"فرع\s*(\d+)", str(status_text))
    if match:
        return f"{int(match.group(1)):02d}"
    return ""


def determine_branch(order_status: str, city: str) -> tuple[str, str]:
    branch_num = extract_branch_from_status(order_status)
    if branch_num:
        return f"Balsam Alula Pharmacy {branch_num}", branch_num

    normalized_city = normalize_city(city)
    delivered_statuses = ["تم التوصيل", "ملغي", "مسترجع", "محذوف"]
    if any(status in normalize_text(order_status) for status in delivered_statuses):
        if normalized_city in {"AL ULA", "ALULA", "AL-ULA"}:
            return "Balsam Alula Pharmacy 09", "09"
        return "Balsam Alula Pharmacy 13", "13"

    return "Balsam Alula Pharmacy 13", "13"


def get_branch_number(pharmacy_name: str) -> str:
    match = re.search(r"(\d{2})$", normalize_text(pharmacy_name))
    return match.group(1) if match else ""


def update_last_access(pharmacy_name: str, pharmacist_name: str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    current_time = now_str()
    cur.execute(
        "UPDATE users SET pharmacist_name = ?, last_login = ? WHERE username = ?",
        (pharmacist_name, current_time, pharmacy_name),
    )
    cur.execute(
        """
        INSERT INTO last_access (pharmacy_name, last_login, pharmacist_name)
        VALUES (?, ?, ?)
        ON CONFLICT(pharmacy_name) DO UPDATE SET
            last_login = excluded.last_login,
            pharmacist_name = excluded.pharmacist_name
        """,
        (pharmacy_name, current_time, pharmacist_name),
    )
    conn.commit()
    conn.close()


def fetch_user(username: str, password: str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT username, role, pharmacist_name FROM users WHERE username = ? AND password = ?",
        (username, password),
    )
    user = cur.fetchone()
    conn.close()
    return user


def get_all_last_logins() -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql_query(
            """
            SELECT pharmacy_name, last_login, pharmacist_name
            FROM last_access
            ORDER BY last_login DESC
            """,
            conn,
        )
    finally:
        conn.close()


def get_latest_upload_summary():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    cur.execute("PRAGMA table_info(uploads)")
    existing_columns = [row[1] for row in cur.fetchall()]
    
    select_cols = ["upload_batch_id", "file_name", "uploaded_by", "uploaded_at", 
                   "total_cases", "total_additions", "total_returns", 
                   "total_orphan_salla", "total_orphan_abc", 
                   "total_branch_mismatch", "total_special_review"]
    
    if "is_locked" in existing_columns:
        select_cols.append("is_locked")
    else:
        select_cols.append("0 as is_locked")
        
    if "session_name" in existing_columns:
        select_cols.append("session_name")
    else:
        select_cols.append("'' as session_name")
    
    query = f"""
        SELECT {', '.join(select_cols)}
        FROM uploads
        WHERE is_active = 1
        ORDER BY uploaded_at DESC
        LIMIT 1
    """
    
    try:
        cur.execute(query)
        row = cur.fetchone()
        conn.close()
        return row
    except Exception:
        conn.close()
        return None


def build_item_key(row: pd.Series) -> str:
    parts = [
        normalize_text(row.get("pharmacy_name")),
        normalize_text(row.get("order_number")),
        normalize_text(row.get("sku")),
        normalize_text(row.get("case_type")),
    ]
    return "||".join(parts)


def prepare_salla_frame(df_salla: pd.DataFrame) -> pd.DataFrame:
    df = df_salla.copy()
    df["order_number"] = df["رقم الطلب"].apply(normalize_order_number)
    df["sku"] = df["SKU"].apply(normalize_sku)
    df["product_name"] = df["اسم المنتج"].apply(normalize_text)
    df["quantity"] = pd.to_numeric(df["الكمية"], errors="coerce").fillna(0)
    df["customer_name"] = df["اسم العميل"].apply(normalize_text)
    df["customer_phone"] = df["رقم الجوال"].apply(normalize_text)
    df["city"] = df["المدينة"].apply(normalize_text)
    df["order_status"] = df["حالة الطلب"].apply(normalize_text)
    df["order_date"] = df["تاريخ الطلب"].apply(normalize_text)
    df["total_amount"] = pd.to_numeric(df["إجمالي الطلب"], errors="coerce").fillna(0)

    df = df[
        (df["order_number"] != "")
        & (df["sku"] != "")
        & (df["quantity"] != 0)
        & (df["order_status"] != "محذوف")
    ].copy()

    branch_info = df.apply(lambda row: determine_branch(row["order_status"], row["city"]), axis=1)
    df["pharmacy_name"] = branch_info.apply(lambda value: value[0])
    df["branch_number"] = branch_info.apply(lambda value: value[1])

    grouped = (
        df.groupby(["order_number", "sku"], as_index=False)
        .agg(
            {
                "product_name": "first",
                "quantity": "sum",
                "customer_name": "first",
                "customer_phone": "first",
                "city": "first",
                "order_status": "first",
                "order_date": "first",
                "total_amount": "first",
                "pharmacy_name": "first",
                "branch_number": "first",
            }
        )
        .rename(
            columns={
                "product_name": "salla_product_name",
                "quantity": "salla_qty",
                "pharmacy_name": "salla_pharmacy_name",
                "branch_number": "salla_branch_number",
            }
        )
    )
    return grouped


def prepare_abc_frame(df_abc: pd.DataFrame) -> pd.DataFrame:
    df = df_abc.copy()
    
    if "نوع البروفايل" in df.columns:
        df = df[df["نوع البروفايل"].astype(str).str.strip() != EXCLUDED_PROFILE].copy()
    
    if "اسم الصنف" in df.columns:
        df = df[~df["اسم الصنف"].astype(str).str.upper().str.contains("DELIVERY FEE", na=False)].copy()
    
    if "رقم الصنف" in df.columns:
        df = df[df["رقم الصنف"].astype(str).str.strip() != "16133"].copy()

    df["order_number"] = df["رقم الطلب"].apply(normalize_order_number)
    df["sku"] = df["رقم الصنف"].apply(normalize_sku)
    df["abc_product_name"] = df["اسم الصنف"].apply(normalize_text)
    df["abc_qty"] = pd.to_numeric(df["Net Sold Qty"], errors="coerce").fillna(0)
    df["invoice_number"] = df["رقم الفاتورة"].apply(normalize_text)
    df["invoice_date"] = df["التاريخ"].apply(normalize_text)
    df["abc_pharmacy_name"] = df["رقم الصيدلية"].apply(normalize_text)
    df["all_abc_pharmacies"] = df["abc_pharmacy_name"]
    df["abc_branch_number"] = df["abc_pharmacy_name"].apply(get_branch_number)
    df["profile_type"] = df["نوع البروفايل"].apply(normalize_text) if "نوع البروفايل" in df.columns else ""
    df["profile_type_from_abc"] = df["نوع البروفايل"].apply(normalize_text) if "نوع البروفايل" in df.columns else ""
    
    if "Receipt Classification" in df.columns:
        df["receipt_classification"] = df["Receipt Classification"].apply(normalize_text)
    else:
        df["receipt_classification"] = ""

    df = df[
        (df["sku"] != "")
        & (df["order_number"] != "")
    ].copy()

    grouped = (
        df.groupby(["order_number", "sku"], as_index=False)
        .agg(
            {
                "abc_qty": "sum",
                "invoice_number": "first",
                "invoice_date": "first",
                "abc_product_name": "first",
                "abc_pharmacy_name": "first",
                "abc_branch_number": "first",
                "profile_type": lambda values: " | ".join(
                    sorted({normalize_text(value) for value in values if normalize_text(value)})
                ),
                "profile_type_from_abc": lambda values: " | ".join(
                    sorted({normalize_text(value) for value in values if normalize_text(value)})
                ),
                "receipt_classification": lambda values: " | ".join(
                    sorted({normalize_text(value) for value in values if normalize_text(value)})
                ),
                "all_abc_pharmacies": lambda values: " | ".join(
                    sorted({normalize_text(value) for value in values if normalize_text(value)})
                ),
            }
        )
    )
    grouped["other_branch_details"] = grouped.apply(
        lambda row: (
            f"تم بيع نفس الطلب/الصنف في فروع أخرى: {row['all_abc_pharmacies']}"
            if " | " in row["all_abc_pharmacies"]
            else ""
        ),
        axis=1,
    )
    return grouped


def classify_cases(df_salla: pd.DataFrame, df_abc: pd.DataFrame) -> pd.DataFrame:
    salla_grouped = prepare_salla_frame(df_salla)
    abc_grouped = prepare_abc_frame(df_abc)
    max_salla_order_dt = pd.to_datetime(salla_grouped["order_date"], errors="coerce").max() if not salla_grouped.empty else pd.NaT

    merged = pd.merge(
        salla_grouped,
        abc_grouped,
        on=["order_number", "sku"],
        how="outer",
        indicator=True,
    )

    for column in [
        "salla_qty",
        "abc_qty",
        "total_amount",
    ]:
        if column in merged.columns:
            merged[column] = pd.to_numeric(merged[column], errors="coerce").fillna(0)

    text_columns = [
        "salla_product_name",
        "abc_product_name",
        "customer_name",
        "customer_phone",
        "city",
        "order_status",
        "order_date",
        "invoice_number",
        "invoice_date",
        "salla_pharmacy_name",
        "abc_pharmacy_name",
        "salla_branch_number",
        "abc_branch_number",
        "profile_type",
        "profile_type_from_abc",
        "receipt_classification",
        "all_abc_pharmacies",
        "other_branch_details",
    ]
    for column in text_columns:
        if column not in merged.columns:
            merged[column] = ""
        merged[column] = merged[column].fillna("").astype(str)

    merged["product_name"] = merged["salla_product_name"]
    missing_product = merged["product_name"].eq("")
    merged.loc[missing_product, "product_name"] = merged.loc[missing_product, "abc_product_name"]

    merged["pharmacy_name"] = merged["salla_pharmacy_name"]
    missing_pharmacy = merged["pharmacy_name"].fillna("").eq("")
    merged.loc[missing_pharmacy, "pharmacy_name"] = merged.loc[missing_pharmacy, "abc_pharmacy_name"]

    merged["branch_number"] = merged["salla_branch_number"]
    missing_branch = merged["branch_number"].fillna("").eq("")
    merged.loc[missing_branch, "branch_number"] = merged.loc[missing_branch, "abc_branch_number"]

    merged["difference"] = merged["salla_qty"] - merged["abc_qty"]
    merged["case_type"] = ""
    merged["case_reason"] = ""
    merged["invoice_datetime"] = pd.to_datetime(merged["invoice_date"], errors="coerce")

    special_mask = merged["order_number"].isin(SPECIAL_ORDER_NUMBERS)
    merged.loc[special_mask, "case_type"] = "special_review"
    merged.loc[special_mask, "case_reason"] = "رقم طلب خاص لا يمكن الاعتماد عليه في الترحيل التلقائي."

    addition_mask = (
        (merged["case_type"] == "")
        & (merged["_merge"] == "both")
        & (merged["salla_qty"] > merged["abc_qty"])
        & (merged["salla_qty"] > 0)
    )
    merged.loc[addition_mask, "case_type"] = "addition"
    merged.loc[addition_mask, "case_reason"] = "كمية الطلب أعلى من كمية الفاتورة."

    return_mask = (
        (merged["case_type"] == "")
        & (merged["_merge"] == "both")
        & (merged["abc_qty"] > merged["salla_qty"])
    )
    merged.loc[return_mask, "case_type"] = "return"
    merged.loc[return_mask, "case_reason"] = "كمية الفاتورة أعلى من كمية الطلب."

    orphan_salla_mask = (
        (merged["case_type"] == "")
        & (merged["_merge"] == "left_only")
        & (merged["salla_qty"] > 0)
    )
    merged.loc[orphan_salla_mask, "case_type"] = "orphan_salla"
    merged.loc[orphan_salla_mask, "case_reason"] = "سطر طلب موجود في سلة ولم يُعثر على سطر مطابق له في ABC."

    orphan_abc_mask = (
        (merged["case_type"] == "")
        & (merged["_merge"] == "right_only")
        & (merged["abc_qty"] != 0)
    )
    merged.loc[orphan_abc_mask, "case_type"] = "orphan_abc"
    merged.loc[orphan_abc_mask, "case_reason"] = "سطر فاتورة موجود في ABC ولم يُعثر على سطر مطابق له في سلة."

    if pd.notna(max_salla_order_dt):
        post_cutoff_mask = (
            (merged["case_type"] == "orphan_abc")
            & merged["invoice_datetime"].notna()
            & (merged["invoice_datetime"] > max_salla_order_dt)
        )
        merged.loc[post_cutoff_mask, "case_type"] = "post_cutoff_abc"
        merged.loc[post_cutoff_mask, "case_reason"] = (
            f"تاريخ الفاتورة أحدث من آخر تاريخ طلب موجود في سلة داخل الملف ({max_salla_order_dt.strftime('%Y-%m-%d %H:%M')})."
        )

    result = merged[merged["case_type"] != ""].copy()
    result["case_label"] = result["case_type"].map(CASE_LABELS)
    result["item_key"] = result.apply(build_item_key, axis=1)
    result["status"] = STATUS_PENDING
    result["performed_by"] = ""
    result["performed_at"] = ""

    ordered_columns = [
        "item_key",
        "order_number",
        "invoice_number",
        "sku",
        "product_name",
        "salla_product_name",
        "abc_product_name",
        "pharmacy_name",
        "salla_pharmacy_name",
        "abc_pharmacy_name",
        "branch_number",
        "salla_qty",
        "abc_qty",
        "difference",
        "case_type",
        "case_label",
        "case_reason",
        "customer_name",
        "customer_phone",
        "city",
        "order_status",
        "order_date",
        "invoice_date",
        "profile_type",
        "profile_type_from_abc",
        "receipt_classification",
        "all_abc_pharmacies",
        "other_branch_details",
        "total_amount",
    ]
    
    for col in ordered_columns:
        if col not in result.columns:
            result[col] = ""
    
    return result[ordered_columns]


def create_session_from_upload(upload_batch_id: str, file_name: str, uploaded_by: str):
    """إنشاء جلسة جديدة مع اسم قابل للقراءة"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    session_name = datetime.now().strftime("%Y-%m-%d %H:%M")
    cur.execute("""
        UPDATE uploads 
        SET session_name = ?, is_locked = 0, locked_by = '', locked_at = '', is_active = 1
        WHERE upload_batch_id = ?
    """, (session_name, upload_batch_id))
    
    cur.execute("""
        UPDATE uploads SET is_active = 0 
        WHERE upload_batch_id != ?
    """, (upload_batch_id,))
    
    conn.commit()
    conn.close()
    return session_name


def get_all_sessions() -> pd.DataFrame:
    """الحصول على جميع الجلسات السابقة"""
    conn = sqlite3.connect(DB_PATH)
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(uploads)")
        existing_columns = [row[1] for row in cur.fetchall()]
        
        select_cols = ["upload_batch_id", "file_name", "uploaded_by", "uploaded_at", 
                       "total_cases", "total_additions", "total_returns"]
        
        if "session_name" in existing_columns:
            select_cols.append("session_name")
        else:
            select_cols.append("'' as session_name")
            
        if "is_locked" in existing_columns:
            select_cols.append("is_locked")
        else:
            select_cols.append("0 as is_locked")
            
        if "locked_by" in existing_columns:
            select_cols.append("locked_by")
        else:
            select_cols.append("'' as locked_by")
            
        if "locked_at" in existing_columns:
            select_cols.append("locked_at")
        else:
            select_cols.append("'' as locked_at")
            
        if "is_active" in existing_columns:
            select_cols.append("is_active")
        else:
            select_cols.append("0 as is_active")
        
        query = f"""
            SELECT {', '.join(select_cols)}
            FROM uploads
            ORDER BY uploaded_at DESC
        """
        
        df = pd.read_sql_query(query, conn)
        
        required_cols = ['upload_batch_id', 'file_name', 'uploaded_by', 'uploaded_at', 
                         'total_cases', 'total_additions', 'total_returns']
        for col in required_cols:
            if col not in df.columns:
                df[col] = ''
        
        return df
    except Exception as e:
        st.warning(f"خطأ في تحميل الجلسات: {str(e)[:100]}")
        return pd.DataFrame()
    finally:
        conn.close()


def get_session_items(upload_batch_id: str) -> pd.DataFrame:
    """الحصول على عناصر جلسة محددة"""
    conn = sqlite3.connect(DB_PATH)
    try:
        return pd.read_sql_query("""
            SELECT order_number, sku, product_name, case_label, status, performed_by
            FROM reconciliation_items
            WHERE upload_batch_id = ?
        """, conn, params=(upload_batch_id,))
    finally:
        conn.close()


def lock_session(upload_batch_id: str, locked_by: str):
    """قفل الجلسة لمنع التعديل"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE uploads 
        SET is_locked = 1, locked_by = ?, locked_at = ?
        WHERE upload_batch_id = ?
    """, (locked_by, now_str(), upload_batch_id))
    conn.commit()
    conn.close()


def unlock_session(upload_batch_id: str):
    """فتح الجلسة للسماح بالتعديل"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE uploads 
        SET is_locked = 0, locked_by = '', locked_at = ''
        WHERE upload_batch_id = ?
    """, (upload_batch_id,))
    conn.commit()
    conn.close()


def activate_session(upload_batch_id: str):
    """تفعيل جلسة معينة (جعلها الجلسة النشطة)"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    cur.execute("UPDATE uploads SET is_active = 0")
    cur.execute("UPDATE uploads SET is_active = 1 WHERE upload_batch_id = ?", (upload_batch_id,))
    
    conn.commit()
    conn.close()


def hide_item_from_pharmacy(item_key: str, hidden_by: str):
    """إخفاء عنصر من ظهوره للصيدلية"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE reconciliation_items 
        SET hidden_from_pharmacy = 1, hidden_by = ?, hidden_at = ?
        WHERE item_key = ?
    """, (hidden_by, now_str(), item_key))
    conn.commit()
    conn.close()


def unhide_item_from_pharmacy(item_key: str):
    """إظهار عنصر مرة أخرى للصيدلية"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE reconciliation_items 
        SET hidden_from_pharmacy = 0, hidden_by = '', hidden_at = ''
        WHERE item_key = ?
    """, (item_key,))
    conn.commit()
    conn.close()


def reopen_case(order_number: str, sku: str, pharmacy_name: str, case_type: str):
    """إعادة فتح حالة مغلقة"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE reconciliation_items
        SET status = 'قيد المتابعة', performed_by = '', performed_at = ''
        WHERE active = 1 AND order_number = ? AND sku = ? AND pharmacy_name = ? AND case_type = ?
    """, (order_number, sku, pharmacy_name, case_type))
    conn.commit()
    conn.close()


def reopen_case_by_item_key(item_key: str):
    """إعادة فتح حالة مغلقة باستخدام المفتاح"""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        UPDATE reconciliation_items
        SET status = 'قيد المتابعة', performed_by = '', performed_at = ''
        WHERE item_key = ?
    """, (item_key,))
    conn.commit()
    conn.close()


def get_completed_items(pharmacy_name: str = None) -> pd.DataFrame:
    """الحصول على العناصر المكتملة"""
    conn = sqlite3.connect(DB_PATH)
    query = """
        SELECT order_number, invoice_number, sku, product_name, case_type, case_label,
               performed_by, performed_at, status, item_key, pharmacy_name
        FROM reconciliation_items
        WHERE active = 1 AND status = 'تم'
    """
    params = []
    if pharmacy_name:
        query += " AND pharmacy_name = ?"
        params.append(pharmacy_name)
    query += " ORDER BY performed_at DESC"
    
    try:
        df = pd.read_sql_query(query, conn, params=params if params else None)
        return df
    finally:
        conn.close()


def get_all_pharmacy_pharmacists():
    """الحصول على جميع الصيادلة المسجلين"""
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql_query("""
            SELECT username, pharmacist_name, last_login
            FROM users
            WHERE role = 'pharmacy' AND pharmacist_name != ''
            ORDER BY last_login DESC
        """, conn)
        return df
    finally:
        conn.close()


# ========== تحديث الأرصدة ==========
def update_balances(abc_file, salla_file):
    """تحديث أرصدة الفروع بناءً على ملف ABC"""
    try:
        df_abc = pd.read_excel(abc_file, skiprows=4)
        df_salla = pd.read_excel(salla_file)
        
        def get_abc_col(branch_num):
            return pd.to_numeric(df_abc.iloc[:, branch_num + 1], errors='coerce').fillna(0)
        
        item_key = df_abc.iloc[:, 0]
        
        tabuk_calc = np.floor(((get_abc_col(8) + get_abc_col(10) + get_abc_col(11) + get_abc_col(12) +
                               get_abc_col(14) + get_abc_col(15) + get_abc_col(16) + get_abc_col(17)) / 2) + get_abc_col(13))
        
        f9_calc = np.floor(((get_abc_col(1) + get_abc_col(3)) / 2) + get_abc_col(9))
        
        def create_map(values):
            return dict(zip(item_key, values.astype(int)))
        
        maps = {
            'tabuk': create_map(tabuk_calc),
            'f9': create_map(f9_calc),
            'f1': create_map(get_abc_col(1)), 'f2': create_map(get_abc_col(2)),
            'f3': create_map(get_abc_col(3)), 'f4': create_map(get_abc_col(4)),
            'f5': create_map(get_abc_col(5)), 'f6': create_map(get_abc_col(6)),
            'f7': create_map(get_abc_col(7)), 'f8': create_map(get_abc_col(8)),
            'f10': create_map(get_abc_col(10)), 'f11': create_map(get_abc_col(11)),
            'f12': create_map(get_abc_col(12)), 'f14': create_map(get_abc_col(14)),
            'f15': create_map(get_abc_col(15)), 'f16': create_map(get_abc_col(16)),
            'f17': create_map(get_abc_col(17))
        }
        
        df_updated = df_salla.copy()
        salla_id_col = 3
        
        col_mapping = {
            5: 'tabuk', 7: 'f8', 9: 'f9', 11: 'f11', 13: 'f15', 15: 'f16',
            17: 'f10', 21: 'f12', 23: 'f14', 25: 'f1', 27: 'f2', 29: 'f3',
            31: 'f4', 33: 'f5', 35: 'f6', 37: 'f7', 39: 'f17'
        }
        
        for col_idx, map_name in col_mapping.items():
            df_updated.iloc[:, col_idx] = df_updated.iloc[:, salla_id_col].map(maps[map_name]).fillna(0).astype(int)
        
        cols_to_check = list(col_mapping.keys())
        old_data = df_salla.iloc[:, cols_to_check].apply(pd.to_numeric, errors='coerce').fillna(0).astype(int)
        new_data = df_updated.iloc[:, cols_to_check]
        
        is_different = (new_data.values != old_data.values).any(axis=1)
        has_balance = new_data.sum(axis=1) > 0
        
        df_final = df_updated[is_different & has_balance]
        
        return df_final, len(df_final)
    except Exception as e:
        return None, str(e)


def persist_reconciliation_results(results: pd.DataFrame, uploaded_file_name: str, uploaded_by: str):
    upload_batch_id = uuid.uuid4().hex
    timestamp = now_str()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("BEGIN")

    cur.execute(
        """
        INSERT INTO uploads (
            upload_batch_id, file_name, uploaded_by, uploaded_at, total_cases,
            total_additions, total_returns, total_orphan_salla, total_orphan_abc,
            total_branch_mismatch, total_special_review, is_locked, is_active
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 1)
        """,
        (
            upload_batch_id,
            uploaded_file_name,
            uploaded_by,
            timestamp,
            len(results),
            int((results["case_type"] == "addition").sum()),
            int((results["case_type"] == "return").sum()),
            int((results["case_type"] == "orphan_salla").sum()),
            int((results["case_type"].isin(["orphan_abc", "post_cutoff_abc"])).sum()),
            int((results["case_type"] == "branch_mismatch").sum()),
            int((results["case_type"] == "special_review").sum()),
        ),
    )

    existing_map = {}
    existing_rows = cur.execute(
        "SELECT item_key, status, performed_by, performed_at, first_seen_at, pharmacist_note FROM reconciliation_items"
    ).fetchall()
    for item_key, status, performed_by, performed_at, first_seen_at, pharmacist_note in existing_rows:
        existing_map[item_key] = {
            "status": status,
            "performed_by": performed_by,
            "performed_at": performed_at,
            "first_seen_at": first_seen_at or timestamp,
            "pharmacist_note": pharmacist_note or "",
        }

    for _, row in results.iterrows():
        previous = existing_map.get(row["item_key"], {})
        cur.execute(
            """
            INSERT INTO reconciliation_items (
                item_key, upload_batch_id, order_number, invoice_number, sku, product_name,
                salla_product_name, abc_product_name, pharmacy_name, salla_pharmacy_name,
                abc_pharmacy_name, branch_number, salla_qty, abc_qty, difference,
                case_type, case_label, case_reason, status, performed_by, performed_at,
                customer_name, customer_phone, city, order_status, order_date,
                invoice_date, profile_type, profile_type_from_abc, receipt_classification, 
                all_abc_pharmacies, other_branch_details, pharmacist_note, total_amount, 
                first_seen_at, last_seen_at, active, hidden_from_pharmacy
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 0)
            ON CONFLICT(item_key) DO UPDATE SET
                upload_batch_id = excluded.upload_batch_id,
                order_number = excluded.order_number,
                invoice_number = excluded.invoice_number,
                sku = excluded.sku,
                product_name = excluded.product_name,
                salla_product_name = excluded.salla_product_name,
                abc_product_name = excluded.abc_product_name,
                pharmacy_name = excluded.pharmacy_name,
                salla_pharmacy_name = excluded.salla_pharmacy_name,
                abc_pharmacy_name = excluded.abc_pharmacy_name,
                branch_number = excluded.branch_number,
                salla_qty = excluded.salla_qty,
                abc_qty = excluded.abc_qty,
                difference = excluded.difference,
                case_type = excluded.case_type,
                case_label = excluded.case_label,
                case_reason = excluded.case_reason,
                customer_name = excluded.customer_name,
                customer_phone = excluded.customer_phone,
                city = excluded.city,
                order_status = excluded.order_status,
                order_date = excluded.order_date,
                invoice_date = excluded.invoice_date,
                profile_type = excluded.profile_type,
                profile_type_from_abc = excluded.profile_type_from_abc,
                receipt_classification = excluded.receipt_classification,
                all_abc_pharmacies = excluded.all_abc_pharmacies,
                other_branch_details = excluded.other_branch_details,
                pharmacist_note = reconciliation_items.pharmacist_note,
                total_amount = excluded.total_amount,
                first_seen_at = reconciliation_items.first_seen_at,
                last_seen_at = excluded.last_seen_at,
                active = 1,
                hidden_from_pharmacy = reconciliation_items.hidden_from_pharmacy
            """,
            (
                row["item_key"],
                upload_batch_id,
                row["order_number"],
                row["invoice_number"],
                row["sku"],
                row["product_name"],
                row["salla_product_name"],
                row["abc_product_name"],
                row["pharmacy_name"],
                row["salla_pharmacy_name"],
                row["abc_pharmacy_name"],
                row["branch_number"],
                numeric_value(row["salla_qty"]),
                numeric_value(row["abc_qty"]),
                numeric_value(row["difference"]),
                row["case_type"],
                row["case_label"],
                row["case_reason"],
                previous.get("status", STATUS_PENDING),
                previous.get("performed_by", ""),
                previous.get("performed_at", ""),
                row["customer_name"],
                row["customer_phone"],
                row["city"],
                row["order_status"],
                row["order_date"],
                row["invoice_date"],
                row.get("profile_type", ""),
                row.get("profile_type_from_abc", ""),
                row.get("receipt_classification", ""),
                row.get("all_abc_pharmacies", ""),
                row.get("other_branch_details", ""),
                previous.get("pharmacist_note", ""),
                numeric_value(row["total_amount"]),
                previous.get("first_seen_at", timestamp),
                timestamp,
            ),
        )

    cur.execute(
        """
        UPDATE reconciliation_items
        SET active = CASE WHEN upload_batch_id = ? THEN 1 ELSE 0 END
        """,
        (upload_batch_id,),
    )

    conn.commit()
    conn.close()
    
    create_session_from_upload(upload_batch_id, uploaded_file_name, uploaded_by)
    
    return upload_batch_id


def process_excel(uploaded_file, uploaded_by: str):
    df_salla = pd.read_excel(uploaded_file, sheet_name="سلة")
    df_abc = pd.read_excel(uploaded_file, sheet_name="abc")
    results = classify_cases(df_salla, df_abc)
    upload_batch_id = persist_reconciliation_results(results, uploaded_file.name, uploaded_by)
    return results, upload_batch_id


def fetch_active_items(pharmacy_name: str | None = None, include_hidden: bool = False) -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    
    cur = conn.cursor()
    cur.execute("""
        SELECT upload_batch_id FROM uploads 
        WHERE is_active = 1
        ORDER BY uploaded_at DESC LIMIT 1
    """)
    active_session = cur.fetchone()
    
    if not active_session:
        conn.close()
        return pd.DataFrame()
    
    active_batch_id = active_session[0]
    
    cur.execute("PRAGMA table_info(uploads)")
    existing_columns = [row[1] for row in cur.fetchall()]
    has_lock_column = "is_locked" in existing_columns
    
    if has_lock_column:
        cur.execute("SELECT is_locked FROM uploads WHERE upload_batch_id = ?", (active_batch_id,))
        lock_result = cur.fetchone()
        is_locked = lock_result[0] if lock_result else 0
    else:
        is_locked = 0
    
    cur.execute("PRAGMA table_info(reconciliation_items)")
    item_columns = [row[1] for row in cur.fetchall()]
    has_hidden_column = "hidden_from_pharmacy" in item_columns
    
    query = """
        SELECT order_number, invoice_number, sku, product_name, pharmacy_name, branch_number,
               salla_qty, abc_qty, difference, case_type, case_label, case_reason, status,
               performed_by, performed_at, customer_name, customer_phone, city, order_status,
               order_date, invoice_date, total_amount, first_seen_at, last_seen_at,
               profile_type, profile_type_from_abc, receipt_classification, 
               all_abc_pharmacies, other_branch_details, pharmacist_note, item_key,
               abc_pharmacy_name,
               ? as is_locked
    """
    
    if has_hidden_column:
        query += ", hidden_from_pharmacy"
    else:
        query += ", 0 as hidden_from_pharmacy"
    
    query += """
        FROM reconciliation_items
        WHERE active = 1 AND upload_batch_id = ? AND status != 'تم'
    """
    params = [1 if is_locked else 0, active_batch_id]
    
    if pharmacy_name:
        query += " AND pharmacy_name = ?"
        params.append(pharmacy_name)
        
        if not include_hidden and has_hidden_column:
            query += " AND (hidden_from_pharmacy = 0 OR hidden_from_pharmacy IS NULL)"
    
    query += " ORDER BY case_type, order_number DESC, sku"
    
    try:
        df = pd.read_sql_query(query, conn, params=params)
        for col in ['abc_pharmacy_name', 'profile_type', 'profile_type_from_abc', 'receipt_classification', 'all_abc_pharmacies', 'other_branch_details', 'pharmacist_note', 'item_key']:
            if col not in df.columns:
                df[col] = ''
        return df
    except Exception as e:
        st.warning(f"خطأ في جلب البيانات: {str(e)[:100]}")
        return pd.DataFrame()
    finally:
        conn.close()


def mark_case_done(order_number: str, sku: str, pharmacy_name: str, case_type: str, performed_by: str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE reconciliation_items
        SET status = ?, performed_by = ?, performed_at = ?
        WHERE active = 1 AND order_number = ? AND sku = ? AND pharmacy_name = ? AND case_type = ?
        """,
        (STATUS_DONE, performed_by, now_str(), order_number, sku, pharmacy_name, case_type),
    )
    conn.commit()
    conn.close()


def save_case_note(order_number: str, sku: str, pharmacy_name: str, case_type: str, note: str):
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE reconciliation_items
        SET pharmacist_note = ?
        WHERE active = 1 AND order_number = ? AND sku = ? AND pharmacy_name = ? AND case_type = ?
        """,
        (note, order_number, sku, pharmacy_name, case_type),
    )
    conn.commit()
    conn.close()


def status_pill(status: str) -> str:
    if status == STATUS_DONE:
        return '<span class="pill pill-completed">مغلق</span>'
    return '<span class="pill pill-amber">قيد المتابعة</span>'


def case_pill(case_type: str) -> str:
    mapping = {
        "addition": "pill-blue",
        "return": "pill-red",
        "orphan_salla": "pill-amber",
        "orphan_abc": "pill-slate",
        "branch_mismatch": "pill-red",
        "special_review": "pill-slate",
    }
    css_class = mapping.get(case_type, "pill-slate")
    return f'<span class="pill {css_class}">{CASE_LABELS.get(case_type, case_type)}</span>'


def status_alert_pill(order_status: str) -> str:
    label = cancel_status_label(order_status)
    return f'<span class="pill pill-cancel">{label}</span>' if label else ""


def payment_alert_pill(order_status: str) -> str:
    return '<span class="pill pill-payment">بانتظار الدفع</span>' if is_pending_payment_status(order_status) else ""


def render_metrics(df: pd.DataFrame):
    cols = st.columns(6)
    metrics = [
        ("إجمالي الحالات", len(df)),
        ("إضافات", int((df["case_type"] == "addition").sum())),
        ("إرجاعات", int((df["case_type"] == "return").sum())),
        ("طلبات بدون فاتورة", int((df["case_type"] == "orphan_salla").sum())),
        ("فواتير بدون طلب", int((df["case_type"] == "orphan_abc").sum())),
        ("تم إنجازها", int((df["status"] == STATUS_DONE).sum())),
    ]
    for col, (label, value) in zip(cols, metrics):
        with col:
            st.markdown(
                f"""
                <div class="metric-box">
                    <div style="font-size:0.88rem;color:#5a7380;">{label}</div>
                    <div style="font-size:2rem;font-weight:800;color:#16425b;">{value}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def render_case_cards(df: pd.DataFrame, allow_actions: bool, pharmacist_name: str, pharmacy_name: str, is_admin: bool = False):
    if df.empty:
        st.success("لا توجد حالات في هذا القسم.")
        return

    for idx, row in df.iterrows():
        with st.container():
            col1, col2 = st.columns([4, 1])
            
            with col1:
                subcol1, subcol2, subcol3, subcol4 = st.columns(4)
                with subcol1:
                    st.markdown(f"**📋 رقم الطلب:** {row['order_number']}")
                    st.markdown(f"**🏷️ SKU:** {row['sku']}")
                    diff_value = numeric_value(row['difference'])
                    st.markdown(f"**📊 الفرق:** {'+' if diff_value > 0 else ''}{diff_value}")
                with subcol2:
                    st.markdown(f"**📦 المنتج:** {row['product_name'][:40]}...")
                    st.markdown(f"**🏥 الفرع:** {row['pharmacy_name'] or 'غير محدد'}")
                with subcol3:
                    invoice_display = row['invoice_number'] or 'غير متوفر'
                    abc_pharmacy = row.get('abc_pharmacy_name', '')
                    # استخراج اسم الصيدلي بدلاً من رقم الصيدلية
                    pharmacist_display = row.get('pharmacist_from_abc', '') or row.get('pharmacist', '') or abc_pharmacy
                    if pharmacist_display and '|' in pharmacist_display:
                        pharmacist_display = pharmacist_display.split('|')[0]
                    # تنظيف الاسم (إزالة كلمات زائدة)
                    if pharmacist_display:
                        # محاولة استخراج اسم الصيدلي من النص
                        if 'PHARMACIST' in pharmacist_display.upper():
                            parts = pharmacist_display.split()
                            for part in parts:
                                if len(part) > 2 and not part.isdigit():
                                    pharmacist_display = part
                                    break
                        elif len(pharmacist_display) > 20:
                            pharmacist_display = pharmacist_display[:20]
                    st.markdown(f"**🧾 رقم الفاتورة/الصيدلي:** {invoice_display}/{pharmacist_display if pharmacist_display else '?'}")
                    st.markdown(f"**📅 تاريخ الطلب:** {row['order_date'][:16] if row['order_date'] else 'غير متوفر'}")
                with subcol4:
                    st.markdown(f"**✅ الحالة:** {status_pill(row['status'])}", unsafe_allow_html=True)
                    st.markdown(f"**🧾 تاريخ الفاتورة:** {row['invoice_date'][:16] if row['invoice_date'] else 'غير متوفر'}")
            
            with col2:
                if is_admin and row.get('hidden_from_pharmacy', 0) == 1:
                    if st.button("👁️ إظهار", key=f"unhide_{idx}_{row['sku']}", use_container_width=True):
                        unhide_item_from_pharmacy(row['item_key'])
                        st.rerun()
                elif is_admin and row.get('hidden_from_pharmacy', 0) == 0:
                    if st.button("🙈 إخفاء", key=f"hide_{idx}_{row['sku']}", use_container_width=True):
                        hide_item_from_pharmacy(row['item_key'], st.session_state.username)
                        st.rerun()
            
            with st.expander("📋 تفاصيل إضافية"):
                detail_cols = st.columns(3)
                with detail_cols[0]:
                    st.markdown(f"**كمية سلة:** {int(row['salla_qty']) if pd.notna(row['salla_qty']) else 0}")
                    st.markdown(f"**كمية ABC:** {int(row['abc_qty']) if pd.notna(row['abc_qty']) else 0}")
                with detail_cols[1]:
                    st.markdown(f"**حالة الطلب:** {row['order_status'] or 'غير متوفرة'}")
                    st.markdown(f"**نوع البروفايل:** {row.get('profile_type', '') or row.get('profile_type_from_abc', '') or 'غير متوفر'}")
                with detail_cols[2]:
                    st.markdown(f"**تصنيف البيع:** {row.get('receipt_classification', '') or 'غير متوفر'}")
                    st.markdown(f"**فروع ABC:** {row.get('all_abc_pharmacies', '') or row.get('abc_pharmacy_name', '') or 'غير متوفر'}")
                st.markdown(f"**التفصيل:** {row.get('case_reason', '')}")
            
            note_key = f"note_{row['case_type']}_{row['order_number']}_{row['sku']}_{idx}"
            note_value = st.text_area(
                "📝 ملحوظة الصيدلي",
                value=row.get("pharmacist_note", "") or "",
                key=note_key,
                height=60,
            )
            
            btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 4])
            with btn_col1:
                if st.button("💾 حفظ", key=f"save_{note_key}", use_container_width=True):
                    save_case_note(
                        order_number=row["order_number"],
                        sku=row["sku"],
                        pharmacy_name=pharmacy_name,
                        case_type=row["case_type"],
                        note=note_value,
                    )
                    st.rerun()
            
            if allow_actions and row["status"] != STATUS_DONE and row["case_type"] in {"addition", "return", "orphan_salla", "orphan_abc"}:
                button_label = "✅ تأكيد" if row["case_type"] in {"addition", "orphan_salla"} else "🔄 تأكيد"
                with btn_col2:
                    if st.button(button_label, key=f"done_{note_key}", use_container_width=True):
                        save_case_note(
                            order_number=row["order_number"],
                            sku=row["sku"],
                            pharmacy_name=pharmacy_name,
                            case_type=row["case_type"],
                            note=note_value,
                        )
                        mark_case_done(
                            order_number=row["order_number"],
                            sku=row["sku"],
                            pharmacy_name=pharmacy_name,
                            case_type=row["case_type"],
                            performed_by=pharmacist_name,
                        )
                        st.rerun()
            
            if is_admin and row["status"] == STATUS_DONE:
                with btn_col2:
                    if st.button("🔓 إعادة فتح", key=f"reopen_{note_key}", use_container_width=True):
                        reopen_case(
                            order_number=row["order_number"],
                            sku=row["sku"],
                            pharmacy_name=pharmacy_name,
                            case_type=row["case_type"],
                        )
                        st.rerun()
            
            st.markdown("---")

def render_completed_table(df: pd.DataFrame, is_admin: bool = False):
    """عرض جدول المكتملات"""
    if df.empty:
        st.info("لا توجد طلبات مكتملة بعد.")
        return
    
    display_df = df.copy()
    display_df = display_df.rename(columns={
        "order_number": "رقم الطلب",
        "invoice_number": "رقم الفاتورة",
        "sku": "SKU",
        "product_name": "المنتج",
        "case_label": "نوع الإجراء",
        "performed_by": "تم بواسطة",
        "performed_at": "تاريخ الإكمال",
        "status": "الحالة",
        "pharmacy_name": "الصيدلية"
    })
    
    cols_to_show = ["رقم الطلب", "رقم الفاتورة", "SKU", "المنتج", "نوع الإجراء", "تم بواسطة", "تاريخ الإكمال"]
    
    if is_admin:
        cols_to_show.insert(1, "الصيدلية")
        st.dataframe(display_df[cols_to_show], use_container_width=True)
        
        for idx, row in display_df.iterrows():
            if st.button(f"🔓 إعادة فتح الطلب {row['رقم الطلب']}", key=f"reopen_completed_{idx}"):
                if 'item_key' in df.iloc[idx]:
                    reopen_case_by_item_key(df.iloc[idx]['item_key'])
                    st.rerun()
    else:
        st.dataframe(display_df[cols_to_show], use_container_width=True)


def prepare_display_df(df: pd.DataFrame) -> pd.DataFrame:
    display_df = df.copy()
    display_df = display_df.rename(
        columns={
            "order_number": "رقم الطلب",
            "invoice_number": "رقم الفاتورة",
            "sku": "SKU",
            "product_name": "الصنف",
            "pharmacy_name": "الفرع",
            "salla_qty": "كمية سلة",
            "abc_qty": "كمية ABC",
            "difference": "الفرق",
            "case_label": "نوع الحالة",
            "status": "الحالة",
            "performed_by": "تم بواسطة",
            "performed_at": "تاريخ التنفيذ",
            "customer_name": "العميل",
            "customer_phone": "جوال العميل",
            "city": "المدينة",
            "order_status": "حالة الطلب",
            "order_date": "تاريخ الطلب",
            "invoice_date": "تاريخ الفاتورة",
            "profile_type": "نوع البروفايل",
            "receipt_classification": "تصنيف البيع",
            "all_abc_pharmacies": "الفروع الظاهرة في ABC",
            "other_branch_details": "ملاحظة الفروع",
            "pharmacist_note": "ملحوظة الصيدلي",
            "case_reason": "تفصيل الحالة",
            "first_seen_at": "أول ظهور",
            "last_seen_at": "آخر تحديث",
        }
    )
    return display_df


def export_tabs_to_excel(dataframes_by_sheet: dict[str, pd.DataFrame]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    output = BytesIO()
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "ملخص"
    default_sheet.append(["ملاحظة"])
    default_sheet.append(["لا توجد بيانات مطابقة للتصدير حسب الفلاتر الحالية."])

    wrote_any_sheet = False
    seen_sheet_names = set()

    for sheet_name, sheet_df in dataframes_by_sheet.items():
        if sheet_df is None:
            continue

        if PandasStyler is not None and isinstance(sheet_df, PandasStyler):
            export_df = sheet_df.data.copy()
        elif isinstance(sheet_df, pd.Series):
            export_df = sheet_df.to_frame()
        elif isinstance(sheet_df, pd.DataFrame):
            export_df = sheet_df.copy()
        else:
            export_df = pd.DataFrame(sheet_df)

        safe_sheet_name = (str(sheet_name).strip() or "Sheet")[:31]
        if safe_sheet_name in seen_sheet_names:
            suffix = 2
            base_name = safe_sheet_name[:28]
            while f"{base_name}_{suffix}" in seen_sheet_names:
                suffix += 1
            safe_sheet_name = f"{base_name}_{suffix}"
        seen_sheet_names.add(safe_sheet_name)

        worksheet = workbook.create_sheet(title=safe_sheet_name)
        if export_df.empty:
            worksheet.append(["لا توجد بيانات"])
        else:
            for row in dataframe_to_rows(export_df, index=False, header=True):
                worksheet.append(row)
        wrote_any_sheet = True

    if wrote_any_sheet and "ملخص" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["ملخص"])

    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def render_balances_updater():
    """صفحة تحديث الأرصدة"""
    st.markdown(
        """
        <div class="hero">
            <h1>🔄 تحديث أرصدة الفروع</h1>
            <p>رفع ملفات ABC و Salla لتحديث الأرصدة بناءً على المعادلات المحددة</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    col1, col2 = st.columns(2)
    with col1:
        abc_file = st.file_uploader("📊 رفع ملف ABC (يبدأ من الصف 5)", type=["xlsx"], key="abc_balances")
    with col2:
        salla_file = st.file_uploader("📋 رفع ملف Salla", type=["xlsx"], key="salla_balances")
    
    if abc_file and salla_file:
        if st.button("🔄 تنفيذ تحديث الأرصدة", use_container_width=True):
            with st.spinner("جاري تحديث الأرصدة..."):
                result_df, result = update_balances(abc_file, salla_file)
                if result_df is not None:
                    st.success(f"✅ تم التحديث بنجاح! عدد الأصناف المحدثة: {result}")
                    st.dataframe(result_df.head(20), use_container_width=True)
                    
                    output = BytesIO()
                    result_df.to_excel(output, index=False)
                    output.seek(0)
                    st.download_button(
                        "📥 تحميل ملف Salla المحدث",
                        data=output,
                        file_name=f"salla_updated_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    st.error(f"❌ خطأ في التحديث: {result}")


def render_pharmacy_monitoring():
    """صفحة مراقبة تعديلات الصيدليات"""
    st.markdown(
        """
        <div class="hero">
            <h1>👥 مراقبة تعديلات الصيدليات</h1>
            <p>متابعة إنجازات الصيادلة وتعديلاتهم على الطلبات</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    
    # عرض الصيادلة المسجلين
    st.markdown('<div class="section-title">👤 الصيادلة المسجلون</div>', unsafe_allow_html=True)
    pharmacists_df = get_all_pharmacy_pharmacists()
    if not pharmacists_df.empty:
        st.dataframe(pharmacists_df, use_container_width=True)
    else:
        st.info("لا يوجد صيادلة مسجلون بعد")
    
    st.markdown("---")
    
    # عرض التعديلات
    st.markdown('<div class="section-title">📋 سجل التعديلات</div>', unsafe_allow_html=True)
    
    conn = sqlite3.connect(DB_PATH)
    try:
        adjustments_df = pd.read_sql_query("""
            SELECT order_number, sku, product_name, pharmacy_name, case_type, 
                   status, performed_by, performed_at, pharmacist_note
            FROM reconciliation_items
            WHERE performed_by != '' AND status = 'تم'
            ORDER BY performed_at DESC
            LIMIT 100
        """, conn)
        
        if not adjustments_df.empty:
            adjustments_df = adjustments_df.rename(columns={
                "order_number": "رقم الطلب",
                "sku": "SKU",
                "product_name": "المنتج",
                "pharmacy_name": "الصيدلية",
                "case_type": "نوع الإجراء",
                "status": "الحالة",
                "performed_by": "تم بواسطة",
                "performed_at": "تاريخ التنفيذ",
                "pharmacist_note": "ملحوظة"
            })
            st.dataframe(adjustments_df, use_container_width=True)
            
            # زر تصدير
            output = BytesIO()
            adjustments_df.to_excel(output, index=False)
            output.seek(0)
            st.download_button(
                "📥 تصدير سجل التعديلات إلى Excel",
                data=output,
                file_name=f"pharmacy_adjustments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("لا توجد تعديلات مسجلة بعد")
    finally:
        conn.close()


def render_admin_dashboard():
    st.markdown(
        """
        <div class="hero">
            <h1>لوحة التحكم الإدارية</h1>
            <p>مطابقة أكثر دقة بين سلة و ABC مع فصل الحالات الفعلية عن السطور غير المربوطة وحفظ الإنجاز بين كل رفعة وأخرى.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    latest = get_latest_upload_summary()
    if latest:
        batch_id, file_name, uploaded_by, uploaded_at, total_cases, additions, returns, orphan_salla, orphan_abc, branch_mismatch, special_review, is_locked, session_name = latest
        lock_status = "🔒 مقفلة" if is_locked else "🔓 مفتوحة"
        st.markdown(
            f"""
            <div class="note-card">
                <strong>الجلسة النشطة:</strong> {session_name or 'غير مسماة'} &nbsp; | &nbsp;
                <strong>الملف:</strong> {file_name} &nbsp; | &nbsp;
                <strong>بواسطة:</strong> {uploaded_by} &nbsp; | &nbsp;
                <strong>التاريخ:</strong> {uploaded_at[:16] if uploaded_at else ''} &nbsp; | &nbsp;
                <strong>الحالة:</strong> {lock_status}
            </div>
            """,
            unsafe_allow_html=True,
        )

    refresh_col, info_col = st.columns([1, 4])
    with refresh_col:
        if st.button("تحديث الصفحة", use_container_width=True):
            st.rerun()
    with info_col:
        st.caption("يعرض آخر الإجراءات المنفذة وآخر دخول للصيادلة والحالات المحدثة بعد الرفع.")

    with st.expander("رفع ملف الطلبات والفواتير", expanded=True):
        uploaded_file = st.file_uploader("اختر ملف Excel", type=["xlsx"], key="reconciliation")
        if uploaded_file and st.button("معالجة الملف و ترحيل الحالات", use_container_width=True):
            with st.spinner("جاري قراءة الملف وتصنيف الحالات بدقة..."):
                results, upload_batch_id = process_excel(uploaded_file, st.session_state.username)
            st.success(f"✅ تمت المعالجة بنجاح. تم إنشاء جلسة جديدة: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            st.session_state.last_processed_preview = results.head(20)
            st.rerun()

    st.markdown('<div class="section-title">📋 إدارة الجلسات السابقة</div>', unsafe_allow_html=True)
    
    sessions_df = get_all_sessions()
    if not sessions_df.empty:
        for _, session in sessions_df.iterrows():
            col1, col2, col3, col4, col5 = st.columns([2.5, 2.5, 2, 1.5, 2])
            
            session_name_val = session.get('session_name', '')
            if not session_name_val or pd.isna(session_name_val):
                session_name_val = session['upload_batch_id'][:8]
            
            with col1:
                st.markdown(f"""
                <div class="session-card">
                    <strong>📅 {session_name_val}</strong><br>
                    <small>{session['file_name'][:35] if session['file_name'] else ''}</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                is_active = session.get('is_active', 0)
                active_badge = "✅ نشطة" if is_active else "⏸ غير نشطة"
                st.markdown(f"""
                <div class="session-card">
                    <small>👤 {session['uploaded_by']}<br>
                    📅 {session['uploaded_at'][:16] if session['uploaded_at'] else ''}<br>
                    {active_badge}</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="session-card">
                    <small>📊 {int(session.get('total_cases', 0))} حالة<br>
                    ➕ {int(session.get('total_additions', 0))} | ➖ {int(session.get('total_returns', 0))}</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                is_locked = session.get('is_locked', 0)
                lock_class = "lock-closed" if is_locked else "lock-open"
                lock_text = "مقفلة" if is_locked else "مفتوحة"
                st.markdown(f"""
                <div class="session-card" style="text-align: center;">
                    <span class="lock-badge {lock_class}">🔒 {lock_text}</span>
                    <br><small>{session.get('locked_by', '')[:15] if session.get('locked_by') else ''}</small>
                </div>
                """, unsafe_allow_html=True)
            
            with col5:
                btn1, btn2, btn3 = st.columns(3)
                with btn1:
                    if not is_locked:
                        if st.button(f"🔒 قفل", key=f"lock_{session['upload_batch_id']}", use_container_width=True):
                            lock_session(session['upload_batch_id'], st.session_state.username)
                            st.rerun()
                    else:
                        if st.button(f"🔓 فتح", key=f"unlock_{session['upload_batch_id']}", use_container_width=True):
                            unlock_session(session['upload_batch_id'])
                            st.rerun()
                
                with btn2:
                    if not is_active:
                        if st.button(f"⭐ تفعيل", key=f"activate_{session['upload_batch_id']}", use_container_width=True):
                            activate_session(session['upload_batch_id'])
                            st.rerun()
                
                with btn3:
                    if st.button(f"👁️ عرض", key=f"view_{session['upload_batch_id']}", use_container_width=True):
                        st.session_state.view_session_id = session['upload_batch_id']
                        st.rerun()
        
        st.markdown("---")
    
    if st.session_state.get('view_session_id'):
        st.markdown(f'<div class="section-title">📄 عرض الجلسة المحددة</div>', unsafe_allow_html=True)
        
        session_items = get_session_items(st.session_state.view_session_id)
        if not session_items.empty:
            st.dataframe(session_items, use_container_width=True)
        
        if st.button("إغلاق العرض", use_container_width=True):
            del st.session_state.view_session_id
            st.rerun()

    df = fetch_active_items(include_hidden=True)
    if df.empty:
        st.info("لا توجد بيانات فعالة بعد. ارفع الملف من الأعلى لبدء التحليل.")
        return

    render_metrics(df)
    
    # فلتر الصيدلية في تبويب تم الانتهاء
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        branch_options = ["الكل"] + sorted(df["pharmacy_name"].dropna().astype(str).unique().tolist())
        selected_branch = st.selectbox("فلتر الفرع", branch_options)
    with col2:
        performer_values = sorted({value for value in df["performed_by"].fillna("").astype(str).tolist() if value.strip()})
        selected_performer = st.selectbox("فلتر المنفذ", ["الكل"] + performer_values)
    with col3:
        date_from = st.date_input("من تاريخ", value=None)
    with col4:
        date_to = st.date_input("إلى تاريخ", value=None)

    filtered_df = df.copy()
    if selected_branch != "الكل":
        filtered_df = filtered_df[filtered_df["pharmacy_name"] == selected_branch]
    if selected_performer != "الكل":
        filtered_df = filtered_df[filtered_df["performed_by"] == selected_performer]
    if date_from or date_to:
        pivot_dates = pd.to_datetime(
            filtered_df["performed_at"].replace("", pd.NA).fillna(filtered_df["invoice_date"]).fillna(filtered_df["order_date"]),
            errors="coerce",
        )
        if date_from:
            filtered_df = filtered_df[pivot_dates.dt.date >= date_from]
            pivot_dates = pd.to_datetime(
                filtered_df["performed_at"].replace("", pd.NA).fillna(filtered_df["invoice_date"]).fillna(filtered_df["order_date"]),
                errors="coerce",
            )
        if date_to:
            filtered_df = filtered_df[pivot_dates.dt.date <= date_to]

    admin_df = prepare_display_df(filtered_df)

    st.markdown('<div class="section-title">👥 آخر دخول للصيدليات</div>', unsafe_allow_html=True)
    last_logins = get_all_last_logins()
    if not last_logins.empty:
        cols = st.columns(4)
        for idx, (_, row) in enumerate(last_logins.head(8).iterrows()):
            with cols[idx % 4]:
                st.markdown(
                    f"""
                    <div class="note-card">
                        <strong>{row['pharmacy_name'][-10:]}</strong><br>
                        <span style="color:#58707a;">{row['pharmacist_name'] or 'غير مسجل'}</span><br>
                        <span style="color:#58707a;">{row['last_login'][:16] if row['last_login'] else 'لم يدخل بعد'}</span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

    additions_admin = admin_df[(filtered_df["case_type"] == "addition") & (~filtered_df["order_status"].apply(is_cancelled_or_returned_status))]
    returns_admin = admin_df[(filtered_df["case_type"] == "return") & (~filtered_df["order_status"].apply(is_cancelled_or_returned_status))]
    orphan_salla_admin = admin_df[(filtered_df["case_type"] == "orphan_salla") & (~filtered_df["order_status"].apply(is_cancelled_or_returned_status))]
    orphan_abc_admin = admin_df[(filtered_df["case_type"] == "orphan_abc") & (~filtered_df["order_status"].apply(is_cancelled_or_returned_status))]
    post_cutoff_admin = admin_df[filtered_df["case_type"] == "post_cutoff_abc"]
    payment_pending_admin = admin_df[filtered_df["order_status"].apply(is_pending_payment_status)]
    cancelled_admin = prepare_display_df(filtered_df[filtered_df["order_status"].apply(is_cancelled_or_returned_status)])
    
    # تبويب تم الانتهاء مع فلتر الصيدلية
    completed_df = get_completed_items()
    if selected_branch != "الكل":
        completed_df = completed_df[completed_df["pharmacy_name"] == selected_branch]

    export_bytes = export_tabs_to_excel(
        {
            "الإضافات": additions_admin,
            "الإرجاعات": returns_admin,
            "طلبات بدون فاتورة": orphan_salla_admin,
            "فواتير بدون طلب": orphan_abc_admin,
            "فواتير بعد آخر طلب": post_cutoff_admin,
            "بانتظار الدفع": payment_pending_admin,
            "الملغي_والمسترجع": cancelled_admin,
            "✅ تم الانتهاء": prepare_display_df(completed_df),
        }
    )
    st.download_button(
        "📥 تصدير كل التبويبات إلى Excel",
        data=export_bytes,
        file_name=f"balsam_reconciliation_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(
        ["الإضافات", "الإرجاعات", "طلبات بدون فاتورة", "فواتير بدون طلب", 
         "فواتير بعد آخر طلب", "بانتظار الدفع", "الملغي/المسترجع", "✅ تم الانتهاء"]
    )

    def styled_frame(input_df):
        def row_style(row):
            case_type = row.get("نوع الحالة", "")
            order_status = row.get("حالة الطلب", "")
            status = row.get("الحالة", "")
            
            if status == "تم":
                color = "background-color: #d4edda"
            elif is_cancelled_or_returned_status(order_status):
                color = "background-color: #ffe5e5"
            elif is_pending_payment_status(order_status):
                color = "background-color: #fff4d6"
            elif case_type == "إرجاع":
                color = "background-color: #ffe0df"
            elif case_type == "إضافة":
                color = "background-color: #dff1ff"
            else:
                color = "background-color: #ffe9cc"
            return [color] * len(row)

        return input_df.style.apply(row_style, axis=1)

    with tab1:
        st.dataframe(styled_frame(additions_admin), use_container_width=True)
    with tab2:
        st.dataframe(styled_frame(returns_admin), use_container_width=True)
    with tab3:
        st.dataframe(styled_frame(orphan_salla_admin), use_container_width=True)
    with tab4:
        st.dataframe(styled_frame(orphan_abc_admin), use_container_width=True)
    with tab5:
        st.dataframe(styled_frame(post_cutoff_admin), use_container_width=True)
    with tab6:
        st.dataframe(styled_frame(payment_pending_admin), use_container_width=True)
    with tab7:
        st.dataframe(styled_frame(cancelled_admin), use_container_width=True)
    with tab8:
        render_completed_table(completed_df, is_admin=True)


def render_pharmacy_dashboard():
    pharmacy_name = st.session_state.username
    pharmacist_name = st.session_state.pharmacist_name or ""
    branch_number = get_branch_number(pharmacy_name)

    st.markdown(
        f"""
        <div class="hero">
            <h1>{pharmacy_name}</h1>
            <p>فرع رقم {branch_number} | الصيدلي: {pharmacist_name}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if not pharmacist_name:
        st.warning("⚠️ الرجاء إدخال اسم الصيدلي من القائمة الجانبية أولاً")
        return

    if st.button("🔄 تحديث الصفحة", use_container_width=True):
        st.rerun()

    df = fetch_active_items(pharmacy_name, include_hidden=False)
    
    if df.empty:
        st.info("لا توجد حالات نشطة لهذا الفرع حاليًا.")
        completed_df = get_completed_items(pharmacy_name)
        if not completed_df.empty:
            st.markdown("---")
            st.markdown('<div class="section-title">✅ الطلبات المكتملة</div>', unsafe_allow_html=True)
            render_completed_table(completed_df, is_admin=False)
        return

    is_locked = df['is_locked'].iloc[0] == 1 if not df.empty else False
    
    if is_locked:
        st.warning("🔒 هذه الجلسة مقفلة ولا يمكن إجراء تعديلات عليها.")
        allow_actions = False
    else:
        allow_actions = True

    render_metrics(df)

    active_non_cancelled = ~df["order_status"].apply(is_cancelled_or_returned_status)
    active_non_payment = ~df["order_status"].apply(is_pending_payment_status)
    active_operational = active_non_cancelled & active_non_payment
    
    additions_df = df[(df["case_type"] == "addition") & active_operational].copy()
    returns_df = df[(df["case_type"] == "return") & active_operational].copy()
    orphan_salla_df = df[(df["case_type"] == "orphan_salla") & active_operational].copy()
    orphan_abc_df = df[(df["case_type"] == "orphan_abc") & active_operational].copy()
    post_cutoff_df = df[df["case_type"] == "post_cutoff_abc"].copy()
    cancelled_df = df[df["order_status"].apply(is_cancelled_or_returned_status)].copy()
    payment_pending_df = df[df["order_status"].apply(is_pending_payment_status)].copy()
    review_df = df[(df["case_type"].isin(["branch_mismatch", "special_review"])) & active_operational].copy()

    pharmacy_export = export_tabs_to_excel(
        {
            "الإضافات": prepare_display_df(additions_df),
            "الإرجاعات": prepare_display_df(returns_df),
            "طلبات بدون فاتورة": prepare_display_df(orphan_salla_df),
            "فواتير بدون طلب": prepare_display_df(orphan_abc_df),
            "فواتير بعد آخر طلب": prepare_display_df(post_cutoff_df),
            "بانتظار الدفع": prepare_display_df(payment_pending_df),
            "الملغي_والمسترجع": prepare_display_df(cancelled_df if not cancelled_df.empty else review_df),
            "✅ تم الانتهاء": prepare_display_df(get_completed_items(pharmacy_name)),
        }
    )
    st.download_button(
        "📥 تصدير تبويبات الفرع إلى Excel",
        data=pharmacy_export,
        file_name=f"{pharmacy_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(
        ["الإضافات", "الإرجاعات", "طلبات بدون فاتورة", "فواتير بدون طلب", 
         "فواتير بعد آخر طلب", "بانتظار الدفع", "الملغي/المسترجع", "✅ تم الانتهاء"]
    )

    with tab1:
        render_case_cards(additions_df, allow_actions, pharmacist_name, pharmacy_name, is_admin=False)
    with tab2:
        render_case_cards(returns_df, allow_actions, pharmacist_name, pharmacy_name, is_admin=False)
    with tab3:
        render_case_cards(orphan_salla_df, allow_actions, pharmacist_name, pharmacy_name, is_admin=False)
    with tab4:
        render_case_cards(orphan_abc_df, allow_actions, pharmacist_name, pharmacy_name, is_admin=False)
    with tab5:
        render_case_cards(post_cutoff_df, False, pharmacist_name, pharmacy_name, is_admin=False)
    with tab6:
        render_case_cards(payment_pending_df, False, pharmacist_name, pharmacy_name, is_admin=False)
    with tab7:
        render_case_cards(cancelled_df if not cancelled_df.empty else review_df, False, pharmacist_name, pharmacy_name, is_admin=False)
    with tab8:
        completed_df = get_completed_items(pharmacy_name)
        render_completed_table(completed_df, is_admin=False)


# ========== INITIALIZATION ==========
ensure_database()


for key, default_value in {
    "logged_in": False,
    "username": "",
    "user_role": "",
    "pharmacist_name": "",
    "last_processed_preview": None,
    "view_session_id": None,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default_value


# ========== SIDEBAR LOGIN ==========
with st.sidebar:
    st.title("نظام بلسم العلا")
    st.caption("مطابقة طلبات سلة والفواتير")
    st.markdown("---")

    if not st.session_state.logged_in:
        username = st.text_input("اسم المستخدم")
        password = st.text_input("كلمة المرور", type="password")
        if st.button("دخول", use_container_width=True):
            user = fetch_user(username, password)
            if user:
                st.session_state.logged_in = True
                st.session_state.username = user[0]
                st.session_state.user_role = user[1]
                st.session_state.pharmacist_name = user[2] or ""
                st.rerun()
            else:
                st.error("بيانات الدخول غير صحيحة.")
    else:
        st.success(f"مرحباً {st.session_state.username}")
        
        if st.session_state.user_role == "pharmacy" and not st.session_state.pharmacist_name:
            st.markdown("---")
            pharmacist_input = st.text_input("👤 اسم الصيدلي", key="pharmacist_name_input")
            if st.button("💾 حفظ الاسم", use_container_width=True):
                if pharmacist_input.strip():
                    st.session_state.pharmacist_name = pharmacist_input.strip()
                    update_last_access(st.session_state.username, st.session_state.pharmacist_name)
                    st.success("✅ تم حفظ الاسم")
                    st.rerun()
        
        if st.session_state.user_role == "admin":
            st.markdown("---")
            st.markdown("### 📂 أدوات المدير")
            
            # تبويب تحديث الأرصدة في القائمة الجانبية
            if st.button("🔄 تحديث الأرصدة", use_container_width=True):
                st.session_state.page = "balances"
                st.rerun()
            
            # تبويب مراقبة تعديلات الصيدليات
            if st.button("👥 مراقبة التعديلات", use_container_width=True):
                st.session_state.page = "monitoring"
                st.rerun()
            
            # العودة إلى لوحة التحكم الرئيسية
            if st.button("📊 لوحة التحكم", use_container_width=True):
                st.session_state.page = "dashboard"
                st.rerun()
        
        st.markdown("---")
        if st.button("🚪 تسجيل خروج", use_container_width=True):
            for key in ["logged_in", "username", "user_role", "pharmacist_name", "last_processed_preview", "view_session_id", "page"]:
                st.session_state[key] = False if key == "logged_in" else None
            st.rerun()


# ========== MAIN CONTENT ==========
if not st.session_state.logged_in:
    st.markdown(
        """
        <div class="hero">
            <h1>نظام بلسم العلا لمراقبة إدخالات الفواتير</h1>
            <p>يعرض الإضافات والإرجاعات الفعلية، ويفصل السطور غير المربوطة وحالات اختلاف الفرع، ويحافظ على حالة كل فرع بين عمليات الرفع.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="note-card">
            <strong>ما الذي يجب فعله في هذا البرنامج؟</strong><br>
            1. المراجعة الدقيقة للاضافات والارجاعات.<br>
            2. تأكيد إضافة الطلبات التي لم يتم ادخالها على ABC.<br>
            3. تأكيد إرجاع الطلبات التي دخلت على ABC بالخطأ.
        </div>
        """,
        unsafe_allow_html=True,
    )
elif st.session_state.user_role == "pharmacy":
    if not st.session_state.pharmacist_name:
        st.info("👈 الرجاء إدخال اسم الصيدلي من القائمة الجانبية")
    else:
        render_pharmacy_dashboard()
else:  # admin
    # التحكم في الصفحة المختارة
    if not hasattr(st.session_state, 'page') or st.session_state.page == "dashboard":
        render_admin_dashboard()
    elif st.session_state.page == "balances":
        render_balances_updater()
    elif st.session_state.page == "monitoring":
        render_pharmacy_monitoring()
    else:
        render_admin_dashboard()


st.markdown("---")
st.markdown(
    """
    <div style="text-align:center;color:#607783;padding:0.6rem 0 0.8rem;">
        نظام بلسم العلا لمطابقة الطلبات والفواتير © 2026
    </div>
    """,
    unsafe_allow_html=True,
)
